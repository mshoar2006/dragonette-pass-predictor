"""
Dragonette pass predictor — core module.

Predicts Wyvern Dragonette (DRAG01-05) imaging opportunities over an AOI
polygon supplied as a KMZ. Method matches the validated 2026-07 workflow:

 1. AOI polygon read from KMZ; shoelace centroid; WGS84 geodetic -> ECEF.
 2. TLEs from Celestrak GP API by NORAD CATNR (cached, TTL).
 3. SGP4 propagation (TEME), coarse grid then golden-section refinement
    of the off-nadir minimum to 0.1 s -> TCA.
 4. TEME<->ECEF via GMST (IAU 1982) Z-rotation only.
 5. Off-nadir = angle at spacecraft between geocentric nadir and LOS to
    the AOI centroid; sign from orbit normal (LOS.(r x v)^). This natural
    right-of-track sense IS Wyvern's convention, so NO flip is applied
    (validated against their supplied sheet, July 2026 -- see METHOD.md).
 6. Sun elevation at AOI at TCA: Astronomical Almanac low-precision solar
    position (~0.01 deg), geometric, no refraction.
 7. Filters: standard |off-nadir| <= 20 deg and sun >= 20 deg;
    marginal 20-30 deg off-nadir or 15-20 deg sun.

No web-framework code in this file: it is imported by both cli.py and app.py.
"""

from __future__ import annotations

import html
import io
import json
import math
import os
import re
import threading
import time
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Callable

import numpy as np
from sgp4.api import Satrec, jday

# --------------------------------------------------------------------------
# Satellite catalogue (NORAD IDs per Celestrak SATCAT, as of 2026-07).
# Edit here when Wyvern launches new Dragonettes.
# --------------------------------------------------------------------------
SATELLITES: dict[str, int] = {
    "DRAG01": 56225,  # Dragonette-001 / EPICHyper-1, Transporter-7
    "DRAG02": 56995,  # Dragonette-002 / EPICHyper-2, Transporter-8
    "DRAG03": 58848,  # Dragonette-003 / EPICHyper-3, Transporter-9
    "DRAG04": 63254,  # Dragonette-004 / Loft YAM-8, Transporter-13
    "DRAG05": 66694,  # Dragonette-005 / Loft YAM-9, Transporter-15
}

# Operational status lives ONLY here so commissioning DRAG05 is a one-line flip.
# [SESSION 2026-07-14, the mission contact] DRAG05 (NORAD 66694) is NOT yet operational:
# predicted, but presented as non-taskable — badged, kept out of headline counts,
# and never mixed inline with operational satellites (CLAUDE.md hard constraint 3).
# [PLACEHOLDER] DRAG05 commissioning date — confirm with Wyvern, then set True.
OPERATIONAL: dict[str, bool] = {name: True for name in SATELLITES}
OPERATIONAL["DRAG05"] = False


# Dragonette sensor / quality parameters [LITERATURE: eoPortal EPICHyper /
# Dragonette; Wyvern data-product guide, 2026]. Used only for advisory
# acquisition-quality metrics (A1) — never as hard access filters (constraint 6).
GSD_NADIR_M = 5.3            # ground sample distance at nadir, metres
SWATH_KM = 20.0             # swath width at nadir
GLINT_HIGH_DEG = 20.0       # sun-glint angle below this = high specular risk (water)
GLINT_CAUTION_DEG = 40.0    # 20–40° = caution

# --------------------------------------------------------------------------
# Sensor profiles — "a second constellation is a config, not a fork" (SPEC.md)
# --------------------------------------------------------------------------
# A profile answers a genuinely different question per sensor class:
#
#   Dragonette is **agile**: it rolls to your AOI on request. The question is
#   "can I task it?", and the envelope (|off-nadir| <= 20 deg) is how far it is
#   willing to lean.
#
#   Landsat and Sentinel-2 are **not agile** — fixed nadir push-brooms that image
#   whatever falls inside their field of view, systematically, whether you ask or
#   not. The question is "when will it image my AOI anyway?", and the envelope is
#   the FOV half-angle: a hard optical limit, not a preference. There is no
#   marginal band, because there is no tasking negotiation to stretch.
#
# That is why min_sun defaults to 0 for them: they acquire the daylit side on
# their own schedule, so a low-sun pass is a real acquisition. Sun elevation is
# still reported per pass so the user can judge usability themselves.
# [SESSION 2026-07-15]
@dataclass(frozen=True)
class SensorProfile:
    key: str
    display: str
    satellites: dict[str, int]
    operational: dict[str, bool]
    max_off_nadir_deg: float
    marginal_off_nadir_deg: float
    min_sun_elev_deg: float
    marginal_sun_elev_deg: float
    swath_km: float
    gsd_m: float
    agile: bool
    note: str = ""


def fov_half_angle_deg(swath_km: float, alt_km: float, re_km: float = 6371.0) -> float:
    """Look half-angle subtending a given GROUND swath, spherical Earth.

    Derived rather than asserted, because NASA publishes Landsat's swath but not
    its FOV in degrees. Validated three ways for Sentinel-2 [VERIFIED 2026-07-15]:
    this returns 10.43 deg from ESA's stated 290 km / 786 km; ESA separately states
    a 20.6 deg FOV (10.30 deg half); and the max off-nadir measured across 400 real
    S2 acquisitions was 10.46 deg. For Landsat it returns 7.47 deg from NASA's
    185 km / 705 km, against the conventional 15 deg FOV (7.50 deg half).
    """
    lam = (swath_km / 2.0) / re_km                       # Earth-central half-angle
    return math.degrees(math.atan2(re_km * math.sin(lam),
                                   re_km + alt_km - re_km * math.cos(lam)))


def ground_half_swath_km(alt_km: float, fov_half_deg: float,
                         re_km: float = 6371.0) -> float | None:
    """Ground half-swath subtended by a FOV half-angle at `alt_km`, or None if the
    look angle passes the limb. Exact inverse of `fov_half_angle_deg`:
    `lambda(eta) = asin((1 + h/Re)·sin eta) − eta`, ground distance `Re·lambda`.

    Why a fixed-FOV push-broom's ground swath is NOT a constant [VERIFIED
    2026-07-16]: it scales with the spacecraft's *geodetic* altitude, which swings
    **704.6 → 731.7 km** for Landsat-9 over a single orbit — the orbit is circular
    geocentrically, but the ellipsoid it flies over is not. So the true half-swath
    runs 92.48 → 96.04 km against the nominal 92.5.

    This matters because a profile describes one optic two ways — `swath_km` (a
    ground distance) and `max_off_nadir_deg` (an angle) — and they cannot both be
    constant. Holding `swath_km` fixed let the access gate and the footprint
    disagree by up to 3.5 km, which reported an AOI that was wholly inside the FOV
    as 0% covered. Deriving the swath from the gate makes them agree by
    construction. [LITERATURE — B1 lambda(eta) relation.]"""
    e = math.radians(fov_half_deg)
    s = (1.0 + alt_km / re_km) * math.sin(e)
    if s >= 1.0:                     # look direction misses the Earth entirely
        return None
    return re_km * (math.asin(s) - e)


DRAGONETTE = SensorProfile(
    key="dragonette", display="Wyvern Dragonette (DRAG01-05)",
    satellites=SATELLITES, operational=OPERATIONAL,
    max_off_nadir_deg=20.0, marginal_off_nadir_deg=30.0,
    min_sun_elev_deg=20.0, marginal_sun_elev_deg=15.0,
    swath_km=SWATH_KM, gsd_m=GSD_NADIR_M, agile=True,
    note="Agile/taskable: rolls to the AOI. Envelope |off-nadir| <= 20 deg, "
         "sun >= 20 deg [REPORT — Wyvern sheet max 19.9 deg]. Marginal band "
         "20-30 deg / 15-20 deg is a stretch tier, not a Wyvern commitment.")

# NORAD IDs [VERIFIED 2026-07-15 vs the live Celestrak `resource` group].
# Specs [VERIFIED 2026-07-15]: swath 185 km, 30 m multispectral / 15 m pan,
# 705 km altitude, 16-day repeat, LTDN 10:12 +/- 5 min, inclination 98.2 deg
# — NASA science.nasa.gov/mission/landsat/oli and USGS usgs.gov/landsat-missions/landsat-9.
# Both Landsat-8 and Landsat-9 are OPERATIONAL. [VERIFIED 2026-07-20 vs the live
# earth-search STAC (landsat-c2-l2): Landsat-8 has 14,632 L2 scenes in the trailing
# 30 days, latest 2026-07-10, ~3.0M lifetime.] An earlier build flagged L8
# non-operational off "no scene since 2026-06-30", which was Collection-2 L2
# processing latency, NOT an outage — corrected here.
LANDSAT = SensorProfile(
    key="landsat", display="Landsat 8/9 (OLI)",
    satellites={"LANDSAT8": 39084, "LANDSAT9": 49260},
    operational={"LANDSAT8": True, "LANDSAT9": True},
    max_off_nadir_deg=round(fov_half_angle_deg(185.0, 705.0), 2),
    marginal_off_nadir_deg=round(fov_half_angle_deg(185.0, 705.0), 2),
    min_sun_elev_deg=0.0, marginal_sun_elev_deg=0.0,
    swath_km=185.0, gsd_m=30.0, agile=False,
    note="Fixed nadir push-broom — images whatever falls in its FOV on its own "
         "16-day cycle; not taskable. Envelope is the FOV half-angle derived from "
         "the 185 km swath at 705 km. Both LANDSAT8 and LANDSAT9 operational "
         "[VERIFIED 2026-07-20 vs earth-search STAC].")

# Specs [VERIFIED 2026-07-15]: swath 290 km, 20.6 deg FOV, 10/20/60 m GSD,
# 786 km mean altitude, LTDN 10:30 MLST, 5-day constellation revisit (10-day per
# satellite) — ESA Copernicus SentiWiki sentiwiki.copernicus.eu/web/s2-mission.
# Sentinel-2A is still acquiring [VERIFIED 2026-07-15 — scenes 05-10, 05-30].
SENTINEL2 = SensorProfile(
    key="sentinel2", display="Sentinel-2 A/B/C (MSI)",
    satellites={"SENTINEL2A": 40697, "SENTINEL2B": 42063, "SENTINEL2C": 60989},
    operational={"SENTINEL2A": True, "SENTINEL2B": True, "SENTINEL2C": True},
    max_off_nadir_deg=10.3, marginal_off_nadir_deg=10.3,   # ESA's stated 20.6 deg FOV
    min_sun_elev_deg=0.0, marginal_sun_elev_deg=0.0,
    swath_km=290.0, gsd_m=10.0, agile=False,
    note="Fixed nadir push-broom — images systematically; not taskable. Envelope "
         "is ESA's stated 20.6 deg FOV (10.3 deg half-angle); the geometric "
         "derivation from 290 km @ 786 km gives 10.43 deg and the max measured "
         "across 400 real acquisitions was 10.46 deg.")

PROFILES: dict[str, SensorProfile] = {p.key: p for p in (DRAGONETTE, LANDSAT, SENTINEL2)}


COMBINED_KEY = "all"     # the "show every constellation at once" pseudo-sensor
COMBINED_DISPLAY = "All sensors (Dragonette + Landsat + Sentinel-2)"


def get_profile(key: str | None) -> SensorProfile:
    """Look up a sensor profile by key; None/empty => Dragonette (the default)."""
    if not key:
        return DRAGONETTE
    try:
        return PROFILES[key.strip().lower()]
    except KeyError:
        raise ValueError(
            f"unknown sensor {key!r}; choose one of {', '.join(sorted(PROFILES))}") from None


def combined_roster() -> tuple[dict[str, int], dict[str, bool]]:
    """Union of every profile's satellites + operational map, ordered Dragonette,
    Landsat, Sentinel-2 — the lane roster for the combined 'all sensors' view."""
    sats: dict[str, int] = {}
    op: dict[str, bool] = {}
    for p in (DRAGONETTE, LANDSAT, SENTINEL2):
        sats.update(p.satellites)
        op.update(p.operational)
    return sats, op

CELESTRAK_URL = "https://celestrak.org/NORAD/elements/gp.php?CATNR={catnr}&FORMAT=TLE"
DEFAULT_CACHE = Path.home() / ".cache" / "dragonette_tles.json"
# Celestrak asks clients to identify themselves and blocks generic scripted
# agents. [SESSION 2026-07-15 — mirrors the UA fetch_real_data.py already sends.]
CELESTRAK_UA = {"User-Agent": "dragonette-predictor/2.0 (research)"}
# After a failed fetch, serve the stale cache rather than re-hitting Celestrak,
# for this long. Celestrak refreshes ~2-hourly and asks for <=2-3 polls/file/day;
# an un-cooled retry loop across concurrent requests is how a soft throttle
# becomes an IP ban. [SESSION 2026-07-15]
FETCH_RETRY_COOLDOWN_S = 15 * 60.0
# Serialises the fetch+cache-write so N concurrent requests cause 1 fetch, not N.
_FETCH_LOCK = threading.Lock()
# A rise in semi-major axis faster than this ⇒ thrust, not drag. Drag can only
# *lower* a, so any sustained rise is unambiguous; the threshold only has to clear
# fit noise. [VERIFIED 2026-07-15 over ~1 d of real Celestrak elements: DRAG01/02/
# 03/05 decayed 6-18 m/day, DRAG04 rose 100 m/day. 30 m/day sits clear of both.]
MANOEUVRE_DA_RISE_KM_PER_DAY = 0.03

# Off-nadir sign convention. [REPORT — corrected 2026-07-14 against Wyvern's
# actual sheet "Wyvern Simulated Passes … June 24–July 24 2026", Wyvern]
# Wyvern's sign equals the NATURAL right-of-track sense (LOS·(r×v)̂), so NO flip.
# Verified on 5 robust Site A passes (off-nadir ≥6°, magnitude within ~1°,
# their "End Datetime" 18–100 s after our TCA): signs match column-for-column.
# The prior default True inverted every sign vs Wyvern — see VALIDATION.md
# "Sign resolution 2026-07-14". Do not change without re-validating against a
# fresh Wyvern signed sheet.
SIGN_FLIP_TO_MATCH_WYVERN = False

# WGS84
_A = 6378.137          # km
_E2 = 6.69437999014e-3
_B = _A * math.sqrt(1.0 - _E2)     # semi-minor axis, km

DEG = math.pi / 180.0



# --------------------------------------------------------------------------
# KMZ / KML parsing
# --------------------------------------------------------------------------
@dataclass
class AOI:
    name: str
    vertices_lonlat: list[tuple[float, float]]   # (lon, lat) degrees
    centroid_lon: float
    centroid_lat: float
    terrain_alt_m: float

    @property
    def centroid_ecef(self) -> np.ndarray:
        return geodetic_to_ecef(self.centroid_lat, self.centroid_lon,
                                self.terrain_alt_m / 1000.0)

    @property
    def vertices_ecef(self) -> np.ndarray:
        return np.array([
            geodetic_to_ecef(lat, lon, self.terrain_alt_m / 1000.0)
            for lon, lat in self.vertices_lonlat
        ])


class AmbiguousPolygonError(ValueError):
    """Raised when a KMZ holds >1 polygon and the caller did not disambiguate.

    [SESSION 2026-07-14, the mission contact] SiteA.kmz carries AOI 1, AOI 2 and
    SITEA_100sqkm; silently taking the first is a known footgun. Callers
    must name a polygon or opt into --all-polygons; carries `.names` so the
    CLI/API can echo the choices.
    """
    def __init__(self, names: list[str]):
        self.names = names
        super().__init__(
            "KMZ contains multiple polygons; specify one with --polygon / "
            "polygon= (substring match), or use --all-polygons to predict each. "
            "Polygons: " + ", ".join(names))


# A KMZ is an emailed file from a partner — untrusted, and `/predict` is
# unauthenticated. Without a bound, `zf.read()` decompresses whatever the zip
# claims: a ~50 KB archive whose inner doc.kml inflates to gigabytes (trivial on
# repetitive XML, ratios ~1000:1) walks the worker straight into an OOM on a
# single request. Bound the declared size, the expansion ratio, and the member
# count. [SESSION 2026-07-15]
MAX_UPLOAD_BYTES = 10 * 1024 * 1024        # a real 100 km2 KMZ is a few KB
MAX_KML_BYTES = 50 * 1024 * 1024           # generous for a hand-drawn mega-polygon
MAX_COMPRESSION_RATIO = 200.0
MAX_ZIP_MEMBERS = 64


def _kml_text(data: bytes) -> str:
    """Return the KML text from a KMZ (zip) or a bare KML byte string.

    Raises ValueError on anything that looks like a decompression bomb rather
    than expanding it. Never calls extract()/extractall(), so there is no
    zip-slip surface — member names only ever select, never become paths.
    """
    if len(data) > MAX_UPLOAD_BYTES:
        raise ValueError(
            f"KMZ/KML is {len(data) / 1e6:.1f} MB; limit is "
            f"{MAX_UPLOAD_BYTES / 1e6:.0f} MB. A real AOI polygon is a few KB.")
    if data[:2] == b"PK":  # zip => KMZ
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            infos = zf.infolist()
            if len(infos) > MAX_ZIP_MEMBERS:
                raise ValueError(
                    f"KMZ contains {len(infos)} members; limit is {MAX_ZIP_MEMBERS}.")
            kml_infos = [i for i in infos if i.filename.lower().endswith(".kml")]
            if not kml_infos:
                raise ValueError("KMZ contains no .kml file")
            info = kml_infos[0]
            if info.file_size > MAX_KML_BYTES:
                raise ValueError(
                    f"KMZ inner {info.filename!r} declares {info.file_size / 1e6:.1f} MB; "
                    f"limit is {MAX_KML_BYTES / 1e6:.0f} MB (decompression bomb?)")
            if (info.compress_size > 0
                    and info.file_size / info.compress_size > MAX_COMPRESSION_RATIO):
                raise ValueError(
                    f"KMZ inner {info.filename!r} expands "
                    f"{info.file_size / info.compress_size:.0f}x (limit "
                    f"{MAX_COMPRESSION_RATIO:.0f}x) — refusing to decompress it.")
            return zf.read(info).decode("utf-8", errors="replace")
    return data.decode("utf-8", errors="replace")


_NAME_MAX_LEN = 120
_CDATA_RE = re.compile(r"^\s*<!\[CDATA\[(.*?)\]\]>\s*$", re.S)


def _clean_placemark_name(raw: str) -> str:
    """Normalise an untrusted KML <name> into a safe display/identifier string.

    A KMZ is an emailed file, so the name is attacker-controlled, and it reaches
    two rendering sinks: the SPA map popup and the xlsx. Google Earth wraps names
    in CDATA as a matter of course, so `<![CDATA[<img src=x onerror=...>]]>`
    survives a bare `<name>(.*?)</name>` scrape and executes when interpolated
    into innerHTML. Strip the CDATA wrapper, resolve entities, then remove any
    markup that survives — a KML placemark name has no legitimate use for tags.

    Note this deliberately does NOT neutralise a leading =/+/-/@ (Excel formula
    injection): that is context-specific and is handled at the xlsx sink by
    `_xlsx_safe`, so a legitimately named polygon keeps its identity here and
    stays matchable by `polygon_name`. [SESSION 2026-07-15]
    """
    s = (raw or "").strip()
    m = _CDATA_RE.match(s)
    if m:
        s = m.group(1)
    s = html.unescape(s)                       # &lt;img ...&gt; -> <img ...>
    s = re.sub(r"<[^>]*>", "", s)              # drop any markup
    s = s.replace("<", "").replace(">", "")    # and any unbalanced angle bracket
    s = "".join(ch for ch in s if ch.isprintable())
    s = " ".join(s.split())                    # collapse whitespace/newlines
    return s[:_NAME_MAX_LEN] or "AOI"


def _polygon_placemarks(kml: str) -> list[tuple[str, str]]:
    """Return (name, block) for every Placemark that contains a <Polygon>.

    Names are passed through `_clean_placemark_name` — they are untrusted input.

    Falls back to treating the whole document as one bare polygon block if it
    has a <Polygon> but no Placemark wrapper (some exporters do this)."""
    placemarks = re.findall(r"<Placemark\b.*?</Placemark>", kml, re.S)
    out: list[tuple[str, str]] = []
    for pm in placemarks:
        if "<Polygon" not in pm:
            continue
        m = re.search(r"<name>(.*?)</name>", pm, re.S)
        out.append((_clean_placemark_name(m.group(1)) if m else "AOI", pm))
    if not out and "<Polygon" in kml:  # bare polygon document
        m = re.search(r"<name>(.*?)</name>", kml, re.S)
        out.append((_clean_placemark_name(m.group(1)) if m else "AOI", kml))
    return out


def list_polygons(data: bytes) -> list[str]:
    """Names of all polygon-bearing placemarks in a KMZ/KML, document order."""
    return [name for name, _ in _polygon_placemarks(_kml_text(data))]


def parse_kmz(data: bytes, terrain_alt_m: float = 0.0,
              polygon_name: str | None = None) -> AOI:
    """Extract the named (or sole) Polygon outer boundary from a KMZ or KML.

    Raises AmbiguousPolygonError if the file has >1 polygon and no
    `polygon_name` is given — never silently takes the first. [SESSION]"""
    kml = _kml_text(data)
    polys = _polygon_placemarks(kml)
    if not polys:
        raise ValueError("No <Polygon> found in KMZ/KML")

    if polygon_name:
        matches = [(n, b) for n, b in polys if polygon_name.lower() in n.lower()]
        if not matches:
            raise ValueError(
                f"No polygon named like '{polygon_name}' found. "
                f"Polygons: {', '.join(n for n, _ in polys)}")
        chosen_name, chosen_block = matches[0]
    elif len(polys) > 1:
        raise AmbiguousPolygonError([n for n, _ in polys])
    else:
        chosen_name, chosen_block = polys[0]

    outer = re.search(
        r"<outerBoundaryIs>.*?<coordinates>(.*?)</coordinates>",
        chosen_block, re.S)
    if not outer:  # some writers omit outerBoundaryIs
        outer = re.search(r"<coordinates>(.*?)</coordinates>", chosen_block, re.S)
    if not outer:
        raise ValueError("Polygon has no <coordinates>")

    verts: list[tuple[float, float]] = []
    for token in outer.group(1).split():
        parts = token.split(",")
        if len(parts) >= 2:
            verts.append((float(parts[0]), float(parts[1])))
    if len(verts) >= 2 and verts[0] == verts[-1]:
        verts = verts[:-1]
    if len(verts) < 3:
        raise ValueError("Polygon has fewer than 3 vertices")

    clon, clat = _shoelace_centroid(verts)
    return AOI(chosen_name, verts, clon, clat, terrain_alt_m)


def _wrap180(lon: float) -> float:
    """Wrap a longitude to (-180, 180]."""
    return (lon + 180.0) % 360.0 - 180.0


def _shoelace_centroid(verts: list[tuple[float, float]]) -> tuple[float, float]:
    """Planar shoelace centroid of a (lon, lat) ring. Returns (lon, lat).

    Two corrections over the naive form [SESSION 2026-07-15]:

    1. **Longitudes are unwrapped onto a branch centred on the first vertex**
       before summing, and re-wrapped after. Raw lon/lat made a ring straddling
       the antimeridian average to the *opposite side of the planet*: a
       Fiji-scale AOI spanning 179.95°E → 179.95°W returned lon ≈ 0.0 — an error
       of ~19,000 km, silently, with every downstream pass, sun angle and cloud
       lookup then computed for a point in the Atlantic.

    2. **Coordinates are shifted to a local origin** before the sum. The cross
       terms are ~lon·lat (≈ 4,100 at these sites) while the differences carrying
       the signal are ~1e-4, so the subtraction destroyed ~8 significant digits.
       Measured before the fix, at Site A: a 10 km AOI was fine (0.0 m), but a
       197 m AOI erred 0.6 m, a **20 m AOI erred 53 m — larger than the AOI
       itself, and 10× the 5.3 m GSD** — and a 2 m AOI erred 9.3 km. The old
       degeneracy guard (`|a| < 1e-12` in deg²) only rescued polygons *below* the
       danger zone, not within it. Shifting restores full precision at every
       scale.

    The degeneracy test is now relative to the ring's own extent, so it means
    "collinear/zero-area" rather than "smaller than an arbitrary absolute size".
    """
    n = len(verts)
    if n == 0:
        raise ValueError("Polygon has no vertices")

    lon0 = verts[0][0]
    xs = [lon0 + _wrap180(x - lon0) for x, _ in verts]      # unwrap onto one branch
    ys = [y for _, y in verts]
    ox, oy = xs[0], ys[0]                                   # shift to a local origin
    xs = [x - ox for x in xs]
    ys = [y - oy for y in ys]

    a = cx = cy = 0.0
    for i in range(n):
        x0, y0 = xs[i], ys[i]
        x1, y1 = xs[(i + 1) % n], ys[(i + 1) % n]
        cross = x0 * y1 - x1 * y0
        a += cross
        cx += (x0 + x1) * cross
        cy += (y0 + y1) * cross

    scale = max(max(xs) - min(xs), max(ys) - min(ys))
    if scale <= 0.0 or abs(a) < 1e-12 * scale * scale:      # collinear / zero-area
        return _wrap180(sum(xs) / n + ox), sum(ys) / n + oy
    a *= 0.5
    return _wrap180(cx / (6 * a) + ox), cy / (6 * a) + oy


# --------------------------------------------------------------------------
# Geodesy / time / frames
# --------------------------------------------------------------------------
def geodetic_to_ecef(lat_deg: float, lon_deg: float, alt_km: float) -> np.ndarray:
    lat, lon = lat_deg * DEG, lon_deg * DEG
    n = _A / math.sqrt(1.0 - _E2 * math.sin(lat) ** 2)
    return np.array([
        (n + alt_km) * math.cos(lat) * math.cos(lon),
        (n + alt_km) * math.cos(lat) * math.sin(lon),
        (n * (1.0 - _E2) + alt_km) * math.sin(lat),
    ])


def gmst_rad(jd_ut1: np.ndarray | float) -> np.ndarray | float:
    """IAU 1982 GMST, radians."""
    t = (np.asarray(jd_ut1, dtype=float) - 2451545.0) / 36525.0
    g = (280.46061837
         + 360.98564736629 * (np.asarray(jd_ut1, dtype=float) - 2451545.0)
         + 0.000387933 * t * t - t ** 3 / 38710000.0)
    return np.remainder(g, 360.0) * DEG


def ecef_to_teme(r_ecef: np.ndarray, theta: np.ndarray) -> np.ndarray:
    """Rotate ECEF vector(s) into TEME given GMST theta (rad).
    theta may be scalar or (N,); r_ecef (3,) or (N,3)."""
    c, s = np.cos(theta), np.sin(theta)
    x, y, z = r_ecef[..., 0], r_ecef[..., 1], r_ecef[..., 2]
    xt = c * x - s * y
    yt = s * x + c * y
    zt = np.broadcast_to(z, np.shape(xt))
    return np.stack([xt, yt, zt], axis=-1)


def teme_to_ecef(r_teme: np.ndarray, theta: float) -> np.ndarray:
    c, s = math.cos(theta), math.sin(theta)
    return np.array([c * r_teme[0] + s * r_teme[1],
                     -s * r_teme[0] + c * r_teme[1],
                     r_teme[2]])


def geodetic_up(lat_deg: float, lon_deg: float) -> np.ndarray:
    lat, lon = lat_deg * DEG, lon_deg * DEG
    return np.array([math.cos(lat) * math.cos(lon),
                     math.cos(lat) * math.sin(lon),
                     math.sin(lat)])


def ecef_to_geodetic_latlon(r_ecef: np.ndarray) -> tuple[float, float]:
    """WGS84 geodetic latitude & longitude (deg) from an ECEF vector (km).
    Bowring's closed-form; ample accuracy (<1e-8 deg) for nadir geometry."""
    x, y, z = float(r_ecef[0]), float(r_ecef[1]), float(r_ecef[2])
    lon = math.atan2(y, x)
    p = math.hypot(x, y)
    b = _A * math.sqrt(1.0 - _E2)                # semi-minor axis
    ep2 = _E2 / (1.0 - _E2)                       # second eccentricity^2
    th = math.atan2(z * _A, p * b)
    lat = math.atan2(z + ep2 * b * math.sin(th) ** 3,
                     p - _E2 * _A * math.cos(th) ** 3)
    return math.degrees(lat), math.degrees(lon)


def nadir_unit_teme(r_teme: np.ndarray, theta: float,
                    ellipsoid: bool = False) -> np.ndarray:
    """Unit nadir direction (pointing 'down') at the spacecraft, in TEME.

    ellipsoid=False: geocentric nadir (-r̂) — the validated default.
    ellipsoid=True:  WGS84 geodetic nadir (negative ellipsoid normal at the
    sub-satellite point). The two differ by up to ~0.2° at mid-latitudes; the
    geodetic form is the physically correct 'nadir' for an off-nadir/roll angle
    but shifts values off the geocentric baseline, so it is opt-in. [SESSION]"""
    if not ellipsoid:
        return -r_teme / np.linalg.norm(r_teme)
    r_ecef = teme_to_ecef(r_teme, theta)
    lat, lon = ecef_to_geodetic_latlon(r_ecef)
    up_teme = ecef_to_teme(geodetic_up(lat, lon), theta)
    return -up_teme / np.linalg.norm(up_teme)


def ellipsoid_intersect(p_ecef: np.ndarray, d_hat: np.ndarray) -> np.ndarray | None:
    """First WGS84-ellipsoid intersection of the ray p + t·d (ECEF km), or None
    if it misses Earth. Analytic (scale to unit sphere). [LITERATURE — B1]"""
    k = np.array([1.0 / _A, 1.0 / _A, 1.0 / _B])
    pk, dk = p_ecef * k, d_hat * k
    a = float(dk @ dk); b = 2.0 * float(pk @ dk); c = float(pk @ pk) - 1.0
    disc = b * b - 4 * a * c
    if disc < 0 or a == 0:
        return None
    t = (-b - math.sqrt(disc)) / (2 * a)             # near root = first hit
    if t < 0:                                        # intersection is behind us
        return None
    return p_ecef + t * d_hat


_OMEGA_E = 7.292115e-5     # Earth rotation rate, rad/s


def swath_footprint_lonlat(r_t: np.ndarray, v_t: np.ndarray, theta: float,
                           aoi: "AOI", swath_km: float = SWATH_KM,
                           agile: bool = True,
                           fov_half_deg: float | None = None) -> list[tuple[float, float]]:
    """Ground footprint polygon (lon,lat): a *ground* rectangle oriented along the
    satellite ground track, spanning the AOI along-track (min 6 km) (B1). Its
    cross-track width is `swath_km` (a fixed ground distance, the standard EO
    convention) unless `fov_half_deg` is given, in which case it is derived from
    the sensor's FOV at its instantaneous altitude — see below.

    Where the swath sits cross-track depends on whether the sensor can roll. This
    is the honest footprint model that IMPROVEMENTS.md's sensor-profile section
    predicted adding the push-brooms would force. [SESSION 2026-07-16]

    - `agile=True` (Dragonette): the sensor rolls to the AOI on request, so the
      boresight points at the AOI centroid and the swath is centred there. What
      coverage then measures is "does my AOI fit cross-track" — which is why it
      reads 1.0 for every AOI narrower than 20 km. Mislabelled, not wrong; see
      SPEC.md Non-goals.
    - `agile=False` (Landsat/Sentinel-2): a fixed nadir push-broom images
      whatever falls under it, so the swath is centred on the **ground track**,
      not on the AOI. Centring it on the AOI would report full coverage for an
      AOI sitting at the very edge of the swath — a plausible number that cannot
      be false, which is this project's dominant bug class. The along-track
      station stays at the AOI (a push-broom's strip is continuous along-track,
      so along-track is never the limiting dimension); only the cross-track
      centre moves onto the track.

    `fov_half_deg` (non-agile only) derives the half-swath from the sensor's FOV
    at its *instantaneous* altitude, instead of trusting the nominal `swath_km`.
    Pass it whenever the caller also gates access on that angle, so the gate and
    the footprint cannot contradict each other — see `ground_half_swath_km`.
    """
    c_lat, c_lon = aoi.centroid_lat, aoi.centroid_lon
    center = geodetic_to_ecef(c_lat, c_lon, aoi.terrain_alt_m / 1000.0)
    sat = teme_to_ecef(r_t, theta)
    v_ecef = teme_to_ecef(v_t, theta) - np.cross([0.0, 0.0, _OMEGA_E], sat)
    up = geodetic_up(c_lat, c_lon)
    track = v_ecef - (v_ecef @ up) * up              # ground-track heading at AOI
    n = np.linalg.norm(track)
    if n < 1e-9:
        return []
    track /= n
    cross = np.cross(up, track); cross /= np.linalg.norm(cross)   # cross-track

    verts = aoi.vertices_ecef
    extent = float(np.linalg.norm(verts.max(axis=0) - verts.min(axis=0))) if len(verts) else 6.0
    half_at = max(extent, 6.0) / 2.0                 # along-track half-length, km
    half_ct = swath_km / 2.0                          # ground cross-track half, km

    if not agile:
        # Slide the centre cross-track onto the ground track, keeping the AOI's
        # along-track station. The track is the locus of *geodetic* sub-satellite
        # points — the standard definition, and the one that matches a
        # local-vertical-pointing push-broom's boresight. Measured against the
        # geocentric sub-point (Landsat-9, one orbit): they differ by up to
        # 2.1 km, but ~2.07 km of that is ALONG-track, which a continuous strip
        # does not care about; the cross-track component that would actually move
        # the swath edge peaks at 0.62 km. [SESSION 2026-07-16]
        ssp_lat, ssp_lon = ecef_to_geodetic_latlon(sat)
        ssp = geodetic_to_ecef(ssp_lat, ssp_lon, 0.0)
        center = center - float((center - ssp) @ cross) * cross
        if fov_half_deg is not None:
            # |sat − ssp| IS the geodetic altitude: ssp is the foot of the
            # ellipsoid normal through the spacecraft.
            hs = ground_half_swath_km(float(np.linalg.norm(sat - ssp)), fov_half_deg)
            if hs is None:
                return []
            half_ct = hs

    corners = []
    for sa, sc in [(-1, -1), (-1, 1), (1, 1), (1, -1)]:
        pt = center + sa * half_at * track + sc * half_ct * cross
        lat, lon = ecef_to_geodetic_latlon(pt)
        corners.append((lon, lat))
    return corners


def aoi_coverage_fraction(footprint_lonlat: list[tuple[float, float]],
                          aoi: "AOI") -> float | None:
    """Fraction of the AOI polygon covered by the swath footprint (0–1), area-
    correct via a local azimuthal-equidistant projection. None if unavailable."""
    if not footprint_lonlat or len(aoi.vertices_lonlat) < 3:
        return None
    try:
        from shapely.geometry import Polygon
        from shapely.ops import transform
        import pyproj
        proj = pyproj.Transformer.from_crs(
            "EPSG:4326",
            f"+proj=aeqd +lat_0={aoi.centroid_lat} +lon_0={aoi.centroid_lon} +units=m",
            always_xy=True).transform
        aoi_poly = transform(proj, Polygon(aoi.vertices_lonlat))
        fp_poly = transform(proj, Polygon(footprint_lonlat))
        if aoi_poly.area <= 0:
            return None
        return round(min(1.0, aoi_poly.intersection(fp_poly).area / aoi_poly.area), 3)
    except Exception:
        return None


def _kasten_young_airmass(zenith_deg: float) -> float:
    """Relative optical air mass (Kasten–Young 1989); robust near the horizon."""
    z = min(zenith_deg, 90.0)
    denom = math.cos(z * DEG) + 0.50572 * (96.07995 - z) ** -1.6364
    return 1.0 / denom if denom > 0 else 40.0


def acquisition_geometry(r_t: np.ndarray, v_t: np.ndarray, theta: float,
                         site_ecef: np.ndarray, aoi_lat: float, aoi_lon: float,
                         off_nadir_deg: float, sun_el: float, sun_az: float,
                         gsd_nadir_m: float = GSD_NADIR_M) -> dict:
    """Advisory acquisition-quality geometry for one pass (A1). All angles in
    degrees; effective GSD in metres. Pure geometry, reuses the state predict()
    already has. [LITERATURE — standard EO/astro relations.]"""
    up = geodetic_up(aoi_lat, aoi_lon)                       # ECEF local up
    sat_ecef = teme_to_ecef(r_t, theta)
    to_sat = sat_ecef - site_ecef
    d = float(np.linalg.norm(to_sat))
    # view (incidence) zenith at the target = 90 − satellite elevation
    sat_el = 90.0 - math.degrees(math.acos(
        max(-1.0, min(1.0, float(np.dot(to_sat, up)) / d))))
    view_zenith = 90.0 - sat_el
    # view azimuth (target→satellite), N clockwise, in the local ENU frame
    east = np.array([-math.sin(aoi_lon * DEG), math.cos(aoi_lon * DEG), 0.0])
    north = np.cross(up, east)
    view_az = math.degrees(math.atan2(float(np.dot(to_sat, east)),
                                      float(np.dot(to_sat, north)))) % 360.0
    thv, ths = view_zenith * DEG, (90.0 - sun_el) * DEG
    dphi = (view_az - sun_az) * DEG
    # phase (Sun–target–sensor) and sun-glint (LOS vs specular) angles
    cos_phase = math.cos(ths) * math.cos(thv) + math.sin(ths) * math.sin(thv) * math.cos(dphi)
    phase = math.degrees(math.acos(max(-1.0, min(1.0, cos_phase))))
    cos_glint = math.cos(thv) * math.cos(ths) - math.sin(thv) * math.sin(ths) * math.cos(dphi)
    glint = math.degrees(math.acos(max(-1.0, min(1.0, cos_glint))))
    # Effective ground sample distance (secant law). `gsd_nadir_m` comes from the
    # sensor profile, so Landsat reports 30 m and Sentinel-2 10 m rather than
    # Dragonette's 5.3 m. [SESSION 2026-07-15]
    # (The old `if sun_el > -90` guard here was dead — elevation is always > -90 —
    # and the parameter `alt_km` was accepted and never read. Both removed.)
    eta = abs(off_nadir_deg) * DEG
    gsd_eff = gsd_nadir_m / (math.cos(eta) * math.cos(thv))
    # ascending/descending from the sign of the northward velocity (TEME and
    # ECEF share the Z axis, so v_t[2] > 0 ⇒ moving north ⇒ ascending)
    node = "ascending" if v_t[2] >= 0 else "descending"
    # ground-track heading: Earth-relative velocity projected onto local ENU
    v_ecef = teme_to_ecef(v_t, theta) - np.cross([0, 0, 7.292115e-5], sat_ecef)
    sub_up = sat_ecef / np.linalg.norm(sat_ecef)
    slon = math.atan2(sat_ecef[1], sat_ecef[0])
    e2 = np.array([-math.sin(slon), math.cos(slon), 0.0])
    n2 = np.cross(sub_up, e2)
    heading = math.degrees(math.atan2(float(np.dot(v_ecef, e2)),
                                      float(np.dot(v_ecef, n2)))) % 360.0
    return {
        "view_zenith_deg": round(view_zenith, 1),
        "phase_deg": round(phase, 1),
        "sun_glint_deg": round(glint, 1),
        "glint_risk": ("high" if glint < GLINT_HIGH_DEG
                       else "caution" if glint < GLINT_CAUTION_DEG else "low"),
        "effective_gsd_m": round(gsd_eff, 2),
        "sun_azimuth_deg": round(sun_az, 1),
        "view_azimuth_deg": round(view_az, 1),
        "airmass_view": round(_kasten_young_airmass(view_zenith), 2),
        "airmass_sun": round(_kasten_young_airmass(90.0 - sun_el), 2),
        "ground_track_heading_deg": round(heading, 1),
        "node": node,
    }


def _quality_badge(off_nadir_mag: float, sun_el: float, glint: float) -> str:
    """Collapse acquisition geometry into good/marginal/poor (A2, advisory).
    Glint only downgrades if the AOI is water/wet — unknown here, so it warns
    but never forces 'poor' on its own. [SESSION — thresholds to confirm.]"""
    if sun_el < 15.0:
        return "poor"                       # too dark for optical
    good = off_nadir_mag <= 12.0 and sun_el >= 20.0
    marg = off_nadir_mag <= 20.0 and sun_el >= 15.0
    return "good" if good else "marginal" if marg else "poor"


# Upper bound on a requested window. Shared by cli.py and app.py so the two
# surfaces cannot disagree about what is accepted. [SESSION 2026-07-15]
MAX_WINDOW_DAYS = 31.0


def parse_start_utc(s: str | None) -> datetime | None:
    """Parse an ISO-8601 window start to UTC. None/empty => None (i.e. 'now').

    An offset-aware string is **converted** to UTC; a naive one is assumed UTC.

    [SESSION 2026-07-15] cli.py previously did `fromisoformat(s).replace(
    tzinfo=utc)`, which parsed the offset and then overwrote it rather than
    converting: `--start 2026-08-01T00:00:00+10:00` became 00:00Z, a silent 10 h
    shift for exactly the Australia/Brisbane users this tool targets.
    """
    if not s:
        return None
    dt = datetime.fromisoformat(s.strip())
    return dt.astimezone(timezone.utc) if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


def validate_window(days: float) -> None:
    """Raise ValueError unless 0 < days <= MAX_WINDOW_DAYS."""
    if not (0.0 < days <= MAX_WINDOW_DAYS):
        raise ValueError(f"days must be within (0, {MAX_WINDOW_DAYS:g}]; got {days:g}")


def dt_to_jd(dt: datetime) -> tuple[float, float]:
    dt = dt.astimezone(timezone.utc)
    jd, fr = jday(dt.year, dt.month, dt.day, dt.hour, dt.minute,
                  dt.second + dt.microsecond / 1e6)
    return jd, fr


def jd_to_dt(jd: float, fr: float) -> datetime:
    days = jd - 2440587.5 + fr  # Unix epoch JD
    return datetime.fromtimestamp(days * 86400.0, tz=timezone.utc)


# --------------------------------------------------------------------------
# Solar position (Astronomical Almanac low-precision, geometric, no refraction)
# --------------------------------------------------------------------------
def sun_position_deg(dt: datetime, lat_deg: float, lon_deg: float
                     ) -> tuple[float, float, float]:
    """(elevation°, azimuth° from N clockwise, true-solar-time hours).

    Astronomical Almanac low-precision solar position — geometric, no
    refraction; ~0.01° over 1950–2050. [LITERATURE — U.S. Naval Observatory,
    Astronomical Almanac, "Low precision formulae for the Sun".]

    Replaces the Spencer/NOAA-approximate series [SESSION 2026-07-15]. That
    series keyed off day-of-year with no year term, so it could not track the
    leap-year cycle: it drifted up to 0.45° in elevation (enough to mis-tier a
    pass against the sun ≥ 20° floor) and disagreed with the validated
    regression baseline by ~0.2°, while the baseline itself agrees with this
    algorithm to ~0.002°. Azimuth now derives from the same hour angle via
    atan2, which removes two further defects in the old branch: an azimuth
    computed from South but returned as if measured from North (every value
    reflected about the N–S axis), and an unwrapped hour angle that inverted
    the AM/PM branch for any |lon| ≳ 45°. Pinned by test_regression_baseline.py.
    """
    dt = dt.astimezone(timezone.utc)
    jd, fr = dt_to_jd(dt)
    n = (jd - 2451545.0) + fr                            # days from J2000.0
    L = math.radians((280.460 + 0.9856474 * n) % 360.0)  # mean longitude
    g = math.radians((357.528 + 0.9856003 * n) % 360.0)  # mean anomaly
    lam = (L + math.radians(1.915) * math.sin(g)
           + math.radians(0.020) * math.sin(2.0 * g))    # ecliptic longitude
    eps = math.radians(23.439 - 4.0e-7 * n)              # obliquity of ecliptic
    ra = math.atan2(math.cos(eps) * math.sin(lam), math.cos(lam))
    decl = math.asin(math.sin(eps) * math.sin(lam))
    # Local hour angle, wrapped to ±π so the AM/PM sense is always correct.
    ha = float(gmst_rad(jd + fr)) + lon_deg * DEG - ra
    ha = (ha + math.pi) % (2.0 * math.pi) - math.pi
    lat = lat_deg * DEG
    sin_el = (math.sin(lat) * math.sin(decl)
              + math.cos(lat) * math.cos(decl) * math.cos(ha))
    el = math.asin(max(-1.0, min(1.0, sin_el)))
    # Azimuth from N, clockwise. atan2 resolves the quadrant directly, so no
    # AM/PM branch and no acos domain clamp are needed.
    az = math.atan2(-math.cos(decl) * math.sin(ha),
                    math.sin(decl) * math.cos(lat)
                    - math.cos(decl) * math.sin(lat) * math.cos(ha))
    tst = (math.degrees(ha) / 15.0 + 12.0) % 24.0        # true (apparent) solar
    return math.degrees(el), math.degrees(az) % 360.0, tst


def sun_elevation_deg(dt: datetime, lat_deg: float, lon_deg: float) -> float:
    return sun_position_deg(dt, lat_deg, lon_deg)[0]


# --------------------------------------------------------------------------
# TLE acquisition
# --------------------------------------------------------------------------
@dataclass
class TLE:
    name: str
    catnr: int
    line1: str
    line2: str
    fetched_utc: str

    @property
    def epoch_utc(self) -> datetime:
        sat = Satrec.twoline2rv(self.line1, self.line2)
        return jd_to_dt(sat.jdsatepoch, sat.jdsatepochF)


_MU_EARTH = 398600.4418          # km^3 s^-2


def _semi_major_km(sat: Satrec) -> float:
    """Semi-major axis from the SGP4 mean motion (Kozai), km."""
    n = sat.no_kozai / 60.0                       # rad/min -> rad/s
    return (_MU_EARTH / (n * n)) ** (1.0 / 3.0)


def orbit_change(old: TLE, new: TLE) -> dict | None:
    """Quantify how an object's orbit changed between two element sets.

    Returns None if the pair is unusable (same/reversed epochs, bad elements).

    `manoeuvred` keys off a **rise** in semi-major axis, which is unambiguous:
    atmospheric drag can only lower it, so thrust is the sole explanation. That
    makes the test robust without modelling drag, which varies with altitude and
    solar activity. [SESSION 2026-07-15 — measured over ~1 day of real Celestrak
    elements: DRAG01/02/03/05 all decayed 6-18 m, DRAG04 rose 113 m.]

    `pos_err_km` / `along_track_s` are the actionable numbers: how badly the OLD
    set would have predicted the NEW set's own epoch — i.e. the error a user
    running on `gap_days`-old elements would actually have suffered.
    """
    gap = (new.epoch_utc - old.epoch_utc).total_seconds() / 86400.0
    if gap <= 0:
        return None
    so = Satrec.twoline2rv(old.line1, old.line2)
    sn = Satrec.twoline2rv(new.line1, new.line2)
    if getattr(so, "error", 0) or getattr(sn, "error", 0):
        return None
    da = _semi_major_km(sn) - _semi_major_km(so)
    jd, fr = dt_to_jd(new.epoch_utc)
    eo, ro, _ = so.sgp4(jd, fr)
    en, rn, vn = sn.sgp4(jd, fr)
    if eo != 0 or en != 0:
        return None
    ro, rn, vn = np.asarray(ro), np.asarray(rn), np.asarray(vn)
    d = ro - rn
    along_km = float(np.dot(d, vn / np.linalg.norm(vn)))
    return dict(gap_days=round(gap, 2),
                da_km=round(da, 4),
                da_km_per_day=round(da / gap, 4),
                pos_err_km=round(float(np.linalg.norm(d)), 3),
                along_track_s=round(abs(along_km) / 7.5, 2),
                manoeuvred=bool(da / gap > MANOEUVRE_DA_RISE_KM_PER_DAY))


def _manoeuvre_warning(name: str, ch: dict) -> str:
    return (
        f"{name}: MANOEUVRE detected — semi-major axis rose {ch['da_km'] * 1000:.0f} m "
        f"over {ch['gap_days']:.2f} d between the last two element sets (drag can only "
        f"lower it, so this is thrust). The superseded set mispredicted by "
        f"{ch['pos_err_km']:.1f} km ({ch['along_track_s']:.2f} s along-track). This "
        f"satellite is actively manoeuvring: the TLE-age timing σ assumes free flight "
        f"and will understate error if it burns again after the current epoch. "
        f"Re-fetch immediately before committing a tasking order.")


def _manoeuvre_warnings(cache: dict | None, fresh: dict[str, TLE]) -> list[str]:
    """One warning per satellite whose orbit changed by thrust since the cache.

    Never raises: a manoeuvre check failing must not break TLE acquisition.
    """
    if not cache:
        return []
    out: list[str] = []
    for name, new in fresh.items():
        try:
            prev = _cache_to_tles(cache, {name: new.catnr}).get(name)
            if prev is None:
                continue
            ch = orbit_change(prev, new)
            if ch and ch["manoeuvred"]:
                out.append(_manoeuvre_warning(name, ch))
        except Exception:
            continue
    return out


def fetch_tles(satellites: dict[str, int] = SATELLITES,
               cache_path: Path = DEFAULT_CACHE,
               cache_ttl_hours: float = 8.0,
               offline_file: str | Path | None = None,
               http_get: Callable[[str], str] | None = None,
               ) -> tuple[dict[str, TLE], list[str]]:
    """Return ({name: TLE}, warnings). Order of preference:
       offline_file > fresh cache > Celestrak fetch > stale cache (warned)."""
    warnings: list[str] = []

    if offline_file:
        return _parse_3le_file(Path(offline_file).read_text(), satellites), warnings

    cache = _load_cache(cache_path)
    now = time.time()
    if (cache and now - cache.get("_ts", 0) < cache_ttl_hours * 3600
            and _cache_covers(cache, satellites)):
        return _cache_to_tles(cache, satellites), warnings

    # Negative caching: on failure the cache records `_fail_ts`, and we refuse to
    # re-attempt for FETCH_RETRY_COOLDOWN_S, serving the stale cache instead.
    # Without it, a Celestrak outage turns every request into 5 more requests —
    # the classic retry storm that converts a soft throttle into a hard IP ban,
    # which takes the whole tool offline. Celestrak asks for <=2-3 polls/file/day.
    # [SESSION 2026-07-15]
    if (cache and cache.get("_fail_ts")
            and now - cache["_fail_ts"] < FETCH_RETRY_COOLDOWN_S
            and _cache_covers(cache, satellites)):
        age_h = (now - cache.get("_ts", 0)) / 3600.0
        warnings.append(
            f"Celestrak fetch failed {(now - cache['_fail_ts']) / 60.0:.0f} min ago; "
            f"not retrying for {FETCH_RETRY_COOLDOWN_S / 60.0:.0f} min. Using cached "
            f"TLEs {age_h:.1f} h old. Re-run before committing a tasking order.")
        return _cache_to_tles(cache, satellites), warnings

    if http_get is None:
        import requests
        def http_get(url: str) -> str:            # noqa: E306
            # Celestrak asks clients to identify themselves and blocks generic
            # scripted agents; the default "python-requests/x.y" risks a 403 that
            # would silently degrade every prediction to stale elements.
            # [SESSION 2026-07-15 — fetch_real_data.py already sets one.]
            r = requests.get(url, headers=CELESTRAK_UA, timeout=20)
            r.raise_for_status()
            return r.text

    out: dict[str, TLE] = {}
    fetched_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    with _FETCH_LOCK:                # one fetcher; the rest wait and reuse the result
        cache = _load_cache(cache_path)          # may have been filled while we waited
        if (cache and time.time() - cache.get("_ts", 0) < cache_ttl_hours * 3600
                and _cache_covers(cache, satellites)):
            return _cache_to_tles(cache, satellites), warnings
        try:
            for name, catnr in satellites.items():
                text = http_get(CELESTRAK_URL.format(catnr=catnr))
                lines = [ln.strip() for ln in text.strip().splitlines() if ln.strip()]
                l1 = next((ln for ln in lines if ln.startswith("1 ")), None)
                l2 = next((ln for ln in lines if ln.startswith("2 ")), None)
                if not (l1 and l2):
                    raise ValueError(
                        f"Celestrak returned no TLE for {name} ({catnr}): {text[:80]!r}")
                out[name] = TLE(name, catnr, l1, l2, fetched_at)
            # Compare each fresh set against the one it is about to replace: the
            # cache is the only orbit history we have, so this is the last moment
            # the comparison is possible. [SESSION 2026-07-15 — this is how the
            # DRAG04 burn was found; see IMPROVEMENTS.md A4-bis.]
            warnings.extend(_manoeuvre_warnings(cache, out))
            _save_cache(cache_path, out)
            return out, warnings
        except Exception as exc:  # network down, rate limited, etc.
            _record_fetch_failure(cache_path)
            if cache:
                age_h = (now - cache.get("_ts", 0)) / 3600.0
                warnings.append(
                    f"Celestrak fetch failed ({exc}); using cached TLEs {age_h:.1f} h old. "
                    "Re-run before committing a tasking order.")
                return _cache_to_tles(cache, satellites), warnings
            raise RuntimeError(
                f"Celestrak fetch failed and no cache available: {exc}") from exc


def _parse_3le_file(text: str, satellites: dict[str, int]) -> dict[str, TLE]:
    """Parse a saved TLE/3LE file, matching on NORAD number."""
    lines = [ln.rstrip() for ln in text.splitlines() if ln.strip()]
    by_catnr: dict[int, tuple[str, str]] = {}
    for i, ln in enumerate(lines):
        if ln.startswith("1 ") and i + 1 < len(lines) and lines[i + 1].startswith("2 "):
            by_catnr[int(ln[2:7])] = (ln, lines[i + 1])
    out, ts = {}, datetime.now(timezone.utc).isoformat(timespec="seconds")
    for name, catnr in satellites.items():
        if catnr not in by_catnr:
            raise ValueError(f"TLE file missing NORAD {catnr} ({name})")
        l1, l2 = by_catnr[catnr]
        out[name] = TLE(name, catnr, l1, l2, ts + " (offline file)")
    return out


def _load_cache(path: Path) -> dict | None:
    try:
        return json.loads(path.read_text())
    except Exception:
        return None


def _write_cache_atomic(path: Path, blob: dict) -> None:
    """Write via temp file + os.replace so a concurrent reader never sees a torn
    file (which `_load_cache` would swallow, triggering another full fetch).
    [SESSION 2026-07-15]"""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + f".{os.getpid()}.tmp")
    tmp.write_text(json.dumps(blob))
    os.replace(tmp, path)


def _save_cache(path: Path, tles: dict[str, TLE]) -> None:
    try:
        blob = {"_ts": time.time()}
        for name, t in tles.items():
            blob[name] = {"catnr": t.catnr, "l1": t.line1, "l2": t.line2,
                          "fetched": t.fetched_utc}
        _write_cache_atomic(path, blob)
    except Exception:
        pass  # cache is best-effort


def _record_fetch_failure(path: Path) -> None:
    """Stamp the cache with a failure time, preserving any TLEs already in it.

    This is what stops an outage becoming a retry storm: subsequent calls see
    `_fail_ts` and serve the stale cache instead of re-hitting all five URLs.
    [SESSION 2026-07-15]"""
    try:
        blob = _load_cache(path) or {}
        blob["_fail_ts"] = time.time()
        _write_cache_atomic(path, blob)
    except Exception:
        pass  # best-effort; a failed failure-record must never mask the real error


def _cache_covers(cache: dict, satellites: dict[str, int]) -> bool:
    """True only if the cache holds every requested satellite. The cache is one
    file shared across sensor profiles, so a run that fetched Dragonette leaves it
    'fresh' but without LANDSAT8/9 — the fresh-cache short-circuit must check
    coverage or it serves (or crashes on) the wrong constellation. [SESSION 2026-07-20]"""
    return bool(cache) and all(cache.get(name) for name in satellites)


def _cache_to_tles(cache: dict, satellites: dict[str, int]) -> dict[str, TLE]:
    out = {}
    for name in satellites:
        e = cache.get(name)
        if not e:
            raise RuntimeError(f"Cache missing {name}; delete cache and re-run online")
        out[name] = TLE(name, e["catnr"], e["l1"], e["l2"], e["fetched"])
    return out


# --------------------------------------------------------------------------
# Pass prediction
# --------------------------------------------------------------------------
@dataclass
class Pass:
    satellite: str
    tca_utc: datetime
    off_nadir_deg: float          # signed, Wyvern convention
    sun_elev_deg: float
    max_off_nadir_aoi_deg: float
    slant_range_km: float
    tle_epoch_utc: datetime
    category: str                 # "standard" | "marginal"
    operational: bool = True      # False => satellite not yet commissioned (R5)
    cloud: "CloudInfo | None" = None   # populated by attach_cloud() (R6)
    geometry: dict = field(default_factory=dict)  # A1 acquisition-quality metrics
    node: str = ""                # "ascending" | "descending" (A2)
    local_solar_time_h: float | None = None       # at AOI, hours (A2)
    quality: str = ""             # "good" | "marginal" | "poor" illumination badge (A2)
    timing_sigma_s: float | None = None           # TLE-age timing uncertainty (A4)
    coverage_pct: float | None = None             # fraction of AOI in swath, 0–1 (B1)
    footprint_lonlat: list = field(default_factory=list)  # swath ground polygon (B1)


@dataclass
class Prediction:
    aoi: AOI
    start_utc: datetime
    end_utc: datetime
    passes: list[Pass] = field(default_factory=list)
    marginal: list[Pass] = field(default_factory=list)
    nonoperational: list[Pass] = field(default_factory=list)  # R5: kept separate
    warnings: list[str] = field(default_factory=list)
    params: dict = field(default_factory=dict)
    # {date -> mean total cloud %} within the deterministic forecast horizon
    # (populated by attach_cloud); days outside it are absent => "unknown".
    cloud_daily: dict = field(default_factory=dict)
    summary: dict = field(default_factory=dict)   # B2 revisit / best-window analytics


def campaign_summary(pred: "Prediction", window_days: float = 3.0) -> dict:
    """B2 analytics over the operational pass list: quality-pass count, revisit
    gaps, and the best contiguous `window_days` window for a field campaign.
    Pure Python over data predict() already produced — no propagation."""
    ops = sorted(pred.passes + pred.marginal, key=lambda p: p.tca_utc)
    good = [p for p in ops if p.quality == "good"]
    out: dict = {"total_opportunities": len(ops),
                 "good_quality": len(good),
                 "window_days": pred.params.get("days")}
    # revisit gaps between consecutive quality opportunities
    if len(good) >= 2:
        gaps = [(good[i + 1].tca_utc - good[i].tca_utc).total_seconds() / 3600.0
                for i in range(len(good) - 1)]
        gaps.sort()
        mid = gaps[len(gaps) // 2] if len(gaps) % 2 else \
            (gaps[len(gaps) // 2 - 1] + gaps[len(gaps) // 2]) / 2
        out["median_revisit_h"] = round(mid, 1)
        out["max_gap_h"] = round(max(gaps), 1)
    # best contiguous window: maximise good-quality passes, tie-break on mean
    # off-nadir (lower better) then mean sun (higher better)
    best = None
    span = timedelta(days=window_days)
    for i, p0 in enumerate(good):
        win = [p for p in good if p0.tca_utc <= p.tca_utc <= p0.tca_utc + span]
        if not win:
            continue
        mean_off = sum(abs(p.off_nadir_deg) for p in win) / len(win)
        mean_sun = sum(p.sun_elev_deg for p in win) / len(win)
        score = (len(win), -mean_off, mean_sun)
        if best is None or score > best[0]:
            best = (score, win, mean_off, mean_sun)
    if best:
        _, win, mean_off, mean_sun = best
        out["best_window"] = {
            "start_utc": win[0].tca_utc.isoformat(timespec="seconds"),
            "end_utc": win[-1].tca_utc.isoformat(timespec="seconds"),
            "good_passes": len(win),
            "mean_off_nadir_deg": round(mean_off, 1),
            "mean_sun_elev_deg": round(mean_sun, 1),
            "satellites": sorted({p.satellite for p in win}),
        }
    return out


def merge_predictions(parts: "list[Prediction]") -> "Prediction":
    """Merge same-AOI predictions from several sensor profiles into one, for the
    combined 'all sensors' view. Concatenates the pass buckets (time-sorted) and
    marks the result sensor='all' so the timeline and JSON build a union roster.
    Cloud is attached afterwards, on the merged result. [SESSION 2026-07-20]"""
    base = parts[0]
    by_time = lambda ps: sorted(ps, key=lambda p: p.tca_utc)
    merged = Prediction(
        aoi=base.aoi,
        start_utc=min(p.start_utc for p in parts),
        end_utc=max(p.end_utc for p in parts),
        passes=by_time([p for part in parts for p in part.passes]),
        marginal=by_time([p for part in parts for p in part.marginal]),
        nonoperational=by_time([p for part in parts for p in part.nonoperational]),
        warnings=list(dict.fromkeys(w for part in parts for w in part.warnings)),
        params={**base.params, "sensor": COMBINED_KEY,
                "sensor_display": COMBINED_DISPLAY},
    )
    merged.summary = campaign_summary(merged)
    return merged


def _off_nadir_series(sat: Satrec, jd: np.ndarray, fr: np.ndarray,
                      site_ecef: np.ndarray
                      ) -> tuple[np.ndarray, np.ndarray, np.ndarray, dict[int, int]]:
    """(eta, r, v, sgp4_errors). Errored epochs are pushed to eta=180 so they
    can never look like an opportunity; the error counts are returned so the
    caller can say so out loud rather than silently reporting 'no passes'."""
    err, r, v = sat.sgp4_array(jd, fr)
    r = np.asarray(r); v = np.asarray(v)
    theta = gmst_rad(jd + fr)
    site_teme = ecef_to_teme(site_ecef, theta)           # (N,3)
    rho = site_teme - r
    bad = err != 0
    with np.errstate(invalid="ignore", divide="ignore"):
        cosang = -np.einsum("ij,ij->i", r, rho) / (
            np.linalg.norm(r, axis=1) * np.linalg.norm(rho, axis=1))
    eta = np.degrees(np.arccos(np.clip(cosang, -1.0, 1.0)))
    eta[bad] = 180.0
    codes = {int(c): int((err == c).sum()) for c in np.unique(err[bad])} if bad.any() else {}
    return eta, r, v, codes


def _eta_at(sat: Satrec, jd0: float, fr0: float, t_s: float,
            site_ecef: np.ndarray, nadir_ellipsoid: bool = False) -> float:
    jd, fr = jd0, fr0 + t_s / 86400.0
    err, r, _ = sat.sgp4(jd, fr)
    if err != 0:
        return 180.0
    r = np.asarray(r)
    theta = float(gmst_rad(jd + fr))
    site_teme = ecef_to_teme(site_ecef, theta)
    rho = site_teme - r
    nadir = nadir_unit_teme(r, theta, nadir_ellipsoid)
    c = float(np.dot(nadir, rho) / np.linalg.norm(rho))
    return math.degrees(math.acos(max(-1.0, min(1.0, c))))


def _golden(f: Callable[[float], float], a: float, b: float,
            tol: float = 0.1) -> float:
    invphi = (math.sqrt(5) - 1) / 2
    c, d = b - invphi * (b - a), a + invphi * (b - a)
    fc, fd = f(c), f(d)
    while b - a > tol:
        if fc < fd:
            b, d, fd = d, c, fc
            c = b - invphi * (b - a); fc = f(c)
        else:
            a, c, fc = c, d, fd
            d = a + invphi * (b - a); fd = f(d)
    return (a + b) / 2


def predict(kmz_bytes: bytes,
            days: float = 14.0,
            start_utc: datetime | None = None,
            terrain_alt_m: float = 0.0,
            max_off_nadir_deg: float | None = None,
            min_sun_elev_deg: float | None = None,
            marginal_off_nadir_deg: float | None = None,
            marginal_sun_elev_deg: float | None = None,
            coarse_step_s: float = 20.0,
            polygon_name: str | None = None,
            satellites: dict[str, int] | None = None,
            tles: dict[str, TLE] | None = None,
            offline_tle_file: str | Path | None = None,
            include_nonoperational: bool = True,
            nadir_ellipsoid: bool = False,
            operational: dict[str, bool] | None = None,
            profile: "SensorProfile | str | None" = None,
            ) -> Prediction:
    """Run the full pipeline. Supply `tles` to skip fetching (tests/offline).

    Non-operational satellites (OPERATIONAL[name] is False, e.g. DRAG05) are
    predicted but routed to `pred.nonoperational` — never `passes`/`marginal` —
    so they are never counted or shown as taskable (R5, CLAUDE.md constraint 3).
    Set include_nonoperational=False to drop them entirely.

    nadir_ellipsoid=True measures off-nadir from the WGS84 ellipsoid normal
    instead of geocentric −r̂ (up to ~0.2° difference; physically the correct
    'nadir' but off the validated geocentric baseline — opt-in). [SESSION]

    `profile` selects the sensor (a SensorProfile or its key, e.g. "landsat",
    "sentinel2"); it supplies the satellite set, operational map and access
    envelope. Default is DRAGONETTE, and every explicit argument still wins over
    the profile — so existing callers are unaffected. [SESSION 2026-07-15]"""
    prof = profile if isinstance(profile, SensorProfile) else get_profile(profile)
    # Profile supplies the defaults; an explicit argument always overrides it.
    max_off_nadir_deg = (prof.max_off_nadir_deg if max_off_nadir_deg is None
                         else max_off_nadir_deg)
    min_sun_elev_deg = (prof.min_sun_elev_deg if min_sun_elev_deg is None
                        else min_sun_elev_deg)
    marginal_off_nadir_deg = (prof.marginal_off_nadir_deg if marginal_off_nadir_deg is None
                              else marginal_off_nadir_deg)
    marginal_sun_elev_deg = (prof.marginal_sun_elev_deg if marginal_sun_elev_deg is None
                             else marginal_sun_elev_deg)
    satellites = prof.satellites if satellites is None else satellites

    aoi = parse_kmz(kmz_bytes, terrain_alt_m, polygon_name)
    start = (start_utc or datetime.now(timezone.utc)).astimezone(timezone.utc)
    end = start + timedelta(days=days)

    op_map = operational if operational is not None else prof.operational

    warnings: list[str] = []
    if tles is None:
        tles, warnings = fetch_tles(satellites, offline_file=offline_tle_file)

    jd0, fr0 = dt_to_jd(start)
    n_steps = int(days * 86400.0 / coarse_step_s) + 1
    t_grid = np.arange(n_steps) * coarse_step_s
    jd = np.full(n_steps, jd0)
    fr = fr0 + t_grid / 86400.0

    site_ecef = aoi.centroid_ecef
    verts_ecef = aoi.vertices_ecef

    pred = Prediction(aoi=aoi, start_utc=start, end_utc=end, warnings=list(warnings),
                      params=dict(days=days, terrain_alt_m=terrain_alt_m,
                                  max_off_nadir_deg=max_off_nadir_deg,
                                  min_sun_elev_deg=min_sun_elev_deg,
                                  marginal_off_nadir_deg=marginal_off_nadir_deg,
                                  marginal_sun_elev_deg=marginal_sun_elev_deg,
                                  coarse_step_s=coarse_step_s,
                                  include_nonoperational=include_nonoperational,
                                  nadir_ellipsoid=nadir_ellipsoid,
                                  swath_km=prof.swath_km,
                                  gsd_nadir_m=prof.gsd_m,
                                  sensor=prof.key,
                                  sensor_display=prof.display,
                                  agile=prof.agile))

    for name, tle in tles.items():
        op = op_map.get(name, True)
        if not op and not include_nonoperational:
            continue  # R5: drop non-operational satellites entirely
        sat = Satrec.twoline2rv(tle.line1, tle.line2)
        if getattr(sat, "error", 0):
            pred.warnings.append(
                f"{name}: SGP4 rejected the element set (error {sat.error}); "
                "no passes reported for this satellite — this is a data problem, "
                "not an empty sky.")
            continue
        eta, _, _, sgp4_errs = _off_nadir_series(sat, jd, fr, site_ecef)
        if sgp4_errs:
            # Silently dropping these makes a decayed or mis-parsed element set
            # look exactly like a satellite that genuinely had no opportunities.
            # [SESSION 2026-07-15]
            n_bad = sum(sgp4_errs.values())
            pred.warnings.append(
                f"{name}: SGP4 returned errors on {n_bad}/{len(eta)} epochs "
                f"(codes {sorted(sgp4_errs)}); results for this satellite are "
                "incomplete — check TLE age/validity.")

        interior = np.flatnonzero(
            (eta[1:-1] < eta[:-2]) & (eta[1:-1] <= eta[2:]) & (eta[1:-1] < 70.0)) + 1

        for k in interior:
            f = lambda t: _eta_at(sat, jd0, fr0, t, site_ecef,   # noqa: E731
                                  nadir_ellipsoid)
            t_star = _golden(f, t_grid[k - 1], t_grid[k + 1], tol=0.1)
            tca = start + timedelta(seconds=t_star)

            err, r_t, v_t = sat.sgp4(jd0, fr0 + t_star / 86400.0)
            if err != 0:
                continue
            r_t, v_t = np.asarray(r_t), np.asarray(v_t)
            theta = float(gmst_rad(jd0 + fr0 + t_star / 86400.0))
            rho = ecef_to_teme(site_ecef, theta) - r_t
            eta_mag = f(t_star)
            slant = float(np.linalg.norm(rho))

            # Visibility gate: a satellite behind the Earth also shows a small
            # off-nadir angle (LOS points near-nadir through the planet).
            # Require the satellite above the site's geometric horizon.
            sat_ecef = teme_to_ecef(r_t, theta)
            to_sat = sat_ecef - site_ecef
            up = geodetic_up(aoi.centroid_lat, aoi.centroid_lon)
            sat_elev = 90.0 - math.degrees(math.acos(
                max(-1.0, min(1.0, float(np.dot(to_sat, up)) / np.linalg.norm(to_sat)))))
            if sat_elev <= 0.0:
                continue

            h_hat = np.cross(r_t, v_t)
            h_hat /= np.linalg.norm(h_hat)
            side = math.copysign(1.0, float(np.dot(rho, h_hat)))
            eta_signed = side * eta_mag
            if SIGN_FLIP_TO_MATCH_WYVERN:
                eta_signed = -eta_signed

            nadir = nadir_unit_teme(r_t, theta, nadir_ellipsoid)
            rho_v = ecef_to_teme(verts_ecef, theta) - r_t        # (M,3)
            cosv = rho_v @ nadir / np.linalg.norm(rho_v, axis=1)
            eta_max = float(np.degrees(np.arccos(np.clip(cosv, -1, 1))).max())

            sun, sun_az, lst = sun_position_deg(tca, aoi.centroid_lat, aoi.centroid_lon)

            # Access envelope: off-nadir magnitude within the standard tier and
            # sun above the floor, else the widened marginal tier, else no access.
            if eta_mag <= max_off_nadir_deg and sun >= min_sun_elev_deg:
                std = True
            elif eta_mag <= marginal_off_nadir_deg and sun >= marginal_sun_elev_deg:
                std = False
            else:
                continue

            # A1: advisory acquisition-quality geometry (never gates access).
            # GSD comes from the sensor profile — Landsat 30 m, Sentinel-2 10 m,
            # Dragonette 5.3 m. [SESSION 2026-07-15]
            geom = acquisition_geometry(r_t, v_t, theta, site_ecef,
                                        aoi.centroid_lat, aoi.centroid_lon,
                                        eta_signed, sun, sun_az,
                                        gsd_nadir_m=prof.gsd_m)
            # A2: quality badge from sun elevation, off-nadir and glint.
            badge = _quality_badge(eta_mag, sun, geom["sun_glint_deg"])
            # A4: TLE-age timing uncertainty (along-track drift → timing jitter)
            age_d = max(0.0, (tca - tle.epoch_utc).total_seconds() / 86400.0)
            tsigma = round((1.0 + 2.0 * age_d) / 7.5, 2)     # km→s at ~7.5 km/s
            # B1: swath footprint + AOI coverage fraction, at the sensor's own
            # swath (185 km Landsat / 290 km Sentinel-2 / 20 km Dragonette).
            # For a push-broom, derive the swath from the profile's own FOV rather
            # than its nominal swath_km, so the footprint cannot contradict the
            # access gate above. prof.max_off_nadir_deg (the optic), not the
            # possibly-overridden gate: widening the gate does not widen the lens.
            fp = swath_footprint_lonlat(
                r_t, v_t, theta, aoi, prof.swath_km, agile=prof.agile,
                fov_half_deg=None if prof.agile else prof.max_off_nadir_deg)
            cov = aoi_coverage_fraction(fp, aoi)

            p = Pass(name, tca, round(eta_signed, 1), round(sun, 4),
                     round(eta_max, 1), round(slant), tle.epoch_utc,
                     "standard" if std else "marginal", operational=op,
                     geometry=geom, node=geom["node"],
                     local_solar_time_h=round(lst, 2), quality=badge,
                     timing_sigma_s=tsigma, coverage_pct=cov,
                     footprint_lonlat=[[round(x, 6), round(y, 6)] for x, y in fp])
            if not op:                       # R5: separate, uncounted bucket
                pred.nonoperational.append(p)
            elif std:
                pred.passes.append(p)
            else:
                pred.marginal.append(p)

    pred.passes.sort(key=lambda p: p.tca_utc)
    pred.marginal.sort(key=lambda p: p.tca_utc)
    pred.nonoperational.sort(key=lambda p: p.tca_utc)

    # Age is reported over the satellites that could actually contribute, so the
    # staleness warning cannot fire on a DRAG05 that was excluded. `default=None`
    # because an empty TLE set used to raise a bare `max() iterable argument is
    # empty` instead of returning an empty prediction. [SESSION 2026-07-15]
    contributing = [t for name, t in tles.items()
                    if op_map.get(name, True) or include_nonoperational]
    max_age = max(((start - t.epoch_utc).total_seconds() / 86400.0
                   for t in contributing), default=None)
    if max_age is None:
        pred.warnings.append(
            "No TLEs available for the requested satellites; no passes predicted.")
    elif max_age > 3.0:
        pred.warnings.append(
            f"Oldest TLE is {max_age:.1f} days old at window start; "
            "expect growing along-track timing error.")
    pred.warnings.append(
        "SGP4 timing error grows with TLE age: ~±1 min over days 1–7, up to "
        "several minutes (and a few degrees off-nadir) by day 14. Re-run on "
        "fresh TLEs before committing a tasking order.")
    pred.summary = campaign_summary(pred)        # B2 revisit / best-window
    return pred


# --------------------------------------------------------------------------
# Cloud cover — three-tier scheme (R6; full reference in CLOUD.md)
# --------------------------------------------------------------------------
# Cloud skill decays fast: deterministic to ~day 5, probabilistic (ensemble)
# 5–10 d, climatology beyond. Tier is keyed off lead time from the window start
# (== "now" for a default run). [SESSION 2026-07-14]  All endpoint/param facts
# are [LITERATURE: open-meteo.com docs 2026-07-14] — re-verify before live use.
FORECAST_URL = ("https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}"
                "&hourly=cloud_cover,cloud_cover_low,cloud_cover_mid,cloud_cover_high"
                "&forecast_days=16&timezone=UTC")
# ecmwf_ifs025 = 51 members, 3-hourly, runs to 15 d [VERIFIED open-meteo 2026-07].
# Nearest-hour lookup snaps a TCA to the nearest 3-h step (≤±90 min) — acceptable.
ENSEMBLE_URL = ("https://ensemble-api.open-meteo.com/v1/ensemble?latitude={lat}"
                "&longitude={lon}&hourly=cloud_cover&models=ecmwf_ifs025"
                "&forecast_days=16&timezone=UTC")
CLOUD_OK_THRESHOLD = 30.0        # % total cloud counted as "clear". [SESSION] A
# sensible default and USER-ADJUSTABLE at request time (SPA "Cloud threshold"
# field, CLI --cloud-threshold, API cloud_threshold). Note: Tier-3 climatology is
# computed against it, so regenerate sites_climatology.json if you change it
# (climatology.py; a drift guard flags a stale file).
TIER1_MAX_DAYS = 5.0
TIER2_MAX_DAYS = 15.0            # A5: was 10; ensemble has skill to ~15 d


@dataclass
class CloudInfo:
    tier: int                     # 1 forecast | 2 outlook | 3 climatology | 0 n/a
    label: str                    # "forecast" | "outlook (probabilistic)" | "" | "n/a"
    total: float | None = None    # tier 1: total cloud % at nearest hour
    low: float | None = None
    mid: float | None = None
    high: float | None = None
    p_clear: float | None = None  # tier 2: P(total cloud < threshold) across members
    spread: tuple[float, float] | None = None   # tier 2: (min, max) % across members
    threshold: float = CLOUD_OK_THRESHOLD
    clim_clear_pct: float | None = None          # tier 3 / base rate (climatology)
    total_band: tuple[float, float] | None = None  # tier 1: ± uncertainty band % (B3)

    @property
    def optical_obstruction(self) -> float | None:
        """A3: cloud weighted for OPTICAL imaging. Opaque low+mid fully block a
        scene; thin high cirrus (weight 0.4) often passes but degrades
        hyperspectral radiometry. [LITERATURE/GK — weight to confirm.]"""
        if self.low is None:
            return None
        return round(min(100.0, self.low + self.mid + 0.4 * self.high), 1)

    @property
    def cirrus_flag(self) -> bool:
        """High cloud >20% — hyperspectral radiometry likely degraded. [A3]"""
        return self.high is not None and self.high > 20.0

    @property
    def likely_cloudy(self) -> bool:
        """Visual aid only — never filters passes out. [SESSION]"""
        if self.tier == 1:
            ob = self.optical_obstruction        # A3: judge on optical obstruction
            if ob is not None:
                return ob > 60.0
            if self.total is not None:
                return self.total > 70.0
        if self.tier == 2 and self.p_clear is not None:
            return self.p_clear < 0.3
        return False

    def xlsx_cells(self) -> list:
        """Six cells appended after Slant Range, per CLOUD.md."""
        na = "n/a"
        total = f"{self.total:.0f}" if self.total is not None else na
        lmh = (f"{self.low:.0f}/{self.mid:.0f}/{self.high:.0f}"
               if self.low is not None else na)
        pcl = f"{self.p_clear:.2f}" if self.p_clear is not None else na
        spr = f"{self.spread[0]:.0f}–{self.spread[1]:.0f}" if self.spread else na
        clim = f"{self.clim_clear_pct:.0f}%" if self.clim_clear_pct is not None else na
        return [total, lmh, pcl, spr, self.label or "", clim]

    def json(self) -> dict:
        return {"tier": self.tier, "label": self.label, "total_pct": self.total,
                "low_pct": self.low, "mid_pct": self.mid, "high_pct": self.high,
                "p_clear": self.p_clear, "threshold_pct": self.threshold,
                "spread_pct": list(self.spread) if self.spread else None,
                "clim_clear_pct": self.clim_clear_pct,
                "total_band_pct": list(self.total_band) if self.total_band else None,
                "optical_obstruction_pct": self.optical_obstruction,
                "cirrus_flag": self.cirrus_flag,
                "likely_cloudy": self.likely_cloudy}


CLOUD_COLUMNS = ["Cloud % (total@TCA)", "Cloud low/mid/high %", "P(cloud<thr)",
                 "Cloud spread (min–max %)", "Cloud label", "Hist. clear-sky rate (mon)"]


def _default_cloud_get(url: str) -> str:
    import requests
    r = requests.get(url, timeout=20)
    r.raise_for_status()
    return r.text


def _parse_hourly_times(hourly: dict) -> list[datetime]:
    out = []
    for t in hourly["time"]:                     # naive ISO, timezone=UTC requested
        out.append(datetime.fromisoformat(t).replace(tzinfo=timezone.utc))
    return out


# A pass may sit beyond the end of the series a provider actually returned (the
# requested forecast_days is a ceiling, not a guarantee: ecmwf_ifs025 has come
# back with 11 d against a 16 d request). Snapping to the nearest sample would
# then silently report cloud from days earlier as if it were that pass's
# outlook, so refuse any match further than this from the TCA. Both the
# deterministic and ensemble series are hourly, so a genuine in-range match is
# always within 30 min. [SESSION 2026-07-15 — see _nearest_hour_index tests.]
_CLOUD_MAX_SNAP_H = 1.5


def _nearest_hour_index(times: list[datetime], tca: datetime,
                        max_gap_h: float = _CLOUD_MAX_SNAP_H) -> int | None:
    """Index of the sample nearest `tca`, or None if the series does not
    actually cover it (i.e. the nearest sample is > max_gap_h away)."""
    if not times:
        return None
    tca = tca.astimezone(timezone.utc)
    i = min(range(len(times)), key=lambda i: abs((times[i] - tca).total_seconds()))
    if abs((times[i] - tca).total_seconds()) > max_gap_h * 3600.0:
        return None
    return i


def _series_at(hourly: dict, key: str, idx: int) -> float | None:
    vals = hourly.get(key)
    if not vals or idx >= len(vals):
        return None
    v = vals[idx]
    return None if v is None else float(v)


def _ensemble_member_keys(hourly: dict) -> list[str]:
    # Control ("cloud_cover") + perturbed members ("cloud_cover_memberNN"). [LITERATURE]
    return [k for k in hourly
            if k == "cloud_cover" or k.startswith("cloud_cover_member")]


def _tier_for(lead_days: float) -> int:
    if lead_days < TIER1_MAX_DAYS:
        return 1
    if lead_days < TIER2_MAX_DAYS:
        return 2
    return 3


def attach_cloud(pred: Prediction,
                 threshold: float = CLOUD_OK_THRESHOLD,
                 http_get: Callable[[str], str] | None = None,
                 forecast_json: dict | None = None,
                 ensemble_json: dict | None = None,
                 climatology: dict | None = None,
                 now: datetime | None = None) -> Prediction:
    """Attach CloudInfo to every pass in-place (R6). Never raises, never blocks:
    on any fetch/parse failure the pass gets a tier-0 'n/a' CloudInfo and a
    warning is appended. Exactly one forecast + one ensemble call per AOI —
    batched, following the fetch_tles http_get injection pattern.

    Supply forecast_json/ensemble_json to run fully offline (tests). `now`
    defaults to the window start so tier assignment is deterministic."""
    ref = (now or pred.start_utc).astimezone(timezone.utc)
    all_passes = pred.passes + pred.marginal + pred.nonoperational
    if not all_passes:
        return pred

    tiers_needed = {_tier_for((p.tca_utc - ref).total_seconds() / 86400.0)
                    for p in all_passes}
    lat, lon = pred.aoi.centroid_lat, pred.aoi.centroid_lon

    def _get(url: str) -> str:
        return (http_get or _default_cloud_get)(url)

    fc, ens = forecast_json, ensemble_json
    if fc is None and 1 in tiers_needed:
        try:
            fc = json.loads(_get(FORECAST_URL.format(lat=lat, lon=lon)))
        except Exception as exc:
            pred.warnings.append(f"Cloud forecast (Tier 1) unavailable ({exc}); "
                                 "cloud columns show n/a.")
    if ens is None and 2 in tiers_needed:
        try:
            ens = json.loads(_get(ENSEMBLE_URL.format(lat=lat, lon=lon)))
        except Exception as exc:
            pred.warnings.append(f"Cloud ensemble outlook (Tier 2) unavailable ({exc}); "
                                 "cloud columns show n/a.")

    fc_times = _parse_hourly_times(fc["hourly"]) if fc else None
    ens_times = _parse_hourly_times(ens["hourly"]) if ens else None

    # Daily mean total cloud within the deterministic forecast horizon, for the
    # timeline cloud strip. Only days the forecast actually covers get a value.
    if fc:
        horizon = ref + timedelta(days=TIER1_MAX_DAYS)
        totals = fc["hourly"].get("cloud_cover") or []
        by_day: dict[str, list[float]] = {}
        for t, v in zip(fc_times, totals):
            if v is None or t < ref or t > horizon:
                continue
            by_day.setdefault(t.strftime("%Y-%m-%d"), []).append(float(v))
        pred.cloud_daily = {d: round(sum(vs) / len(vs), 1) for d, vs in by_day.items()}

    clim = climatology or {}
    month = ref.strftime("%b")
    clim_pct = (clim.get(pred.aoi.name, {}) or {}).get(month)

    if fc or ens or clim_pct is not None:
        pred.warnings.append("Weather data by Open-Meteo.com (CC BY 4.0).")

    beyond_horizon = 0
    for p in all_passes:
        lead = (p.tca_utc - ref).total_seconds() / 86400.0
        tier = _tier_for(lead)
        i = None
        if tier == 1 and fc:
            i = _nearest_hour_index(fc_times, p.tca_utc)
        elif tier == 2 and ens:
            i = _nearest_hour_index(ens_times, p.tca_utc)
        if i is None and tier in (1, 2) and (fc if tier == 1 else ens):
            # The provider returned a shorter series than the tier claims to
            # cover. Report n/a rather than inventing a value from the last
            # available hour. [SESSION 2026-07-15]
            beyond_horizon += 1
            p.cloud = CloudInfo(0, "n/a", threshold=threshold, clim_clear_pct=clim_pct)
            continue
        if tier == 1 and fc:
            h = fc["hourly"]
            total = _series_at(h, "cloud_cover", i)
            # B3: honest ± band — deterministic cloud skill decays with lead time
            # (~±10% at day 1 growing to ~±25% by day 5). [SESSION — calibrate via C1.]
            band = None
            if total is not None:
                halfw = 10.0 + 3.0 * min(lead, TIER1_MAX_DAYS)
                band = (round(max(0.0, total - halfw), 0),
                        round(min(100.0, total + halfw), 0))
            p.cloud = CloudInfo(1, "forecast", total=total,
                                low=_series_at(h, "cloud_cover_low", i),
                                mid=_series_at(h, "cloud_cover_mid", i),
                                high=_series_at(h, "cloud_cover_high", i),
                                threshold=threshold, clim_clear_pct=clim_pct,
                                total_band=band)
        elif tier == 2 and ens:
            h = ens["hourly"]
            members = [_series_at(h, k, i) for k in _ensemble_member_keys(h)]
            members = [m for m in members if m is not None]
            if members:
                # B3: soft membership instead of a hard count. Raw ensemble
                # fractions are under-dispersive (over-confident 0/1); a logistic
                # ramp of width ~8% around the threshold smooths the estimate so a
                # member at 29% and one at 31% don't flip the whole probability.
                w = 8.0
                p_clear = sum(1.0 / (1.0 + math.exp((m - threshold) / w))
                              for m in members) / len(members)
                p.cloud = CloudInfo(2, "outlook (probabilistic)",
                                    p_clear=round(p_clear, 3),
                                    spread=(min(members), max(members)),
                                    threshold=threshold, clim_clear_pct=clim_pct)
            else:
                p.cloud = CloudInfo(0, "n/a", threshold=threshold, clim_clear_pct=clim_pct)
        elif tier == 3:
            p.cloud = CloudInfo(3, "", threshold=threshold, clim_clear_pct=clim_pct)
        else:                                    # data missing for the tier we needed
            p.cloud = CloudInfo(0, "n/a", threshold=threshold, clim_clear_pct=clim_pct)
    if beyond_horizon:
        pred.warnings.append(
            f"{beyond_horizon} pass(es) fall beyond the end of the cloud series "
            "actually returned by Open-Meteo; their cloud columns show n/a rather "
            "than a value carried over from the last available hour.")
    return pred


def load_climatology(path: str | Path) -> dict:
    """Load per-site monthly clear-sky rates (R6 Tier 3). Missing/bad => {}."""
    try:
        return json.loads(Path(path).read_text())
    except Exception:
        return {}


# --------------------------------------------------------------------------
# JSON contract (digital-twin front-end; R9)
# --------------------------------------------------------------------------
SCHEMA_VERSION = "2.1"    # 2.1: added per-AOI "sensor" block (roster + operational)


def _pass_json(p: Pass) -> dict:
    d = {
        "satellite": p.satellite,
        "tca_utc": p.tca_utc.isoformat(timespec="seconds"),
        "off_nadir_deg": p.off_nadir_deg,
        "sun_elev_deg": p.sun_elev_deg,
        "max_off_nadir_aoi_deg": p.max_off_nadir_aoi_deg,
        "slant_range_km": p.slant_range_km,
        "tle_epoch_utc": p.tle_epoch_utc.isoformat(timespec="seconds"),
        "category": p.category,
        "operational": p.operational,
        "node": p.node,
        "local_solar_time_h": p.local_solar_time_h,
        "quality": p.quality,
        "timing_sigma_s": p.timing_sigma_s,
        "coverage_pct": p.coverage_pct,
        "footprint_lonlat": p.footprint_lonlat,
        "geometry": p.geometry,
    }
    if p.cloud is not None:
        d["cloud"] = p.cloud.json()
    return d


def _sensor_block(pred: Prediction) -> dict:
    """The sensor roster the front-end draws lanes from — a single profile, or the
    union for the combined 'all' view."""
    key = pred.params.get("sensor")
    if key == COMBINED_KEY:
        sats, op = combined_roster()
        return {"key": COMBINED_KEY,
                "display": pred.params.get("sensor_display", COMBINED_DISPLAY),
                "satellites": list(sats), "operational": op}
    prof = get_profile(key)
    return {"key": prof.key, "display": prof.display,
            "satellites": list(prof.satellites), "operational": prof.operational}


def _one_prediction_json(pred: Prediction) -> dict:
    return {
        "sensor": _sensor_block(pred),
        "aoi": {"name": pred.aoi.name,
                "centroid_lat": round(pred.aoi.centroid_lat, 5),
                "centroid_lon": round(pred.aoi.centroid_lon, 5),
                "terrain_alt_m": pred.aoi.terrain_alt_m,
                "vertices_lonlat": [[round(lon, 6), round(lat, 6)]
                                    for lon, lat in pred.aoi.vertices_lonlat]},
        "window_utc": [pred.start_utc.isoformat(timespec="seconds"),
                       pred.end_utc.isoformat(timespec="seconds")],
        "passes": [_pass_json(p) for p in pred.passes],
        "marginal": [_pass_json(p) for p in pred.marginal],
        "nonoperational": [_pass_json(p) for p in pred.nonoperational],
        "cloud_daily": pred.cloud_daily,          # {date -> mean total cloud %}
        "cloud_horizon_utc": (pred.start_utc + timedelta(days=TIER1_MAX_DAYS)
                              ).isoformat(timespec="seconds") if pred.cloud_daily else None,
        "summary": pred.summary,                  # B2 revisit / best-window
        "warnings": pred.warnings,
        "params": pred.params,
    }


def prediction_json(preds: list[Prediction]) -> dict:
    """Versioned JSON for the DT front-end (R9).

    One AOI keeps the flat top-level shape (aoi/passes/marginal/...) for
    backward compatibility; several AOIs nest under `aois`. Always carries
    `schema_version` so the consumer can branch safely."""
    if len(preds) == 1:
        return {"schema_version": SCHEMA_VERSION, **_one_prediction_json(preds[0])}
    return {"schema_version": SCHEMA_VERSION,
            "aois": [_one_prediction_json(pr) for pr in preds]}


# --------------------------------------------------------------------------
# Timeline chart (R4) — Gantt, one row per satellite, coloured by off-nadir band
# --------------------------------------------------------------------------
# Bars are coloured by off-nadir MAGNITUDE band (near-nadir is best for imaging)
# and labelled with the angle. DRAG05 (non-operational) keeps its band colour
# but is hatched + row-labelled so it never reads as taskable (constraint 3).
# [SESSION 2026-07-14, matches Wyvern-style opportunity charts.]
_OFFNADIR_BANDS = [                      # (upper bound inclusive, colour, label)
    (5.0,  "#22c55e", "0–5° near nadir"),
    (12.0, "#f59e0b", "5–12° moderate roll"),
    (999.0, "#ef4444", ">12° high roll"),
]
# Cloud-strip cell colours by daily mean total cloud %.
_CLOUD_BANDS = [(30.0, "#bcd9f5", "clear"), (70.0, "#9aa7b4", "partly cloudy"),
                (100.1, "#5c6b7f", "cloudy")]
_CLOUD_UNKNOWN = "#e9edf1"               # hatched => beyond forecast horizon
_BAR_HALF_WIDTH_DAYS = 3.0 / 24.0        # ±3 h so an instantaneous TCA is legible


def _offnadir_band(off_deg: float) -> tuple[str, str]:
    a = abs(off_deg)
    for ub, col, lab in _OFFNADIR_BANDS:
        if a <= ub:
            return col, lab
    return _OFFNADIR_BANDS[-1][1], _OFFNADIR_BANDS[-1][2]


def _cloud_cell(pct: float | None):
    """(facecolor, hatch) for a daily cloud-strip cell; None => unknown."""
    if pct is None:
        return _CLOUD_UNKNOWN, "////"
    for ub, col, _ in _CLOUD_BANDS:
        if pct < ub:
            return col, None
    return _CLOUD_BANDS[-1][1], None


def build_timeline_figure(preds: "list[Prediction] | Prediction",
                          tz_name: str = "Australia/Brisbane"):
    """Build the timeline figure (R4). Returns (fig, ax) — ax is the bar axis.

    One row per satellite. Each pass is a single Rectangle centred on its TCA,
    coloured by off-nadir band and labelled with the angle; non-operational
    passes are hatched. When a single AOI carries cloud data, a daily cloud
    strip is drawn above (clear / partly / cloudy / unknown-beyond-horizon).
    Every bar carries gid 'SAT:op|nonop:band' for artist-level tests."""
    # Object-oriented API, not pyplot. `/timeline.png` is a sync def, so FastAPI
    # runs it in its threadpool — and plt.figure()/plt.close() mutate pyplot's
    # global Gcf figure registry, which concurrent requests race on (interleaved
    # or corrupted PNGs, or leaked figures growing without bound). Figure() owns
    # nothing global, so this is thread-safe by construction rather than by luck.
    # [SESSION 2026-07-15]
    import matplotlib.dates as mdates
    from matplotlib.figure import Figure

    if isinstance(preds, Prediction):
        preds = [preds]

    # Lanes come from the active sensor profile (or the union for the combined
    # view), not the Dragonette constant, so the chart draws the right
    # constellation(s) and greys the right non-op sats. [SESSION 2026-07-20]
    key = preds[0].params.get("sensor") if preds else None
    if key == COMBINED_KEY:
        sats_map, operational = combined_roster()
        sats = list(sats_map)
        display = preds[0].params.get("sensor_display", COMBINED_DISPLAY)
    else:
        prof = get_profile(key) if preds else DRAGONETTE
        operational, sats, display = prof.operational, list(prof.satellites), prof.display
    y_of = {name: i for i, name in enumerate(sats)}
    start = min(p.start_utc for p in preds)
    end = max(p.end_utc for p in preds)
    x0, x1 = mdates.date2num(start), mdates.date2num(end)

    cloud_daily = preds[0].cloud_daily if len(preds) == 1 else {}
    strip = bool(cloud_daily)

    fig = Figure(figsize=(11.5, 0.62 * len(sats) + (2.3 if strip else 1.7)))
    if strip:
        gs = fig.add_gridspec(2, 1, height_ratios=[0.6, len(sats)], hspace=0.12)
        cax = fig.add_subplot(gs[0]); ax = fig.add_subplot(gs[1], sharex=cax)
    else:
        ax = fig.add_subplot(111); cax = None

    for pr in preds:
        for op, bucket in ((True, pr.passes), (True, pr.marginal),
                           (False, pr.nonoperational)):
            for p in bucket:
                col, _ = _offnadir_band(p.off_nadir_deg)
                left = mdates.date2num(p.tca_utc) - _BAR_HALF_WIDTH_DAYS
                bar = ax.barh(y_of[p.satellite], 2 * _BAR_HALF_WIDTH_DAYS, left=left,
                              height=0.55, color=col, alpha=0.55 if not op else 0.98,
                              edgecolor="#1b2230", linewidth=0.5,
                              hatch="//" if not op else None)
                band = _offnadir_band(p.off_nadir_deg)[1]
                bar[0].set_gid(f"{p.satellite}:{'nonop' if not op else 'op'}:{band}")
                ax.text(mdates.date2num(p.tca_utc), y_of[p.satellite] - 0.45,
                        f"{abs(p.off_nadir_deg):.1f}", ha="center", va="bottom",
                        fontsize=7.5, color="#33404f")

    ax.set_yticks(range(len(sats)))
    ax.set_yticklabels([f"{s} (non-op)" if not operational.get(s, True) else s
                        for s in sats])
    ax.set_ylim(len(sats) - 0.4, -0.7)           # inverted: DRAG01 at top
    ax.set_xlim(x0, x1)
    ax.xaxis_date()
    ax.xaxis.set_major_locator(mdates.DayLocator(interval=2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
    ax.grid(axis="x", linestyle=":", alpha=0.35)
    ax.set_xlabel("Date (UTC)")

    # ---- cloud strip -----------------------------------------------------
    if strip:
        d = datetime(start.year, start.month, start.day, tzinfo=timezone.utc)
        while mdates.date2num(d) < x1:
            cx = mdates.date2num(d)
            pct = cloud_daily.get(d.strftime("%Y-%m-%d"))
            fc, hatch = _cloud_cell(pct)
            cax.barh(0, 1.0, left=cx, height=1.0, color=fc, hatch=hatch,
                     edgecolor="#cbd5e1", linewidth=0.4)
            d += timedelta(days=1)
        cax.set_xlim(x0, x1); cax.set_ylim(-0.5, 0.5)
        cax.set_yticks([0]); cax.set_yticklabels(["Cloud"], fontsize=8)
        cax.tick_params(labelbottom=False, length=0)
        for s in cax.spines.values():
            s.set_visible(False)
        horizon = start + timedelta(days=TIER1_MAX_DAYS)
        if mdates.date2num(horizon) < x1:
            cax.text((mdates.date2num(horizon) + x1) / 2, 0,
                     "no cloud forecast available this far ahead",
                     ha="center", va="center", fontsize=7.5, color="#5c6b7f")

    # ---- legend ----------------------------------------------------------
    from matplotlib.patches import Patch
    handles = [Patch(facecolor=c, label=lab) for _, c, lab in _OFFNADIR_BANDS]
    nonop_sats = [s for s in sats if not operational.get(s, True)]
    if nonop_sats:
        nonop_lbl = f"non-operational ({', '.join(nonop_sats)})"
        handles.append(Patch(facecolor="#9aa7b4", hatch="//", label=nonop_lbl))
    if strip:
        handles += [Patch(facecolor=_CLOUD_BANDS[0][1], label="clear"),
                    Patch(facecolor=_CLOUD_BANDS[1][1], label="partly cloudy"),
                    Patch(facecolor=_CLOUD_UNKNOWN, hatch="////", label="unknown (beyond horizon)")]
    ncol = 4 if strip else 4
    ax.legend(handles=handles, loc="upper center", bbox_to_anchor=(0.5, -0.14),
              ncol=ncol, frameon=False, fontsize=8.5)

    aois = ", ".join(dict.fromkeys(pr.aoi.name for pr in preds))
    title_ax = cax if strip else ax
    title_ax.set_title(f"{display} imaging opportunities — {aois}\n"
                       f"{start:%Y-%m-%d} to {end:%Y-%m-%d}  ·  number over bar = "
                       "off-nadir angle (°)", fontsize=11)
    if strip:                                    # tight_layout dislikes gridspec+legend
        fig.subplots_adjust(left=0.09, right=0.975, top=0.85, bottom=0.16, hspace=0.14)
    else:
        fig.tight_layout()
    return fig, ax


def render_timeline_png(preds: "list[Prediction] | Prediction",
                        tz_name: str = "Australia/Brisbane",
                        dpi: int = 110) -> bytes:
    """Render the timeline (R4) to PNG bytes; usable in reports and endpoints.

    Thread-safe: the figure is an OO `Figure` with its own Agg canvas and is never
    registered with pyplot, so there is nothing global to close and nothing for
    concurrent requests to race on. [SESSION 2026-07-15]"""
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    fig, _ = build_timeline_figure(preds, tz_name)
    FigureCanvasAgg(fig)                      # attach a canvas; no pyplot involved
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    return buf.getvalue()


def _embed_timeline_sheet(wb, preds, tz_name: str) -> None:
    """Add a 'Timeline' sheet with the rendered PNG embedded (best-effort)."""
    try:
        from openpyxl.drawing.image import Image as XLImage
        png = render_timeline_png(preds, tz_name)
        ws = wb.create_sheet("Timeline")
        img = XLImage(io.BytesIO(png))
        ws.add_image(img, "A1")
    except Exception:
        pass  # a missing plotting backend must never break the spreadsheet


# --------------------------------------------------------------------------
# XLSX report
# --------------------------------------------------------------------------
_PASS_HEADERS = ["Satellite", "Datetime (TCA, UTC)", "Off Nadir Angle (°)",
                 "Sun Angle of Elevation (°)", "Max Off Nadir across AOI (°)",
                 "Slant Range (km)", "TLE Epoch (UTC)"]
# A1/A2/A4 advisory quality columns (never gate access).
QUALITY_COLUMNS = ["Quality", "AOI Coverage %", "Node", "Eff. GSD (m)",
                   "View Zenith (°)", "Sun Glint (°)", "Glint Risk",
                   "Sun Airmass", "TCA ± (s)"]


def _quality_cells(p: Pass) -> list:
    g = p.geometry or {}
    cov = f"{p.coverage_pct * 100:.0f}%" if p.coverage_pct is not None else "n/a"
    return [p.quality or "n/a", cov, (p.node[:3] if p.node else "n/a"),
            g.get("effective_gsd_m", "n/a"), g.get("view_zenith_deg", "n/a"),
            g.get("sun_glint_deg", "n/a"), g.get("glint_risk", "n/a"),
            g.get("airmass_sun", "n/a"),
            p.timing_sigma_s if p.timing_sigma_s is not None else "n/a"]


def _pass_headers(tz_name: str, with_aoi: bool = False,
                  with_cloud: bool = False) -> list[str]:
    """Column headers; `Local (tz)` is inserted after the UTC datetime."""
    h = _PASS_HEADERS[:1] + [_PASS_HEADERS[1], f"Local ({tz_name})"] + _PASS_HEADERS[2:]
    h = h + QUALITY_COLUMNS
    if with_cloud:
        h = h + CLOUD_COLUMNS
    return (["AOI"] + h) if with_aoi else h


def _pass_row(p: Pass, tz, aoi_name: str | None = None,
              with_cloud: bool = False) -> list:
    row = [p.satellite,
           p.tca_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
           p.tca_utc.astimezone(tz).strftime("%Y-%m-%d %H:%M"),
           p.off_nadir_deg, p.sun_elev_deg,
           p.max_off_nadir_aoi_deg, p.slant_range_km,
           p.tle_epoch_utc.strftime("%Y-%m-%dT%H:%M:%SZ")]
    row = row + _quality_cells(p)
    if with_cloud:
        row = row + (p.cloud.xlsx_cells() if p.cloud is not None else ["n/a"] * 6)
    return ([aoi_name] + row) if aoi_name is not None else row


def _cell(ws, row: int, col: int, value, font):
    """Write one cell, never letting untrusted text become a live formula.

    openpyxl infers a leading '=' as a formula, so an AOI name scraped from a
    KMZ (e.g. `=cmd|'/c calc'!A1`) would land in the workbook as an executable
    DDE payload — in a file CLAUDE.md says is circulated to research teams.
    Forcing the cell to text makes Excel render it verbatim instead. Leading
    '+', '-' and '@' need no handling: openpyxl already stores those as text
    [VERIFIED 2026-07-15]. [SESSION 2026-07-15]
    """
    c = ws.cell(row, col, value)
    if isinstance(value, str) and value.startswith("="):
        c.data_type = "s"
    c.font = font
    return c


def _fill_pass_sheet(ws, headers: list[str], rows: list, base, bold) -> None:
    from openpyxl.utils import get_column_letter
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h).font = bold
    for r, vals in enumerate(rows, 2):
        for c, v in enumerate(vals, 1):
            _cell(ws, r, c, v, base)
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 24


def _marginal_note(params: dict) -> str:
    return (f"Passes at {params['max_off_nadir_deg']:.0f}–"
            f"{params['marginal_off_nadir_deg']:.0f}° off-nadir or "
            f"{params['marginal_sun_elev_deg']:.0f}–"
            f"{params['min_sun_elev_deg']:.0f}° sun elevation — "
            "outside Wyvern's standard list; confirm feasibility with Wyvern.")


def _nonop_note() -> str:
    names = [n for n, ok in OPERATIONAL.items() if not ok]
    who = ", ".join(names) if names else "these satellites"
    return (f"{who} NOT yet operational — shown for planning only. NOT a taskable "
            "opportunity and excluded from headline pass counts. Confirm "
            "commissioning with Wyvern before tasking. [SESSION 2026-07-14]")


def _combined_method_rows(pred: Prediction) -> list[tuple[str, str]]:
    """Method sheet for the combined 'all sensors' workbook: one that describes
    every constellation and its own native envelope, rather than a single
    profile. [SESSION 2026-07-20]"""
    aoi = pred.aoi
    sats, _ = combined_roster()

    def envelope(p: SensorProfile) -> str:
        return (f"{p.display}: |off-nadir| ≤ {p.max_off_nadir_deg:g}°"
                + ("" if p.agile else " (FOV half-angle, fixed nadir push-broom)"))

    return [
        ("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")),
        ("Sensor", COMBINED_DISPLAY + " — Dragonette is taskable; Landsat and "
                   "Sentinel-2 are fixed nadir push-brooms imaging on their own cycle"),
        ("AOI polygon", aoi.name),
        ("AOI centroid", f"{abs(aoi.centroid_lat):.5f} {'S' if aoi.centroid_lat < 0 else 'N'}, "
                         f"{abs(aoi.centroid_lon):.5f} {'W' if aoi.centroid_lon < 0 else 'E'}"),
        ("Terrain height assumed (m)", aoi.terrain_alt_m),
        ("Window (UTC)", f"{pred.start_utc:%Y-%m-%d %H:%M} to {pred.end_utc:%Y-%m-%d %H:%M}"),
        ("Satellites (NORAD)", ", ".join(f"{k}={v}" for k, v in sats.items())),
        ("Access filter (per sensor)",
         "  |  ".join(envelope(p) for p in (DRAGONETTE, LANDSAT, SENTINEL2))),
        ("TLE source", "Celestrak GP catalogue, propagated with SGP4"),
        ("Frames", "TEME→ECEF via GMST (IAU 1982) rotation"),
        ("Off-nadir", "Angle at spacecraft between geocentric nadir and LOS to AOI centroid; "
                      "TCA by golden-section to 0.1 s; natural right-of-track sign (Wyvern "
                      "convention, no flip)"),
        ("Sun elevation", "Astronomical Almanac low-precision solar position (~0.01°) at AOI at TCA"),
        ("Caveat", "Geometric access only. Dragonette rows are taskable opportunities; "
                   "Landsat/Sentinel-2 rows are PREDICTED ACQUISITIONS on a fixed cycle, not "
                   "taskable. Cloud cover and operator duty cycle are separate constraints."),
    ]


def _method_rows(pred: Prediction) -> list[tuple[str, str]]:
    aoi = pred.aoi
    # Read the sensor off the prediction, never the Dragonette module constants —
    # a Landsat workbook that lists DRAG01-05 and a 20 deg envelope is worse than
    # no Method sheet at all. [SESSION 2026-07-15]
    if pred.params.get("sensor") == COMBINED_KEY:
        return _combined_method_rows(pred)
    prof = get_profile(pred.params.get("sensor"))
    return [
        ("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")),
        ("Sensor", prof.display + ("" if prof.agile else
                                   "  — fixed nadir push-broom: images systematically "
                                   "on its own cycle; NOT taskable")),
        ("AOI polygon", aoi.name),
        ("AOI centroid", f"{abs(aoi.centroid_lat):.5f} {'S' if aoi.centroid_lat < 0 else 'N'}, "
                         f"{abs(aoi.centroid_lon):.5f} {'W' if aoi.centroid_lon < 0 else 'E'}"),
        ("Terrain height assumed (m)", aoi.terrain_alt_m),
        ("Window (UTC)", f"{pred.start_utc:%Y-%m-%d %H:%M} to {pred.end_utc:%Y-%m-%d %H:%M}"),
        ("Satellites (NORAD)", ", ".join(f"{k}={v}" for k, v in prof.satellites.items())),
        ("Swath / nadir GSD", f"{prof.swath_km:g} km / {prof.gsd_m:g} m"),
        ("TLE source", "Celestrak GP catalogue, propagated with SGP4"),
        ("Frames", "TEME→ECEF via GMST (IAU 1982) rotation"),
        ("Off-nadir", "Angle at spacecraft between geocentric nadir and LOS to AOI centroid; "
                      "TCA by golden-section refinement to 0.1 s; sign is the natural "
                      "right-of-track sense (LOS vs orbit normal r×v), which is Wyvern's "
                      "convention — no flip applied (validated vs Wyvern sheet, Jul 2026)"),
        ("Sun elevation", "Astronomical Almanac low-precision solar position (~0.01°) at "
                          "AOI at TCA, geometric, no refraction"),
        ("Access filter", f"|off-nadir| ≤ {pred.params['max_off_nadir_deg']}°, "
                          f"sun ≥ {pred.params['min_sun_elev_deg']}°"
                          + ("" if prof.agile else
                             "  — the off-nadir limit is this sensor's FOV half-angle "
                             "(a hard optical limit, not a tasking preference); the sun "
                             "floor is daylight, since it images on its own schedule")),
        ("Caveat", ("Geometric access only — Wyvern tasking/scheduling and cloud cover "
                    "are separate constraints.") if prof.agile else
                   ("Geometric access only. These are PREDICTED ACQUISITIONS, not "
                    "tasking opportunities: this sensor images whatever falls in its "
                    "swath on a fixed cycle, and cannot be pointed at your AOI on "
                    "request. Whether a scene is actually delivered/usable depends on "
                    "the operator's own duty cycle and cloud.")),
    ]


def _write_method_sheet(ws, rows: list[tuple[str, str]], base, bold) -> None:
    for r, (k, v) in enumerate(rows, 1):
        ws.cell(r, 1, k).font = bold
        _cell(ws, r, 2, str(v), base)      # row values include the AOI name
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110


def _has_cloud(*buckets: list[Pass]) -> bool:
    return any(p.cloud is not None for b in buckets for p in b)


def write_xlsx(pred: Prediction, tz_name: str = "Australia/Brisbane") -> bytes:
    from zoneinfo import ZoneInfo
    from openpyxl import Workbook
    from openpyxl.styles import Font

    tz = ZoneInfo(tz_name)
    wb = Workbook()
    base = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=10, bold=True)
    wc = _has_cloud(pred.passes, pred.marginal, pred.nonoperational)
    headers = _pass_headers(tz_name, with_cloud=wc)

    ws = wb.active; ws.title = "Passes"
    _fill_pass_sheet(ws, headers,
                     [_pass_row(p, tz, with_cloud=wc) for p in pred.passes], base, bold)

    # Per-satellite summary as formulas so it stays live if rows are edited.
    srow = len(pred.passes) + 4
    ws.cell(srow, 1, "Passes per satellite").font = bold
    sat_names = sorted(SATELLITES)
    for i, name in enumerate(sat_names, 1):
        ws.cell(srow + i, 1, name).font = base
        ws.cell(srow + i, 2, f'=COUNTIF(A2:A{max(len(pred.passes) + 1, 2)},"{name}")').font = base
    ws.cell(srow + len(sat_names) + 1, 1, "Total").font = bold
    ws.cell(srow + len(sat_names) + 1, 2,
            f"=SUM(B{srow + 1}:B{srow + len(sat_names)})").font = bold

    ws2 = wb.create_sheet("Marginal - stretch")
    _fill_pass_sheet(ws2, headers,
                     [_pass_row(p, tz, with_cloud=wc) for p in pred.marginal], base, bold)
    ws2.cell(len(pred.marginal) + 3, 1, _marginal_note(pred.params)).font = base

    if pred.nonoperational:                  # R5: separate, badged, uncounted
        wsn = wb.create_sheet("Non-operational")
        _fill_pass_sheet(wsn, headers,
                         [_pass_row(p, tz, with_cloud=wc) for p in pred.nonoperational],
                         base, bold)
        wsn.cell(len(pred.nonoperational) + 3, 1, _nonop_note()).font = bold

    ws3 = wb.create_sheet("Method")
    _write_method_sheet(ws3, _method_rows(pred)
                        + [("Warning", w) for w in pred.warnings], base, bold)

    _embed_timeline_sheet(wb, pred, tz_name)    # R4

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_xlsx_multi(preds: list[Prediction],
                     tz_name: str = "Australia/Brisbane") -> bytes:
    """Combined workbook for several AOIs/polygons (R7 --all-polygons, R8).

    One 'Passes' and one 'Marginal' sheet with a leading AOI column, plus a
    Method sheet summarising every AOI. A single prediction is delegated to
    write_xlsx so the validated single-AOI format is never disturbed."""
    if len(preds) == 1:
        return write_xlsx(preds[0], tz_name)

    from zoneinfo import ZoneInfo
    from openpyxl import Workbook
    from openpyxl.styles import Font

    tz = ZoneInfo(tz_name)
    wb = Workbook()
    base = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=10, bold=True)
    wc = any(_has_cloud(pr.passes, pr.marginal, pr.nonoperational) for pr in preds)
    headers = _pass_headers(tz_name, with_aoi=True, with_cloud=wc)

    std_rows, marg_rows, nonop_rows = [], [], []
    for pr in preds:
        std_rows += [_pass_row(p, tz, pr.aoi.name, wc) for p in pr.passes]
        marg_rows += [_pass_row(p, tz, pr.aoi.name, wc) for p in pr.marginal]
        nonop_rows += [_pass_row(p, tz, pr.aoi.name, wc) for p in pr.nonoperational]

    ws = wb.active; ws.title = "Passes"
    _fill_pass_sheet(ws, headers, std_rows, base, bold)
    # Per-AOI totals as live formulas over the AOI column (A).
    srow = len(std_rows) + 4
    ws.cell(srow, 1, "Passes per AOI").font = bold
    for i, pr in enumerate(preds, 1):
        _cell(ws, srow + i, 1, pr.aoi.name, base)
        # Criterion is a cell reference, not an interpolated literal: an AOI name
        # containing a double quote (e.g. Paddock "North") would otherwise close
        # the string argument and emit a malformed formula, which Excel reports
        # as unreadable content and repairs by discarding the sheet.
        # [SESSION 2026-07-15]
        _cell(ws, srow + i, 2,
              f"=COUNTIF(A2:A{max(len(std_rows) + 1, 2)},A{srow + i})", base)
    ws.cell(srow + len(preds) + 1, 1, "Total").font = bold
    ws.cell(srow + len(preds) + 1, 2,
            f"=SUM(B{srow + 1}:B{srow + len(preds)})").font = bold

    ws2 = wb.create_sheet("Marginal - stretch")
    _fill_pass_sheet(ws2, headers, marg_rows, base, bold)
    ws2.cell(len(marg_rows) + 3, 1, _marginal_note(preds[0].params)).font = base

    if nonop_rows:                           # R5: separate, badged, uncounted
        wsn = wb.create_sheet("Non-operational")
        _fill_pass_sheet(wsn, headers, nonop_rows, base, bold)
        wsn.cell(len(nonop_rows) + 3, 1, _nonop_note()).font = bold

    ws3 = wb.create_sheet("Method")
    rows: list[tuple[str, str]] = [
        ("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")),
        ("AOIs", ", ".join(pr.aoi.name for pr in preds)),
    ]
    for pr in preds:
        rows.append((f"AOI · {pr.aoi.name}",
                     f"centroid {abs(pr.aoi.centroid_lat):.5f}"
                     f"{'S' if pr.aoi.centroid_lat < 0 else 'N'}, "
                     f"{abs(pr.aoi.centroid_lon):.5f}"
                     f"{'W' if pr.aoi.centroid_lon < 0 else 'E'}; "
                     f"{len(pr.passes)} standard, {len(pr.marginal)} marginal"))
    rows += _method_rows(preds[0])[4:]  # shared method notes (skip per-AOI header rows)
    seen_warn: set[str] = set()
    for pr in preds:
        for w in pr.warnings:
            if w not in seen_warn:
                seen_warn.add(w); rows.append(("Warning", w))
    _write_method_sheet(ws3, rows, base, bold)

    _embed_timeline_sheet(wb, preds, tz_name)   # R4 + R8 campaign timeline

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
