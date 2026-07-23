"""Tests runnable fully offline: synthetic TLEs via the sgp4 exporter.

Live-TLE behaviour (Celestrak fetch, cache, staleness warning) is exercised
with a fake http_get; real-network correctness must be checked on first
local run against Wyvern's own sheet (see README, 'First-run validation').
"""
import io
import math
import sys
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path

import numpy as np
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import passes as P  # noqa: E402
from sgp4.api import Satrec, WGS72, jday  # noqa: E402
from sgp4 import exporter  # noqa: E402

SITEA = (-20.0, 150.0)


# ---------------------------------------------------------------- helpers
def make_kmz(lat: float, lon: float, half_deg: float = 0.045,
             name: str = "TEST_AOI") -> bytes:
    coords = " ".join(
        f"{lon + dx},{lat + dy},0"
        for dx, dy in [(-half_deg, -half_deg), (half_deg, -half_deg),
                       (half_deg, half_deg), (-half_deg, half_deg),
                       (-half_deg, -half_deg)])
    kml = f"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2"><Document><Placemark>
<name>{name}</name><Polygon><outerBoundaryIs><LinearRing>
<coordinates>{coords}</coordinates>
</LinearRing></outerBoundaryIs></Polygon></Placemark></Document></kml>"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("doc.kml", kml)
    return buf.getvalue()


def synth_tle(satnum: int, epoch_dt: datetime, raan_deg: float,
              m_deg: float, incl_deg: float = 97.44,
              mm_rev_day: float = 15.2) -> P.TLE:
    """Sun-synchronous-ish LEO built with sgp4init and exported as a TLE."""
    jd, fr = jday(epoch_dt.year, epoch_dt.month, epoch_dt.day,
                  epoch_dt.hour, epoch_dt.minute, epoch_dt.second)
    epoch_1949 = jd + fr - 2433281.5
    sat = Satrec()
    sat.sgp4init(
        WGS72, "i", satnum, epoch_1949,
        2.0e-5,            # bstar
        0.0, 0.0,          # ndot, nddot (ignored by SGP4)
        0.0008,            # ecco
        90.0 * P.DEG,      # argpo
        incl_deg * P.DEG,  # inclo
        m_deg * P.DEG,     # mo
        mm_rev_day * 2.0 * math.pi / 1440.0,  # no_kozai rad/min
        raan_deg * P.DEG,  # nodeo
    )
    l1, l2 = exporter.export_tle(sat)
    return P.TLE(f"SYN{satnum}", satnum, l1, l2,
                 epoch_dt.isoformat(timespec="seconds"))


def synth_constellation(start: datetime) -> dict[str, P.TLE]:
    epoch = start - timedelta(days=1)
    return {name: synth_tle(99000 + i, epoch, raan_deg=40.0 * i,
                            m_deg=72.0 * i)
            for i, name in enumerate(P.SATELLITES)}


# ---------------------------------------------------------------- unit: math
def test_gmst_j2000():
    jd, fr = jday(2000, 1, 1, 12, 0, 0.0)
    g = math.degrees(float(P.gmst_rad(jd + fr)))
    assert abs(g - 280.4606) < 0.01


def test_ecef_known_points():
    r = P.geodetic_to_ecef(0.0, 0.0, 0.0)
    assert np.allclose(r, [6378.137, 0, 0], atol=1e-6)
    r = P.geodetic_to_ecef(90.0, 0.0, 0.0)
    assert abs(r[2] - 6356.7523) < 1e-3 and abs(r[0]) < 1e-6


def test_teme_ecef_rotation_preserves_norm_and_inverts():
    v = np.array([3000.0, -5000.0, 2000.0])
    th = 1.234
    w = P.ecef_to_teme(v, th)
    assert abs(np.linalg.norm(w) - np.linalg.norm(v)) < 1e-9
    # inverse rotation (theta -> -theta pattern via manual R3)
    c, s = math.cos(th), math.sin(th)
    back = np.array([c * w[0] + s * w[1], -s * w[0] + c * w[1], w[2]])
    assert np.allclose(back, v, atol=1e-9)


def test_solar_equinox_noon_and_night():
    el = P.sun_elevation_deg(datetime(2026, 3, 20, 12, 0, tzinfo=timezone.utc), 0.0, 0.0)
    assert el > 85.0
    el_night = P.sun_elevation_deg(datetime(2026, 3, 20, 0, 0, tzinfo=timezone.utc), 0.0, 0.0)
    assert el_night < -80.0


def test_solar_palmerston_north_winter_noon():
    # 2026-07-14 ~solar noon local (~00:20 UTC), lat -40.35: expect ~26-28 deg
    el = P.sun_elevation_deg(datetime(2026, 7, 14, 0, 20, tzinfo=timezone.utc),
                             -40.35, 175.61)
    assert 22.0 < el < 32.0


def test_sun_azimuth_hemisphere_sense_at_solar_noon():
    """At solar noon the sun is due south seen from the northern hemisphere and
    due north from the southern. Pins the azimuth reference direction.

    The previous model computed azimuth from South but
    returned it as if measured from North, reflecting every value about the N–S
    axis: it put the noon sun at 3.1° (due N) for lat +40 and 175.6° (due S) for
    Site A — both exactly backwards.
    """
    noon = datetime(2026, 3, 20, 12, 0, tzinfo=timezone.utc)   # tst ~11.86 h
    _, az_n, tst_n = P.sun_position_deg(noon, 40.0, 0.0)
    _, az_s, _ = P.sun_position_deg(noon, -20.0, 0.0)
    assert 11.5 < tst_n < 12.0, "fixture instant should sit just before solar noon"
    assert abs(az_n - 180.0) < 5.0, f"N-hemisphere noon sun should be due S, got {az_n}"
    assert min(az_s, 360.0 - az_s) < 8.0, \
        f"S-hemisphere noon sun should be due N, got {az_s}"


def test_sun_azimuth_am_pm_branch_at_high_longitude():
    """The hour angle must be wrapped to ±180° before the AM/PM sense is taken.

    Site A sits at lon 151.4°, so true solar time and UTC
    diverge by ~10 h and the unwrapped hour angle exceeded +180°, inverting the
    branch: 07:00 solar time (morning, sun in the east) was reported at 244°
    (west). Any AOI with |lon| >~ 45° could trigger this.
    """
    lat, lon = -20.0, 150.0
    _, az_am, tst_am = P.sun_position_deg(
        datetime(2026, 7, 15, 21, 0, tzinfo=timezone.utc), lat, lon)
    _, az_pm, tst_pm = P.sun_position_deg(
        datetime(2026, 7, 15, 3, 0, tzinfo=timezone.utc), lat, lon)
    assert tst_am < 12.0 < tst_pm, f"instants must straddle solar noon, got {tst_am}/{tst_pm}"
    assert 0.0 < az_am < 180.0, f"morning sun must lie east of the meridian, got {az_am}"
    assert 180.0 < az_pm < 360.0, f"afternoon sun must lie west of the meridian, got {az_pm}"


# ---------------------------------------------------------------- unit: KMZ
def test_kmz_parse_centroid():
    aoi = P.parse_kmz(make_kmz(*SITEA), terrain_alt_m=400.0)
    assert aoi.name == "TEST_AOI"
    assert abs(aoi.centroid_lat - SITEA[0]) < 1e-6
    assert abs(aoi.centroid_lon - SITEA[1]) < 1e-6
    assert len(aoi.vertices_lonlat) == 4


def test_kmz_bad_input():
    with pytest.raises(ValueError):
        P.parse_kmz(b"<kml>no polygon here</kml>")


# ---------------------------------------------------------------- e2e
START = datetime(2026, 7, 14, 0, 0, tzinfo=timezone.utc)


def _predict(**kw):
    tles = synth_constellation(START)
    defaults = dict(days=14.0, start_utc=START, terrain_alt_m=400.0,
                    min_sun_elev_deg=-90.0, marginal_sun_elev_deg=-90.0,
                    tles=tles)
    defaults.update(kw)
    return P.predict(make_kmz(*SITEA), **defaults)


def test_e2e_finds_and_refines_passes():
    pred = _predict()
    assert len(pred.passes) >= 3, "synthetic constellation should yield passes"
    for p in pred.passes:
        assert abs(p.off_nadir_deg) <= 20.05
        assert p.max_off_nadir_aoi_deg >= abs(p.off_nadir_deg) - 0.05
        assert 400 < p.slant_range_km < 3000
        assert pred.start_utc <= p.tca_utc <= pred.end_utc
    # chronological
    tcas = [p.tca_utc for p in pred.passes]
    assert tcas == sorted(tcas)


def test_refinement_beats_coarse_grid():
    """eta at refined TCA must be <= eta at every coarse sample near it."""
    tles = synth_constellation(START)
    pred = _predict()
    p = pred.passes[0]
    sat = Satrec.twoline2rv(tles[p.satellite].line1, tles[p.satellite].line2)
    site = pred.aoi.centroid_ecef
    jd0, fr0 = P.dt_to_jd(START)
    t_star = (p.tca_utc - START).total_seconds()
    eta_star = P._eta_at(sat, jd0, fr0, t_star, site)
    for dt_s in (-20, -10, 10, 20):
        assert eta_star <= P._eta_at(sat, jd0, fr0, t_star + dt_s, site) + 1e-6


def test_sign_flips_across_ground_track():
    """Same passes, AOI shifted east vs west: signed off-nadir must differ in sign
    for the matching pass when the shift straddles the ground track."""
    tles = synth_constellation(START)
    east = P.predict(make_kmz(SITEA[0], SITEA[1] + 1.2), days=14.0,
                     start_utc=START, min_sun_elev_deg=-90.0,
                     marginal_sun_elev_deg=-90.0, tles=tles,
                     max_off_nadir_deg=60.0, marginal_off_nadir_deg=60.0)
    west = P.predict(make_kmz(SITEA[0], SITEA[1] - 1.2), days=14.0,
                     start_utc=START, min_sun_elev_deg=-90.0,
                     marginal_sun_elev_deg=-90.0, tles=tles,
                     max_off_nadir_deg=60.0, marginal_off_nadir_deg=60.0)
    flipped = 0
    for pe in east.passes:
        for pw in west.passes:
            if (pe.satellite == pw.satellite
                    and abs((pe.tca_utc - pw.tca_utc).total_seconds()) < 120):
                if pe.off_nadir_deg * pw.off_nadir_deg < 0:
                    flipped += 1
    assert flipped >= 1, "expected at least one matched pass with opposite sign"


def test_sun_filter_and_marginal_band():
    all_p = _predict()  # sun filter off
    lit = _predict(min_sun_elev_deg=20.0, marginal_sun_elev_deg=15.0)
    assert len(lit.passes) <= len(all_p.passes)
    for p in lit.passes:
        assert p.sun_elev_deg >= 20.0
    for p in lit.marginal:
        assert (20.0 < abs(p.off_nadir_deg) <= 30.0) or (15.0 <= p.sun_elev_deg < 20.0)


def test_tle_age_warning():
    tles = {k: synth_tle(99100 + i, START - timedelta(days=9),
                         raan_deg=40 * i, m_deg=72 * i)
            for i, k in enumerate(P.SATELLITES)}
    pred = P.predict(make_kmz(*SITEA), days=3.0, start_utc=START,
                     min_sun_elev_deg=-90, marginal_sun_elev_deg=-90, tles=tles)
    assert any("days old" in w for w in pred.warnings)


# ---------------------------------------------------------------- xlsx
def test_xlsx_output(tmp_path):
    from openpyxl import load_workbook
    pred = _predict()
    blob = P.write_xlsx(pred, tz_name="Australia/Brisbane")
    f = tmp_path / "out.xlsx"
    f.write_bytes(blob)
    wb = load_workbook(f)
    # R5 may add a 'Non-operational' sheet (synthetic constellation includes DRAG05).
    assert {"Passes", "Marginal - stretch", "Method"} <= set(wb.sheetnames)
    ws = wb["Passes"]
    assert ws.cell(1, 1).value == "Satellite"
    assert ws.max_row >= len(pred.passes) + 1
    # summary formulas present
    found = any(str(c.value).startswith("=COUNTIF")
                for row in ws.iter_rows() for c in row if c.value)
    assert found


# ---------------------------------------------------------------- TLE fetch
def test_fetch_with_fake_http_and_cache(tmp_path):
    tles = synth_constellation(START)
    served = {c: f"{n}\n{t.line1}\n{t.line2}\n"
              for n, t in tles.items() for c in [t.catnr]}
    calls = []

    def fake_get(url):
        catnr = int(url.split("CATNR=")[1].split("&")[0])
        calls.append(catnr)
        return served[catnr]

    sats = {n: t.catnr for n, t in tles.items()}
    cache = tmp_path / "tle_cache.json"
    got, warn = P.fetch_tles(sats, cache_path=cache, http_get=fake_get)
    assert len(got) == 5 and not warn and len(calls) == 5
    # second call served from cache — no HTTP
    got2, _ = P.fetch_tles(sats, cache_path=cache, http_get=fake_get)
    assert len(calls) == 5
    assert got2["DRAG01"].line1 == got["DRAG01"].line1


def test_fresh_cache_for_one_constellation_is_not_served_for_another(tmp_path):
    """Regression: the TLE cache is a single file shared across sensor profiles.
    A fresh cache full of Dragonette must NOT be served (nor crash with "Cache
    missing") when a different constellation — Landsat / Sentinel-2 — is
    requested; fetch_tles must fetch the roster actually asked for. Found by
    switching sensors in the SPA."""
    tles = synth_constellation(START)
    cache = tmp_path / "tle_cache.json"
    P._save_cache(cache, tles)                     # a FRESH Dragonette-only cache
    assert P._cache_covers(P._load_cache(cache), {n: t.catnr for n, t in tles.items()})

    ref = next(iter(tles.values()))                # reuse valid TLE lines
    served = {70001: f"SATX\n{ref.line1}\n{ref.line2}\n",
              70002: f"SATY\n{ref.line1}\n{ref.line2}\n"}
    calls = []

    def fake_get(url):
        catnr = int(url.split("CATNR=")[1].split("&")[0]); calls.append(catnr)
        return served[catnr]

    other = {"SATX": 70001, "SATY": 70002}
    got, warn = P.fetch_tles(other, cache_path=cache, http_get=fake_get)
    assert set(got) == {"SATX", "SATY"}            # returned the requested roster
    assert sorted(calls) == [70001, 70002]         # fetched, not served from the DRAG cache


def test_fetch_falls_back_to_stale_cache(tmp_path):
    tles = synth_constellation(START)
    sats = {n: t.catnr for n, t in tles.items()}
    cache = tmp_path / "c.json"
    P._save_cache(cache, tles)
    import json
    blob = json.loads(cache.read_text()); blob["_ts"] = 0.0
    cache.write_text(json.dumps(blob))

    def dead(url):  # network down
        raise ConnectionError("no route")

    got, warn = P.fetch_tles(sats, cache_path=cache, http_get=dead)
    assert len(got) == 5 and any("cached TLEs" in w for w in warn)


# ---------------------------------------------------------------- API
def test_api_endpoints(monkeypatch, tmp_path):
    from fastapi.testclient import TestClient
    import app as A

    tles = synth_constellation(START)
    monkeypatch.setattr(P, "fetch_tles", lambda *a, **k: (tles, []))
    client = TestClient(A.app)

    assert client.get("/health").status_code == 200
    assert "satellites" in client.get("/tle-status").json()

    kmz = make_kmz(*SITEA)
    form = {"days": "14", "alt": "400", "tz": "Australia/Brisbane",
            "max_off_nadir": "20", "min_sun": "-90"}
    r = client.post("/predict/json", files={"kmz": ("t.kmz", kmz)}, data=form)
    assert r.status_code == 200
    body = r.json()
    assert len(body["passes"]) >= 3
    assert abs(body["aoi"]["centroid_lat"] - SITEA[0]) < 1e-4

    r = client.post("/predict", files={"kmz": ("t.kmz", kmz)}, data=form)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxml")
    (tmp_path / "api.xlsx").write_bytes(r.content)
    from openpyxl import load_workbook
    assert "Passes" in load_workbook(tmp_path / "api.xlsx").sheetnames

    # bad input -> 422
    r = client.post("/predict/json", files={"kmz": ("t.kmz", b"not a kmz <kml/>")},
                    data=form)
    assert r.status_code == 422
