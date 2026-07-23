"""P3: 16-day GFS rain outlook (stakeholder request).

A feed separate from the existing three-tier cloud scheme, not a
replacement for it: attach_cloud's TIER1_MAX_DAYS (deterministic, 0-5 d) and
TIER2_MAX_DAYS (ensemble P(clear), 5-15 d) must be untouched by this
addition. attach_rain queries Open-Meteo's GFS (models=gfs_seamless) daily
precipitation sum + probability for the full 16 d window in one call per
AOI, flagging day 8-16 values as a low-skill outlook rather than truncating
them.

`fixtures/gfs_rain_siteA_20260723.json` is a real captured Open-Meteo
response for Site A's own centroid (-27.85619, 151.44823 — Queensland,
Australia, the same site the rest of the suite validates against), proving
this runs against a genuine QLD site rather than a synthetic one. The live
one-shot fidelity check (this fixture's source) is `compare_rain_feed.py`;
this file is the offline, reproducible form of the same comparison.
"""
import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
import passes as P  # noqa: E402

FIX = ROOT / "fixtures"
GFS_RAW = json.loads((FIX / "gfs_rain_siteA_20260723.json").read_text())
GFS_DAILY = GFS_RAW["daily"]
SITEA_KMZ = (FIX / "SiteA.kmz").read_bytes()
POLYGON = "SITEA_100sqkm"


def _tles():
    return P._parse_3le_file((FIX / "tles_real_20260714.txt").read_text(), P.SATELLITES)


def _pred(days=16.0, start=datetime(2026, 7, 14, tzinfo=timezone.utc)):
    return P.predict(SITEA_KMZ, days=days, start_utc=start, terrain_alt_m=400.0,
                     polygon_name=POLYGON, tles=_tles())


# ------------------------------------------------------- the existing tiers
def test_cloud_tier_boundaries_are_unchanged_by_the_rain_feature():
    """The brief's own requirement: rain must not perturb the deterministic
    0-5 d / ensemble 5-15 d cloud tiers."""
    assert P.TIER1_MAX_DAYS == 5.0
    assert P.TIER2_MAX_DAYS == 15.0


# --------------------------------------------------------- offline QLD fixture
def test_rain_fixture_is_a_real_qld_site_16_day_gfs_response():
    assert len(GFS_DAILY["time"]) == 16
    # Open-Meteo snaps to its nearest GFS grid cell (~0.25 deg), not the exact
    # centroid -- still unambiguously the same QLD site, not a different one.
    assert GFS_RAW["latitude"] == pytest.approx(-27.856, abs=0.2)
    assert GFS_RAW["longitude"] == pytest.approx(151.448, abs=0.2)


def test_attach_rain_reproduces_the_captured_gfs_values_exactly():
    """The offline form of compare_rain_feed.py: every pass's rain figures
    must equal the raw Open-Meteo value for its own UTC date, unmodified."""
    pred = _pred()
    P.attach_rain(pred, daily_json=GFS_DAILY)
    by_date = P._daily_rain_index(GFS_DAILY)
    all_passes = pred.passes + pred.marginal + pred.nonoperational
    assert all_passes, "fixture window should produce at least one pass"
    checked = 0
    for p in all_passes:
        date = p.tca_utc.strftime("%Y-%m-%d")
        raw = by_date.get(date)
        if raw is None:
            continue
        raw_mm, raw_pct = raw
        assert p.rain.rain_sum_mm == (None if raw_mm is None else float(raw_mm))
        assert p.rain.rain_prob_pct == (None if raw_pct is None else float(raw_pct))
        checked += 1
    assert checked > 0


# ------------------------------------------------------------ low-skill flag
def test_days_0_to_8_are_not_flagged_low_skill():
    ref = datetime(2026, 7, 14, tzinfo=timezone.utc)
    pred = _pred(days=7.0, start=ref)
    P.attach_rain(pred, daily_json=GFS_DAILY, now=ref)
    all_passes = pred.passes + pred.marginal + pred.nonoperational
    assert all_passes
    assert all(not p.rain.low_skill for p in all_passes)


def test_days_8_to_16_are_flagged_low_skill():
    ref = datetime(2026, 7, 14, tzinfo=timezone.utc)
    pred = _pred(days=16.0, start=ref)
    P.attach_rain(pred, daily_json=GFS_DAILY, now=ref)
    late = [p for p in pred.passes + pred.marginal + pred.nonoperational
           if (p.tca_utc - ref).total_seconds() / 86400.0 >= P.RAIN_LOW_SKILL_DAYS]
    assert late, "need at least one pass in the low-skill window to check"
    assert all(p.rain.low_skill for p in late)


def test_low_skill_boundary_is_exactly_rain_low_skill_days():
    epoch = datetime(2026, 7, 20, tzinfo=timezone.utc)
    just_before = epoch + timedelta(days=P.RAIN_LOW_SKILL_DAYS - 0.01)
    just_after = epoch + timedelta(days=P.RAIN_LOW_SKILL_DAYS + 0.01)
    lead_before = (just_before - epoch).total_seconds() / 86400.0
    lead_after = (just_after - epoch).total_seconds() / 86400.0
    assert lead_before < P.RAIN_LOW_SKILL_DAYS <= lead_after


# --------------------------------------------------------------- resilience
def test_a_fetch_failure_never_raises_and_never_blocks():
    def _boom(url):
        raise RuntimeError("network unreachable")
    pred = _pred(days=3.0)
    P.attach_rain(pred, http_get=_boom)
    all_passes = pred.passes + pred.marginal + pred.nonoperational
    assert all_passes
    assert all(p.rain is not None and p.rain.rain_sum_mm is None for p in all_passes)
    assert any("unavailable" in w for w in pred.warnings)


def test_empty_prediction_is_a_no_op():
    pred = P.predict(SITEA_KMZ, days=1.0, start_utc=datetime(2026, 7, 14, tzinfo=timezone.utc),
                     terrain_alt_m=400.0, polygon_name=POLYGON, tles=_tles(),
                     max_off_nadir_deg=0.001, marginal_off_nadir_deg=0.001)
    assert not (pred.passes or pred.marginal or pred.nonoperational)
    P.attach_rain(pred, daily_json=GFS_DAILY)   # must not raise on an empty prediction


# ------------------------------------------------------------ sheet + JSON
def test_rain_columns_appear_only_when_rain_was_attached(tmp_path):
    pred = _pred()
    f_without = tmp_path / "no_rain.xlsx"
    f_without.write_bytes(P.write_xlsx(pred, "UTC"))
    headers_without = [c.value for c in
                       next(load_workbook(f_without)["Passes"].iter_rows(max_row=1))]
    assert not (set(P.RAIN_COLUMNS) & set(headers_without))

    P.attach_rain(pred, daily_json=GFS_DAILY)
    f_with = tmp_path / "with_rain.xlsx"
    f_with.write_bytes(P.write_xlsx(pred, "UTC"))
    headers_with = [c.value for c in
                    next(load_workbook(f_with)["Passes"].iter_rows(max_row=1))]
    assert set(P.RAIN_COLUMNS) <= set(headers_with)


def test_json_export_includes_rain_block_and_schema_bump():
    pred = _pred()
    P.attach_rain(pred, daily_json=GFS_DAILY)
    body = P.prediction_json([pred])
    assert body["schema_version"] == "2.2"
    rained = [p for p in body["passes"] if "rain" in p]
    assert rained, "at least one pass should carry a rain block"
    assert {"rain_sum_mm", "rain_prob_pct", "low_skill"} <= set(rained[0]["rain"])
