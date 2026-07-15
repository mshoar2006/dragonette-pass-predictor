"""Offline regression test against the committed real-TLE baseline.

This is the executable form of the VALIDATION.md 2026-07-14 sign resolution.
It replays `fixtures/tles_real_20260714.txt` over the Site A 100 km² polygon and
compares every row of `fixtures/regression_baseline_siteA_*.xlsx`.

Why it exists [SESSION 2026-07-15]: CLAUDE.md hard constraint 1
(`SIGN_FLIP_TO_MATCH_WYVERN = False`) had no test. `test_sign_flips_across_
ground_track` only asserts east/west passes carry *opposite* signs, which is
invariant under global negation — flipping the constant passed the whole suite.
The assertion that pins it is `off_nadir_deg == approx(ref)` on the **signed**
value: a flip moves every row by 2×, far outside tolerance.

Fully offline and date-stable: real TLEs are read from a fixture, never fetched,
and the window is pinned to the baseline's epoch.
"""
from datetime import datetime, timezone
from pathlib import Path
import sys

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
import passes as P  # noqa: E402

FIX = ROOT / "fixtures"
BASELINE = FIX / "regression_baseline_siteA_2026-07-14_to_2026-07-28.xlsx"
REAL_TLES = FIX / "tles_real_20260714.txt"
START = datetime(2026, 7, 14, 0, 0, tzinfo=timezone.utc)
TERRAIN_M = 400.0
POLYGON = "SITEA_100sqkm"

# Baseline sheet layout: column A carries a satellite-identification preamble
# (rows 1-6); the pass table lives in columns G..N with its header on row 3.
# Indices are asserted against the header text in `_baseline_rows` so a column
# insert fails loudly rather than silently validating the wrong field.
_HDR_ROW = 3
_COL = dict(sat=6, tca=7, local=8, off=9, sun=10, maxoff=11, slant=12, epoch=13)


def _baseline_rows():
    ws = load_workbook(BASELINE)["Pass Predictions"]
    hdr = [c.value for c in ws[_HDR_ROW]]
    assert str(hdr[_COL["sat"]]).startswith("Satellite")
    assert str(hdr[_COL["tca"]]).startswith("Datetime (TCA")
    assert str(hdr[_COL["off"]]).startswith("Off Nadir Angle")
    assert str(hdr[_COL["sun"]]).startswith("Sun Angle of Elevation")
    assert str(hdr[_COL["maxoff"]]).startswith("Max Off Nadir across AOI")
    assert str(hdr[_COL["slant"]]).startswith("Slant Range")
    out = []
    for r in ws.iter_rows(min_row=_HDR_ROW + 1, values_only=True):
        sat = r[_COL["sat"]]
        if not sat or not str(sat).startswith("DRAG"):
            continue
        out.append(dict(
            sat=str(sat),
            tca=datetime.fromisoformat(str(r[_COL["tca"]]).replace("Z", "+00:00")),
            off=float(r[_COL["off"]]),
            sun=float(r[_COL["sun"]]),
            maxoff=float(r[_COL["maxoff"]]),
            slant=float(r[_COL["slant"]]),
        ))
    assert out, "baseline workbook yielded no DRAG rows"
    return out


@pytest.fixture(scope="module")
def prediction():
    tles = P._parse_3le_file(REAL_TLES.read_text(), P.SATELLITES)
    return P.predict((FIX / "SiteA.kmz").read_bytes(), days=14.0,
                     start_utc=START, terrain_alt_m=TERRAIN_M,
                     polygon_name=POLYGON, tles=tles)


def _all_passes(pred):
    # DRAG05 is in the baseline and is routed to `nonoperational` (R5).
    return pred.passes + pred.marginal + pred.nonoperational


def _match(pred, row, tol_s=600.0):
    cands = [p for p in _all_passes(pred)
             if p.satellite == row["sat"]
             and abs((p.tca_utc - row["tca"]).total_seconds()) < tol_s]
    if not cands:
        return None
    return min(cands, key=lambda p: abs((p.tca_utc - row["tca"]).total_seconds()))


def test_sign_convention_constant_is_false():
    """CLAUDE.md hard constraint 1. Pinned explicitly, mirroring the DRAG05
    OPERATIONAL guard, so the intent is greppable and not merely implied."""
    assert P.SIGN_FLIP_TO_MATCH_WYVERN is False


def test_every_baseline_pass_is_reproduced(prediction):
    missing = [r["sat"] + " @ " + r["tca"].isoformat()
               for r in _baseline_rows() if _match(prediction, r) is None]
    assert not missing, f"baseline passes not reproduced: {missing}"


def test_signed_off_nadir_matches_baseline(prediction):
    """Pins CLAUDE.md constraint 1. The value is compared **signed** — this is
    the assertion that fails if SIGN_FLIP_TO_MATCH_WYVERN is ever flipped."""
    for row in _baseline_rows():
        p = _match(prediction, row)
        assert p is not None, f"lost {row['sat']} @ {row['tca']}"
        assert p.off_nadir_deg == pytest.approx(row["off"], abs=0.25), (
            f"{row['sat']} @ {row['tca']}: signed off-nadir "
            f"{p.off_nadir_deg} vs baseline {row['off']}")


def test_tca_matches_baseline(prediction):
    for row in _baseline_rows():
        p = _match(prediction, row)
        d = abs((p.tca_utc - row["tca"]).total_seconds())
        assert d <= 5.0, f"{row['sat']} @ {row['tca']}: TCA drift {d:.2f}s"


def test_sun_elevation_matches_baseline(prediction):
    """The baseline's sun elevations agree with the Astronomical Almanac
    low-precision algorithm to ~0.002°, so this pins solar accuracy, not just
    self-consistency. [SESSION 2026-07-15 — verified against the Almanac.]"""
    for row in _baseline_rows():
        p = _match(prediction, row)
        assert p.sun_elev_deg == pytest.approx(row["sun"], abs=0.05), (
            f"{row['sat']} @ {row['tca']}: sun elevation "
            f"{p.sun_elev_deg} vs baseline {row['sun']}")


def test_slant_and_max_off_nadir_match_baseline(prediction):
    for row in _baseline_rows():
        p = _match(prediction, row)
        assert p.slant_range_km == pytest.approx(row["slant"], abs=5.0)
        assert p.max_off_nadir_aoi_deg == pytest.approx(row["maxoff"], abs=0.25)
