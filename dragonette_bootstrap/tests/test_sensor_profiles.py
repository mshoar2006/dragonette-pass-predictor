"""Sensor profiles: Landsat 8/9 and Sentinel-2 A/B/C as predictable sensors.

The headline test is `test_sentinel2_profile_predicts_real_acquisitions`: it takes
scenes that genuinely happened, and asserts the profile predicts each one. That is
the difference between "the config parses" and "the sensor works".

Fully offline — real captured STAC + Celestrak payloads in fixtures/.
"""
import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
import passes as P  # noqa: E402

FIX = ROOT / "fixtures"
SCENES = json.loads(
    (FIX / "stac_scenes_geometry_validation_20260715.json").read_text())
EO_TLES = P._parse_3le_file((FIX / "tles_landsat_sentinel2_20260715.txt").read_text(),
                            {**P.LANDSAT.satellites, **P.SENTINEL2.satellites})
SITEA_KMZ = (FIX / "SiteA.kmz").read_bytes()
START = datetime(2026, 7, 14, 0, 0, tzinfo=timezone.utc)


def _point_kmz(lat, lon, half=0.02, name="AOI"):
    import io, zipfile
    coords = " ".join(f"{lon + dx},{lat + dy},0" for dx, dy in
                      [(-half, -half), (half, -half), (half, half),
                       (-half, half), (-half, -half)])
    kml = ('<kml xmlns="http://www.opengis.net/kml/2.2"><Document><Placemark>'
           f'<name>{name}</name><Polygon><outerBoundaryIs><LinearRing>'
           f'<coordinates>{coords}</coordinates>'
           '</LinearRing></outerBoundaryIs></Polygon></Placemark></Document></kml>')
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("doc.kml", kml)
    return buf.getvalue()


# ------------------------------------------------------------------ the profiles
def test_profiles_registered_and_lookupable():
    assert set(P.PROFILES) == {"dragonette", "landsat", "sentinel2"}
    assert P.get_profile(None) is P.DRAGONETTE          # default stays Dragonette
    assert P.get_profile("LANDSAT") is P.LANDSAT        # case-insensitive
    assert P.get_profile(" sentinel2 ") is P.SENTINEL2
    with pytest.raises(ValueError, match="unknown sensor"):
        P.get_profile("hubble")


def test_verified_specs_are_carried_in_the_profiles():
    """Specs checked against published sources: Landsat 185 km / 30 m / 705 km
    (NASA, USGS); Sentinel-2 290 km / 10 m / 786 km / 20.6 deg FOV (ESA Copernicus
    SentiWiki)."""
    assert (P.LANDSAT.swath_km, P.LANDSAT.gsd_m) == (185.0, 30.0)
    assert (P.SENTINEL2.swath_km, P.SENTINEL2.gsd_m) == (290.0, 10.0)
    assert P.LANDSAT.satellites == {"LANDSAT8": 39084, "LANDSAT9": 49260}
    assert P.SENTINEL2.satellites == {"SENTINEL2A": 40697, "SENTINEL2B": 42063,
                                      "SENTINEL2C": 60989}


def test_fov_half_angle_derivation_matches_the_published_fov():
    """NASA publishes Landsat's swath but not its FOV, so the envelope is derived.
    Sentinel-2 is the control: ESA states both, and they must agree."""
    s2 = P.fov_half_angle_deg(290.0, 786.0)
    assert s2 == pytest.approx(10.3, abs=0.2), f"derived {s2} vs ESA's stated 10.3"
    ls = P.fov_half_angle_deg(185.0, 705.0)
    assert ls == pytest.approx(7.5, abs=0.1), f"derived {ls} vs the 15 deg FOV (7.5 half)"
    # and the derivation must be monotonic in the obvious directions
    assert P.fov_half_angle_deg(290.0, 786.0) > P.fov_half_angle_deg(185.0, 786.0)
    assert P.fov_half_angle_deg(290.0, 400.0) > P.fov_half_angle_deg(290.0, 786.0)


def test_non_agile_sensors_have_no_marginal_band():
    """A fixed push-broom's envelope is its FOV — a hard optical limit. There is no
    tasking negotiation to stretch, so a 'marginal' tier would be a fiction."""
    for prof in (P.LANDSAT, P.SENTINEL2):
        assert not prof.agile
        assert prof.marginal_off_nadir_deg == prof.max_off_nadir_deg
        assert prof.min_sun_elev_deg == 0.0, "they image the daylit side on their own cycle"
    assert P.DRAGONETTE.agile
    assert P.DRAGONETTE.marginal_off_nadir_deg > P.DRAGONETTE.max_off_nadir_deg


def test_both_landsat_satellites_are_operational():
    """Landsat-8 and Landsat-9 both operational, checked against earth-search
    STAC: L8 has 14,632 L2 scenes in the trailing 30 d, latest 2026-07-10. An
    earlier build flagged L8 non-op off a Collection-2 processing-latency gap, which
    was wrong — this pins the correction."""
    assert P.LANDSAT.operational["LANDSAT8"] is True
    assert P.LANDSAT.operational["LANDSAT9"] is True


# --------------------------------------------------- the Dragonette path is intact
def test_dragonette_remains_the_default_and_is_unchanged():
    tles = P._parse_3le_file((FIX / "tles_real_20260714.txt").read_text(), P.SATELLITES)
    a = P.predict(SITEA_KMZ, days=3.0, start_utc=START, terrain_alt_m=400.0,
                  polygon_name="SITEA_100sqkm", tles=tles)
    b = P.predict(SITEA_KMZ, days=3.0, start_utc=START, terrain_alt_m=400.0,
                  polygon_name="SITEA_100sqkm", tles=tles, profile="dragonette")
    assert [(p.satellite, p.tca_utc, p.off_nadir_deg) for p in a.passes] == \
           [(p.satellite, p.tca_utc, p.off_nadir_deg) for p in b.passes]
    assert a.params["max_off_nadir_deg"] == 20.0 and a.params["min_sun_elev_deg"] == 20.0
    assert a.params["swath_km"] == 20.0 and a.params["gsd_nadir_m"] == 5.3
    assert a.params["sensor"] == "dragonette" and a.params["agile"] is True


def test_explicit_arguments_still_override_the_profile():
    tles = P._parse_3le_file((FIX / "tles_real_20260714.txt").read_text(), P.SATELLITES)
    pred = P.predict(SITEA_KMZ, days=2.0, start_utc=START, terrain_alt_m=400.0,
                     polygon_name="SITEA_100sqkm", tles=tles,
                     profile="dragonette", max_off_nadir_deg=5.0, min_sun_elev_deg=-90.0)
    assert pred.params["max_off_nadir_deg"] == 5.0
    assert all(abs(p.off_nadir_deg) <= 5.0 for p in pred.passes)


# ------------------------------------------------- the profiles against reality
@pytest.mark.parametrize("prof_key,platform_prefix", [
    ("sentinel2", "sentinel-2"), ("landsat", "landsat"),
])
def test_profile_predicts_real_acquisitions(prof_key, platform_prefix):
    """THE test. For scenes that genuinely happened, the profile must predict a
    pass at that time over that point — inside the sensor's own FOV envelope.

    This is what makes the profile real rather than plausible: it is the same
    check as test_geometry_validation.py, but driven through `predict()` and the
    profile's envelope instead of the raw geometry helpers.
    """
    prof = P.get_profile(prof_key)
    scenes = [f for f in SCENES["features"]
              if (f["properties"].get("platform") or "").startswith(platform_prefix)
              and f["properties"].get("proj:centroid")]
    assert scenes, f"fixture should carry {platform_prefix} scenes"

    checked = 0
    for s in scenes[:6]:                       # a handful is plenty; each is a full run
        p = s["properties"]
        c = p["proj:centroid"]
        when = datetime.fromisoformat(p["datetime"].replace("Z", "+00:00"))
        pred = P.predict(_point_kmz(c["lat"], c["lon"]), days=0.5,
                         start_utc=when - timedelta(hours=6),
                         profile=prof, tles=EO_TLES, include_nonoperational=True)
        hit = [x for x in pred.passes + pred.marginal + pred.nonoperational
               if abs((x.tca_utc - when).total_seconds()) < 60.0]
        assert hit, (f"{p['platform']} imaged {c['lat']:.3f},{c['lon']:.3f} at {when} "
                     f"but the {prof_key} profile predicted no pass within 60 s")
        # and it must fall inside the sensor's real field of view
        assert abs(hit[0].off_nadir_deg) <= prof.max_off_nadir_deg + 0.5
        checked += 1
    assert checked >= 3


def test_profile_reports_its_own_optics_not_dragonettes():
    """The bug this test exists for: swath/GSD were module constants, so a Landsat
    run would have reported Dragonette's 20 km swath and 5.3 m GSD."""
    s = SCENES["features"][0]["properties"]
    c = s["proj:centroid"]
    when = datetime.fromisoformat(s["datetime"].replace("Z", "+00:00"))
    kmz = _point_kmz(c["lat"], c["lon"])
    for key, swath, gsd in (("landsat", 185.0, 30.0), ("sentinel2", 290.0, 10.0)):
        pred = P.predict(kmz, days=0.5, start_utc=when - timedelta(hours=6),
                         profile=key, tles=EO_TLES)
        assert pred.params["swath_km"] == swath
        assert pred.params["gsd_nadir_m"] == gsd
        assert pred.params["agile"] is False
        for p in pred.passes:
            # effective GSD is the secant law on the sensor's own nadir GSD
            assert p.geometry["effective_gsd_m"] >= gsd


def test_method_sheet_describes_the_actual_sensor(tmp_path):
    """A Landsat workbook listing DRAG01-05 and a 20 deg envelope would be worse
    than no Method sheet at all."""
    from openpyxl import load_workbook
    s = SCENES["features"][0]["properties"]
    c = s["proj:centroid"]
    when = datetime.fromisoformat(s["datetime"].replace("Z", "+00:00"))
    pred = P.predict(_point_kmz(c["lat"], c["lon"]), days=0.5,
                     start_utc=when - timedelta(hours=6), profile="sentinel2",
                     tles=EO_TLES)
    f = tmp_path / "s2.xlsx"
    f.write_bytes(P.write_xlsx_multi([pred], "UTC"))
    rows = {r[0].value: str(r[1].value) for r in load_workbook(f)["Method"].iter_rows()
            if r[0].value}
    assert "Sentinel-2" in rows["Sensor"]
    assert "NOT taskable" in rows["Sensor"]
    assert "SENTINEL2A=40697" in rows["Satellites (NORAD)"]
    assert "DRAG" not in rows["Satellites (NORAD)"]
    assert "290 km" in rows["Swath / nadir GSD"] and "10 m" in rows["Swath / nadir GSD"]
    assert "PREDICTED ACQUISITIONS" in rows["Caveat"]


def test_cli_rejects_an_unknown_sensor(tmp_path):
    import subprocess
    r = subprocess.run(
        [sys.executable, str(ROOT / "src" / "cli.py"), str(FIX / "SiteA.kmz"),
         "--polygon", "SITEA_100sqkm", "--sensor", "hubble",
         "--tle-file", str(FIX / "demo_tles_synthetic.txt"), "-o", str(tmp_path / "x.xlsx")],
        capture_output=True, text=True)
    assert r.returncode == 2 and "Traceback" not in r.stderr
    assert "unknown sensor" in r.stderr.lower()
