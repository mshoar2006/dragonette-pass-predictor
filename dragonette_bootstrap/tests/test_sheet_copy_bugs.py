"""P2: sheet-writer text must describe the sensor that actually ran, not
Dragonette by default.

Two bugs, both in the xlsx writer rather than the geometry:

1. `_marginal_note` assumed max_off_nadir_deg < marginal_off_nadir_deg,
   which only holds for Dragonette's own agile stretch-tier convention. An
   off-nadir gate overridden wider than a profile's marginal bound rendered
   an inverted band ("Passes at 20-10 deg off-nadir"); a fixed push-broom's
   marginal bound always equals its max (no real tier to describe) so the
   note should be suppressed entirely, not printed as "7-7 deg".
2. The "Passes per satellite" summary in write_xlsx listed the module-level
   Dragonette SATELLITES constant regardless of which sensor actually ran,
   so a Landsat/Sentinel-2 workbook carried an always-empty DRAG01-05 block.

Fully offline — real captured STAC + Celestrak payloads in fixtures/.
"""
import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
import passes as P  # noqa: E402

FIX = ROOT / "fixtures"
SCENES = json.loads(
    (FIX / "stac_scenes_geometry_validation_20260715.json").read_text())
EO_TLES = P._parse_3le_file((FIX / "tles_landsat_sentinel2_20260715.txt").read_text(),
                            {**P.LANDSAT.satellites, **P.SENTINEL2.satellites})


def _point_kmz(lat, lon, half=0.02, name="AOI"):
    import io, zipfile
    coords = " ".join(f"{lon + dx},{lat + dy},0" for dx, dy in
                      [(-half, -half), (half, -half), (half, half),
                       (-half, half), (-half, -half)])
    kml = ('<kml xmlns="http://www.opengis.net/kml/2.2"><Document><Placemark>'
           f'<name>{name}</name><Polygon><outerBoundaryIs><LinearRing>'
           f'<coordinates>{coords}</coordinates>'
           '</LinearRing></outerBoundaryIs></Polygon></Placemark></Document></kml>')
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("doc.kml", kml)
    return buf.getvalue()


def _sentinel2_pred():
    s = SCENES["features"][0]["properties"]
    c = s["proj:centroid"]
    when = datetime.fromisoformat(s["datetime"].replace("Z", "+00:00"))
    return P.predict(_point_kmz(c["lat"], c["lon"]), days=1.0,
                     start_utc=when - timedelta(hours=12), profile="sentinel2",
                     tles=EO_TLES)


# ------------------------------------------------------------------ P2 item 2
def test_per_satellite_summary_lists_the_actual_sensors_roster(tmp_path):
    pred = _sentinel2_pred()
    f = tmp_path / "s2.xlsx"
    f.write_bytes(P.write_xlsx(pred, "UTC"))
    ws = load_workbook(f)["Passes"]
    names = {c.value for row in ws.iter_rows() for c in row
            if c.value in P.SENTINEL2.satellites or c.value in P.SATELLITES}
    assert names & set(P.SENTINEL2.satellites), "expected SENTINEL2A/B/C in the summary"
    assert not (names & set(P.SATELLITES)), \
        f"a Sentinel-2 workbook must not list Dragonette satellites: {names}"


def test_combined_workbook_still_lists_every_constellation(tmp_path):
    """The combined 'all sensors' summary is the one place DRAG names are
    correct — this must keep working."""
    names = sorted(P._satellite_names_for({"sensor": P.COMBINED_KEY}))
    sats, _ = P.combined_roster()
    assert names == sorted(sats)
    assert set(P.SATELLITES) <= set(names)
    assert set(P.LANDSAT.satellites) <= set(names)


# ------------------------------------------------------------------ P2 item 1
def test_marginal_note_is_suppressed_for_fixed_sensors():
    """Landsat/Sentinel-2 have no real marginal tier (their envelope is the
    FOV, a hard optical limit) -- the note must not print a vacuous band."""
    for prof in (P.LANDSAT, P.SENTINEL2):
        params = dict(max_off_nadir_deg=prof.max_off_nadir_deg,
                     marginal_off_nadir_deg=prof.marginal_off_nadir_deg,
                     min_sun_elev_deg=prof.min_sun_elev_deg,
                     marginal_sun_elev_deg=prof.marginal_sun_elev_deg,
                     sensor=prof.key)
        assert P._marginal_note(params) is None


def test_marginal_note_handles_an_inverted_band_without_garbling_it():
    """The exact scenario the bug shipped: max_off_nadir_deg overridden wider
    than marginal_off_nadir_deg (e.g. an unclamped P0 gate on a fixed
    sensor). The note must show the correct ascending order, not '20-10'."""
    params = dict(max_off_nadir_deg=20.0, marginal_off_nadir_deg=10.0,
                 min_sun_elev_deg=20.0, marginal_sun_elev_deg=15.0,
                 sensor="landsat")
    note = P._marginal_note(params)
    assert note is not None
    assert "10" in note and "20" in note
    assert "10–20" in note or "10-20" in note, note
    assert "20–10" not in note and "20-10" not in note, note


def test_marginal_note_still_describes_dragonettes_real_band():
    params = dict(max_off_nadir_deg=P.DRAGONETTE.max_off_nadir_deg,
                 marginal_off_nadir_deg=P.DRAGONETTE.marginal_off_nadir_deg,
                 min_sun_elev_deg=P.DRAGONETTE.min_sun_elev_deg,
                 marginal_sun_elev_deg=P.DRAGONETTE.marginal_sun_elev_deg,
                 sensor="dragonette")
    note = P._marginal_note(params)
    assert note is not None
    assert "Wyvern" in note


def test_marginal_note_does_not_mention_wyvern_for_a_fixed_sensor():
    """Even in the (currently hypothetical) case of a real band on a
    non-agile profile, the operator wording must not default to Wyvern."""
    params = dict(max_off_nadir_deg=5.0, marginal_off_nadir_deg=8.0,
                 min_sun_elev_deg=0.0, marginal_sun_elev_deg=0.0, sensor="landsat")
    note = P._marginal_note(params)
    assert note is not None
    assert "Wyvern" not in note


def test_sentinel2_workbook_has_no_marginal_note(tmp_path):
    pred = _sentinel2_pred()
    f = tmp_path / "s2_marginal.xlsx"
    f.write_bytes(P.write_xlsx(pred, "UTC"))
    ws = load_workbook(f)["Marginal - stretch"]
    footer_cells = [c.value for row in ws.iter_rows() for c in row if c.value]
    assert not any("Wyvern" in str(v) for v in footer_cells)
    assert not any("outside the standard list" in str(v) for v in footer_cells)
