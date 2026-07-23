"""P0: a fixed push-broom's off-nadir gate must never exceed its own optic.

`predict()` used to let a caller-supplied `max_off_nadir_deg` override the
profile default with no clamp for non-agile sensors (Landsat/Sentinel-2). A
gate wider than the instrument's real FOV half-angle let physically
impossible rows land on Passes/Marginal with a good/marginal quality badge,
while AOI coverage (computed from the real FOV) correctly read 0%. Confirmed
2026-07-23 against a real workbook: 12 of 18 Landsat rows at 10.5-13 deg
off-nadir, 0% coverage, listed as passes. A reader would plan around scenes
that will never exist.

Fully offline — real captured STAC + Celestrak payloads in fixtures/.
"""
import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
import passes as P  # noqa: E402

FIX = ROOT / "fixtures"
SCENES = json.loads(
    (FIX / "stac_scenes_geometry_validation_20260715.json").read_text())
EO_TLES = P._parse_3le_file((FIX / "tles_landsat_sentinel2_20260715.txt").read_text(),
                            {**P.LANDSAT.satellites, **P.SENTINEL2.satellites})


def _point_kmz(lat, lon, half=0.02, name="AOI"):
    import io, zipfile
    coords = " ".join(f"{lon + dx},{lat + dy},0" for dx, dy in
                      [(-half, -half), (half, -half), (half, half),
                       (-half, half), (-half, -half)])
    kml = ('<kml xmlns="http://www.opengis.net/kml/2.2"><Document><Placemark>'
           f'<name>{name}</name><Polygon><outerBoundaryIs><LinearRing>'
           f'<coordinates>{coords}</coordinates>'
           '</LinearRing></outerBoundaryIs></Polygon></Placemark></Document></kml>')
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("doc.kml", kml)
    return buf.getvalue()


def _real_landsat_run(max_off_nadir_deg):
    s = SCENES["features"][0]["properties"]
    c = s["proj:centroid"]
    when = datetime.fromisoformat(s["datetime"].replace("Z", "+00:00"))
    return P.predict(_point_kmz(c["lat"], c["lon"]), days=1.0,
                     start_utc=when - timedelta(hours=12), profile="landsat",
                     tles=EO_TLES, max_off_nadir_deg=max_off_nadir_deg)


def test_wide_gate_is_clamped_to_the_sensors_own_fov():
    """A 20 deg gate requested on Landsat (7.47 deg FOV) must not survive into
    the prediction params, must warn, and must not admit any row beyond the
    real FOV -- regardless of what off-nadir a real fixture scene happens to
    produce."""
    pred = _real_landsat_run(max_off_nadir_deg=20.0)
    assert pred.params["max_off_nadir_deg"] == pytest.approx(P.LANDSAT.max_off_nadir_deg)
    assert pred.params["marginal_off_nadir_deg"] == pytest.approx(P.LANDSAT.max_off_nadir_deg)
    assert any("clamped" in w and "FOV half-angle" in w for w in pred.warnings), pred.warnings
    for p in pred.passes + pred.marginal:
        assert abs(p.off_nadir_deg) <= P.LANDSAT.max_off_nadir_deg + 1e-6


def test_gate_within_the_fov_is_left_alone():
    """An honest, already-narrower gate must not warn or be touched."""
    tight = P.LANDSAT.max_off_nadir_deg - 1.0
    pred = _real_landsat_run(max_off_nadir_deg=tight)
    assert pred.params["max_off_nadir_deg"] == pytest.approx(tight)
    assert not any("clamped" in w for w in pred.warnings)


def test_agile_sensor_gate_is_never_clamped():
    """Dragonette's gate is a tasking policy, not an optic -- widening it (the
    marginal/stretch tier) is a legitimate, existing feature and must not
    trigger the fixed-sensor clamp."""
    tles = P._parse_3le_file((FIX / "tles_real_20260714.txt").read_text(), P.SATELLITES)
    pred = P.predict((FIX / "SiteA.kmz").read_bytes(), days=2.0,
                     start_utc=datetime(2026, 7, 14, tzinfo=timezone.utc),
                     terrain_alt_m=400.0, polygon_name="SITEA_100sqkm", tles=tles,
                     profile="dragonette", max_off_nadir_deg=60.0)
    assert pred.params["max_off_nadir_deg"] == 60.0
    assert not any("clamped" in w for w in pred.warnings)


def test_method_sheet_does_not_overstate_the_clamped_gate(tmp_path):
    """The Access filter row must report the clamped value, not the caller's
    original wider request."""
    from openpyxl import load_workbook
    pred = _real_landsat_run(max_off_nadir_deg=20.0)
    f = tmp_path / "landsat_clamped.xlsx"
    f.write_bytes(P.write_xlsx_multi([pred], "UTC"))
    rows = {r[0].value: str(r[1].value) for r in load_workbook(f)["Method"].iter_rows()
            if r[0].value}
    assert "20" not in rows["Access filter"].split("°")[0]
    assert f"{P.LANDSAT.max_off_nadir_deg:g}" in rows["Access filter"]


# --------------------------------------------------- belt-and-braces (sheet writer)
def _synthetic_pass(off_nadir, coverage_pct):
    return P.Pass(satellite="LANDSAT9", tca_utc=datetime(2026, 7, 20, tzinfo=timezone.utc),
                 off_nadir_deg=off_nadir, sun_elev_deg=45.0,
                 max_off_nadir_aoi_deg=off_nadir, slant_range_km=900.0,
                 tle_epoch_utc=datetime(2026, 7, 19, tzinfo=timezone.utc),
                 category="standard", quality="good", coverage_pct=coverage_pct)


def test_zero_coverage_row_is_never_kept_for_a_fixed_sensor():
    """The exact scenario the bug shipped: a synthetic non-agile profile, gate
    20 deg, one pass at 12 deg off-nadir with 0% AOI coverage -- it must be
    routed out of Passes/Marginal, never carry a good/marginal badge."""
    rows = [_synthetic_pass(12.0, 0.0), _synthetic_pass(3.0, 0.65)]
    kept, near_miss = P._partition_near_miss(rows, agile=False)
    assert [p.off_nadir_deg for p in kept] == [3.0]
    assert [p.off_nadir_deg for p in near_miss] == [12.0]


def test_zero_coverage_row_is_not_clamped_away_for_an_agile_sensor():
    """Dragonette can legitimately have low/zero coverage on a marginal-tier
    pass at the edge of a stretch tasking -- agile sensors are not routed."""
    rows = [_synthetic_pass(25.0, 0.0)]
    kept, near_miss = P._partition_near_miss(rows, agile=True)
    assert kept == rows and near_miss == []


def test_near_miss_sheet_carries_no_quality_badge(tmp_path):
    from openpyxl import load_workbook
    aoi = P.AOI(name="Synthetic", vertices_lonlat=[(0.0, 0.0)],
               centroid_lon=0.0, centroid_lat=0.0, terrain_alt_m=0.0)
    pred = P.Prediction(aoi=aoi, start_utc=datetime(2026, 7, 20, tzinfo=timezone.utc),
                        end_utc=datetime(2026, 7, 21, tzinfo=timezone.utc),
                        passes=[_synthetic_pass(12.0, 0.0)],
                        params=dict(agile=False, max_off_nadir_deg=20.0,
                                    marginal_off_nadir_deg=20.0, min_sun_elev_deg=0.0,
                                    marginal_sun_elev_deg=0.0, sensor="landsat"))
    f = tmp_path / "nm.xlsx"
    f.write_bytes(P.write_xlsx(pred, "UTC"))
    wb = load_workbook(f)
    assert P.NEAR_MISS_SHEET in wb.sheetnames
    assert "Passes" in wb.sheetnames
    passes_rows = list(wb["Passes"].iter_rows(min_row=2, values_only=True))
    assert not any(r[0] == "LANDSAT9" for r in passes_rows if r[0])
    nm_ws = wb[P.NEAR_MISS_SHEET]
    header = [c.value for c in next(nm_ws.iter_rows(min_row=1, max_row=1))]
    quality_col = header.index("Quality")
    first_row = next(nm_ws.iter_rows(min_row=2, max_row=2, values_only=True))
    assert first_row[quality_col] == P.NEAR_MISS_QUALITY
