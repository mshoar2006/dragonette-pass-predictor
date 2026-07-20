"""v2 offline tests (R4–R9). Fully offline: real AOI fixtures + the synthetic
demo TLE file; no network. Live-sky correctness is a separate checklist
(VALIDATION.md)."""
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
FIX = ROOT / "fixtures"
sys.path.insert(0, str(SRC))
import passes as P  # noqa: E402

SITEA_KMZ = (FIX / "SiteA.kmz").read_bytes()
SITEC_KMZ = (FIX / "SiteC.kmz").read_bytes()
SITEB_KMZ = (FIX / "SiteB.kmz").read_bytes()
DEMO_TLES = str(FIX / "demo_tles_synthetic.txt")
START = datetime(2026, 7, 14, 0, 0, tzinfo=timezone.utc)


def _tles():
    return P._parse_3le_file(Path(DEMO_TLES).read_text(), P.SATELLITES)


SITEA = (-20.0, 150.0)


def make_kmz(lat, lon, half_deg=0.045, name="TEST_AOI"):
    import io, zipfile
    coords = " ".join(
        f"{lon + dx},{lat + dy},0"
        for dx, dy in [(-half_deg, -half_deg), (half_deg, -half_deg),
                       (half_deg, half_deg), (-half_deg, half_deg),
                       (-half_deg, -half_deg)])
    kml = (f'<kml xmlns="http://www.opengis.net/kml/2.2"><Document><Placemark>'
           f'<name>{name}</name><Polygon><outerBoundaryIs><LinearRing>'
           f'<coordinates>{coords}</coordinates>'
           f'</LinearRing></outerBoundaryIs></Polygon></Placemark></Document></kml>')
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("doc.kml", kml)
    return buf.getvalue()


def _predict(kmz, polygon=None, **kw):
    defaults = dict(days=14.0, start_utc=START, terrain_alt_m=400.0,
                    min_sun_elev_deg=-90.0, marginal_sun_elev_deg=-90.0,
                    polygon_name=polygon, tles=_tles())
    defaults.update(kw)
    return P.predict(kmz, **defaults)


# ---------------------------------------------------------------- R7
def test_list_polygons_counts():
    assert P.list_polygons(SITEA_KMZ) == ["AOI 1", "AOI 2", "SITEA_100sqkm"]
    assert P.list_polygons(SITEC_KMZ) == ["Site C trial site"]
    assert len(P.list_polygons(SITEB_KMZ)) == 2


# ------------------------------------------------- untrusted KMZ name handling
def _named_kmz(raw_name):
    """A KMZ whose <name> is inserted verbatim — i.e. what an emailed KMZ can carry."""
    import io as _io, zipfile as _zf
    coords = " ".join(f"{150.0 + dx},{-20.0 + dy},0" for dx, dy in
                      [(-.02, -.02), (.02, -.02), (.02, .02), (-.02, .02), (-.02, -.02)])
    kml = ('<kml xmlns="http://www.opengis.net/kml/2.2"><Document><Placemark>'
           f'<name>{raw_name}</name><Polygon><outerBoundaryIs><LinearRing>'
           f'<coordinates>{coords}</coordinates>'
           '</LinearRing></outerBoundaryIs></Polygon></Placemark></Document></kml>')
    buf = _io.BytesIO()
    with _zf.ZipFile(buf, "w") as z:
        z.writestr("doc.kml", kml)
    return buf.getvalue()


def test_cdata_wrapped_name_is_stripped_of_markup():
    """Google Earth writes names in CDATA, so `<![CDATA[<img onerror=...>]]>`
    survives a bare <name> scrape. It reached the SPA's popup innerHTML.
    [SESSION 2026-07-15]"""
    data = _named_kmz("<![CDATA[<img src=x onerror=alert(1)>]]>")
    name = P.list_polygons(data)[0]
    assert "<" not in name and ">" not in name
    assert "CDATA" not in name and "onerror" not in name
    assert P.parse_kmz(data, 0.0).name == name


def test_entity_encoded_markup_in_name_is_neutralised():
    data = _named_kmz("&lt;script&gt;alert(1)&lt;/script&gt;Paddock")
    name = P.list_polygons(data)[0]
    assert "<" not in name and ">" not in name and "script" not in name.lower()


def test_ordinary_names_survive_sanitisation_unchanged():
    """Sanitising must not disturb real AOI names — they are also the identifier
    `polygon_name` matches on."""
    assert P.list_polygons(SITEA_KMZ) == ["AOI 1", "AOI 2", "SITEA_100sqkm"]
    for raw in ("SITEA_100sqkm", "AOI 1", "Site C trial site",
                "Côte paddock (north)", "Paddock #3 - west"):
        assert P.list_polygons(_named_kmz(raw))[0] == raw


def test_formula_like_aoi_name_is_not_a_live_formula_in_xlsx(tmp_path):
    """openpyxl infers a leading '=' as a formula, so a polygon named
    `=cmd|'/c calc'!A1` would ship a live DDE payload in a workbook DEVELOPMENT.md
    says is circulated to research teams. [SESSION 2026-07-15]"""
    from openpyxl import load_workbook
    payload = "=cmd|'/c calc'!A1"
    pred = P.predict(_named_kmz(payload), days=2.0, start_utc=START,
                     min_sun_elev_deg=-90.0, marginal_sun_elev_deg=-90.0,
                     max_off_nadir_deg=60.0, marginal_off_nadir_deg=60.0, tles=_tles())
    assert pred.aoi.name == payload          # identity preserved for matching
    f = tmp_path / "inj.xlsx"
    f.write_bytes(P.write_xlsx_multi([pred], "Australia/Brisbane"))
    wb = load_workbook(f)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value == payload:
                    assert c.data_type == "s", "AOI name must be text, not a formula"


def test_quoted_aoi_name_does_not_corrupt_countif(tmp_path):
    """A double quote in a name used to close the COUNTIF string argument and
    emit a malformed formula; Excel then repairs the sheet by discarding it.

    Needs >=2 predictions: write_xlsx_multi delegates to write_xlsx for a single
    AOI, and the per-AOI COUNTIF summary only exists on the multi-AOI path.
    """
    from openpyxl import load_workbook
    kw = dict(days=2.0, start_utc=START, min_sun_elev_deg=-90.0,
              marginal_sun_elev_deg=-90.0, max_off_nadir_deg=60.0,
              marginal_off_nadir_deg=60.0, tles=_tles())
    pred = P.predict(_named_kmz('Paddock "North"'), **kw)
    other = P.predict(_named_kmz("Paddock South"), **kw)
    assert pred.aoi.name == 'Paddock "North"'
    f = tmp_path / "q.xlsx"
    f.write_bytes(P.write_xlsx_multi([pred, other], "Australia/Brisbane"))
    formulas = [c.value for ws in load_workbook(f).worksheets for row in ws.iter_rows()
                for c in row if isinstance(c.value, str) and c.value.startswith("=COUNTIF")]
    assert formulas, "COUNTIF summaries should exist"
    # Per-satellite COUNTIFs legitimately quote our own constants ("DRAG01").
    # What must never appear inside a formula is the untrusted AOI name.
    for fx in formulas:
        assert "Paddock" not in fx, f"COUNTIF must not interpolate the AOI name: {fx}"
    assert any(re.search(r"=COUNTIF\(A\d+:A\d+,A\d+\)", fx) for fx in formulas), \
        "per-AOI COUNTIF should reference the name cell, not inline it"


def test_parse_ambiguous_raises_with_names():
    with pytest.raises(P.AmbiguousPolygonError) as ei:
        P.parse_kmz(SITEA_KMZ, 400.0)
    assert ei.value.names == ["AOI 1", "AOI 2", "SITEA_100sqkm"]


def test_parse_named_substring_and_single():
    aoi = P.parse_kmz(SITEA_KMZ, 400.0, polygon_name="100sqkm")
    assert aoi.name == "SITEA_100sqkm"
    # single-polygon KMZ needs no disambiguation
    assert P.parse_kmz(SITEC_KMZ).name.startswith("Site C")


def test_parse_bad_name_lists_options():
    with pytest.raises(ValueError) as ei:
        P.parse_kmz(SITEA_KMZ, 400.0, polygon_name="does-not-exist")
    assert "AOI 1" in str(ei.value)


def test_predict_ambiguous_raises():
    with pytest.raises(P.AmbiguousPolygonError):
        _predict(SITEA_KMZ)  # no polygon on a 3-polygon KMZ


def test_cli_ambiguous_exits_nonzero(tmp_path):
    r = subprocess.run(
        [sys.executable, str(SRC / "cli.py"), str(FIX / "SiteA.kmz"),
         "--tle-file", DEMO_TLES, "--start", "2026-07-14T00:00:00",
         "-o", str(tmp_path / "x.xlsx")],
        capture_output=True, text=True)
    assert r.returncode == 2
    assert "AOI 1" in r.stderr and "SITEA_100sqkm" in r.stderr


def test_cli_all_polygons_workbook(tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "all.xlsx"
    r = subprocess.run(
        [sys.executable, str(SRC / "cli.py"), str(FIX / "SiteA.kmz"),
         "--all-polygons", "--alt", "400", "--min-sun", "-90",
         "--tle-file", DEMO_TLES, "--start", "2026-07-14T00:00:00",
         "-o", str(out)],
        capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    wb = load_workbook(out)
    assert wb["Passes"].cell(1, 1).value == "AOI"  # combined sheet has AOI col


def test_write_xlsx_multi_structure(tmp_path):
    from openpyxl import load_workbook
    preds = [_predict(SITEA_KMZ, "AOI 1"), _predict(SITEA_KMZ, "SITEA_100sqkm")]
    blob = P.write_xlsx_multi(preds, tz_name="Australia/Brisbane")
    f = tmp_path / "m.xlsx"; f.write_bytes(blob)
    wb = load_workbook(f)
    assert {"Passes", "Marginal - stretch", "Method"} <= set(wb.sheetnames)
    ws = wb["Passes"]
    assert ws.cell(1, 1).value == "AOI" and ws.cell(1, 2).value == "Satellite"
    # per-AOI COUNTIF summary present
    assert any(str(c.value).startswith("=COUNTIF")
               for row in ws.iter_rows() for c in row if c.value)


def test_write_xlsx_multi_single_delegates(tmp_path):
    from openpyxl import load_workbook
    pred = _predict(SITEC_KMZ)
    blob = P.write_xlsx_multi([pred], tz_name="Australia/Brisbane")
    f = tmp_path / "s.xlsx"; f.write_bytes(blob)
    ws = load_workbook(f)["Passes"]
    assert ws.cell(1, 1).value == "Satellite"  # single -> unchanged format


# ---------------------------------------------------------------- R5
def test_operational_dict_isolates_drag05():
    assert P.OPERATIONAL["DRAG05"] is False
    assert all(P.OPERATIONAL[n] for n in P.SATELLITES if n != "DRAG05")


def test_drag05_separated_not_counted():
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=60.0,
                    marginal_off_nadir_deg=60.0)
    assert all(p.satellite != "DRAG05" for p in pred.passes)
    assert all(p.satellite != "DRAG05" for p in pred.marginal)
    assert all(p.satellite == "DRAG05" for p in pred.nonoperational)
    assert all(p.operational is False for p in pred.nonoperational)
    # headline count excludes DRAG05
    assert all(p.operational for p in pred.passes)


def test_exclude_nonoperational_drops_drag05():
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=60.0,
                    marginal_off_nadir_deg=60.0, include_nonoperational=False)
    assert pred.nonoperational == []


def test_json_has_operational_and_nonop():
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=60.0,
                    marginal_off_nadir_deg=60.0)
    body = P.prediction_json([pred])
    assert "nonoperational" in body
    assert all(r["operational"] for r in body["passes"])
    assert all(r["operational"] is False for r in body["nonoperational"])


def test_nonop_sheet_present_and_badged(tmp_path):
    from openpyxl import load_workbook
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=60.0,
                    marginal_off_nadir_deg=60.0)
    blob = P.write_xlsx(pred, tz_name="Australia/Brisbane")
    f = tmp_path / "n.xlsx"; f.write_bytes(blob)
    wb = load_workbook(f)
    assert "Non-operational" in wb.sheetnames
    ws = wb["Non-operational"]
    joined = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "DRAG05" in joined and "not yet operational" in joined.lower()


# ---------------------------------------------------------------- R4
def test_timeline_bar_count_equals_pass_count():
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=45.0,
                    marginal_off_nadir_deg=60.0)
    n = len(pred.passes) + len(pred.marginal) + len(pred.nonoperational)
    fig, ax = P.build_timeline_figure(pred, tz_name="Australia/Brisbane")
    assert len(ax.patches) == n
    import matplotlib.pyplot as plt
    plt.close(fig)


def test_timeline_drag05_hatched_and_bands_by_offnadir():
    import matplotlib.colors as mcolors
    import matplotlib.pyplot as plt
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=60.0,
                    marginal_off_nadir_deg=60.0)
    assert pred.nonoperational, "need DRAG05 bars to assert their styling"
    fig, ax = P.build_timeline_figure(pred, tz_name="Australia/Brisbane")
    nonop = [pt for pt in ax.patches if ":nonop:" in (pt.get_gid() or "")]
    op = [pt for pt in ax.patches if ":op:" in (pt.get_gid() or "")]
    assert len(nonop) == len(pred.nonoperational)
    # DRAG05 stays distinct via hatch (constraint 3); operational bars are not hatched
    for pt in nonop:
        assert pt.get_hatch() == "//"
    for pt in op:
        assert not pt.get_hatch()
    # colour encodes off-nadir band: near-nadir green, >12° red
    bands = {mcolors.to_hex(pt.get_facecolor()) for pt in ax.patches}
    assert "#22c55e" in bands or "#f59e0b" in bands or "#ef4444" in bands
    assert P._offnadir_band(3.0)[0] == "#22c55e"     # 0–5° green
    assert P._offnadir_band(8.0)[0] == "#f59e0b"     # 5–12° amber
    assert P._offnadir_band(-19.0)[0] == "#ef4444"   # |off| >12° red (sign-agnostic)
    plt.close(fig)


def test_timeline_cloud_strip_present_when_cloud():
    import matplotlib.pyplot as plt
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=45.0)
    P.attach_cloud(pred, forecast_json=json.loads(
        (FIX / "openmeteo_forecast_sample.json").read_text()),
        now=P.datetime(2026, 7, 14, tzinfo=P.timezone.utc))
    assert pred.cloud_daily, "attach_cloud should populate a daily cloud summary"
    fig, ax = P.build_timeline_figure(pred, tz_name="Australia/Brisbane")
    assert len(fig.axes) == 2               # bar axis + cloud strip axis
    plt.close(fig)
    # without cloud, no strip axis
    fig2, _ = P.build_timeline_figure(
        _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=45.0))
    assert len(fig2.axes) == 1
    plt.close(fig2)


def test_render_timeline_png_bytes():
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=45.0)
    png = P.render_timeline_png(pred, tz_name="Australia/Brisbane")
    assert png[:8] == b"\x89PNG\r\n\x1a\n" and len(png) > 2000


def test_timeline_sheet_embedded(tmp_path):
    from openpyxl import load_workbook
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=45.0)
    blob = P.write_xlsx(pred, tz_name="Australia/Brisbane")
    f = tmp_path / "t.xlsx"; f.write_bytes(blob)
    wb = load_workbook(f)
    assert "Timeline" in wb.sheetnames
    assert len(wb["Timeline"]._images) == 1


def test_index_serves_spa():
    from fastapi.testclient import TestClient
    import app as A
    r = TestClient(A.app).get("/")
    assert r.status_code == 200 and "text/html" in r.headers["content-type"]
    assert "Dragonette Pass Predictor" in r.text
    # SPA drives the JSON endpoint, renders the timeline client-side (Plotly),
    # plots the AOI on a map (Leaflet), and offers xlsx download
    assert "/predict/json" in r.text and "/predict" in r.text
    assert "Plotly" in r.text and "leaflet" in r.text.lower()


def test_api_timeline_png(monkeypatch):
    from fastapi.testclient import TestClient
    import app as A
    monkeypatch.setattr(P, "fetch_tles", lambda *a, **k: (_tles(), []))
    client = TestClient(A.app)
    r = client.post("/timeline.png",
                    files={"kmz": ("t.kmz", SITEC_KMZ)},
                    data={"days": "14", "alt": "10", "min_sun": "-90"})
    assert r.status_code == 200
    assert r.headers["content-type"] == "image/png"
    assert r.content[:8] == b"\x89PNG\r\n\x1a\n"


# ---------------------------------------------------------------- WGS84 nadir (opt-in)
def test_ecef_geodetic_roundtrip():
    for lat, lon, alt in [(-20.0, 150.0, 0.4), (0.0, 0.0, 0.0),
                          (45.0, -120.0, 0.0), (-40.35, 175.61, 0.02)]:
        r = P.geodetic_to_ecef(lat, lon, alt)
        glat, glon = P.ecef_to_geodetic_latlon(r)
        assert abs(glat - lat) < 1e-6 and abs(((glon - lon + 180) % 360) - 180) < 1e-6


def test_nadir_ellipsoid_optin_small_shift_and_default_unchanged():
    # default (geocentric) must be byte-identical to the validated baseline
    base = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=25.0,
                    marginal_off_nadir_deg=25.0)
    ell = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=25.0,
                   marginal_off_nadir_deg=25.0, nadir_ellipsoid=True)
    assert base.params["nadir_ellipsoid"] is False
    assert ell.params["nadir_ellipsoid"] is True
    g = base.passes + base.marginal
    maxd = 0.0
    for p in ell.passes + ell.marginal:
        m = [q for q in g if q.satellite == p.satellite
             and abs((q.tca_utc - p.tca_utc).total_seconds()) < 120]
        if m:
            maxd = max(maxd, abs(abs(p.off_nadir_deg) - abs(m[0].off_nadir_deg)))
    # geocentric-vs-geodetic nadir differs by up to ~0.2°, never more
    assert 0.0 < maxd <= 0.25


# ---------------------------------------------------------------- Tier A (quality metrics)
def test_acquisition_geometry_fields_and_sanity():
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=25.0,
                    marginal_off_nadir_deg=25.0)
    assert pred.passes
    for p in pred.passes:
        g = p.geometry
        assert set(g) >= {"view_zenith_deg", "sun_glint_deg", "glint_risk",
                          "effective_gsd_m", "airmass_sun", "node",
                          "ground_track_heading_deg", "phase_deg"}
        # view zenith (incidence) always exceeds off-nadir on a curved Earth
        assert g["view_zenith_deg"] >= abs(p.off_nadir_deg) - 0.05
        # effective GSD >= nadir GSD, grows with off-nadir
        assert g["effective_gsd_m"] >= P.GSD_NADIR_M - 0.01
        assert 0 <= g["sun_glint_deg"] <= 180
        assert g["glint_risk"] in ("high", "caution", "low")
        assert p.node in ("ascending", "descending")
        assert 0.0 <= p.local_solar_time_h < 24.0
        assert p.quality in ("good", "marginal", "poor")


def test_effective_gsd_grows_with_offnadir():
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=25.0,
                    marginal_off_nadir_deg=25.0)
    rows = sorted(pred.passes + pred.marginal, key=lambda p: abs(p.off_nadir_deg))
    lo, hi = rows[0], rows[-1]
    assert lo.geometry["effective_gsd_m"] <= hi.geometry["effective_gsd_m"]
    # secant law stays sane within the ~20° envelope (well under +30%)
    assert hi.geometry["effective_gsd_m"] < P.GSD_NADIR_M * 1.3


def test_timing_sigma_grows_with_tle_age():
    # within one 14-day run, later passes are further from the (fixed) TLE epoch
    # ⇒ larger along-track timing uncertainty
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=25.0,
                    marginal_off_nadir_deg=25.0)
    rows = sorted(pred.passes + pred.marginal, key=lambda p: p.tca_utc)
    assert all(p.timing_sigma_s is not None and p.timing_sigma_s > 0 for p in rows)
    assert rows[-1].timing_sigma_s > rows[0].timing_sigma_s


def test_optical_obstruction_and_cirrus_flag():
    c = P.CloudInfo(1, "forecast", total=60, low=10, mid=20, high=80)
    assert c.optical_obstruction == round(10 + 20 + 0.4 * 80, 1)   # 62.0
    assert c.cirrus_flag is True
    clear = P.CloudInfo(1, "forecast", total=5, low=0, mid=0, high=5)
    assert clear.cirrus_flag is False and clear.optical_obstruction == 2.0


def test_tier2_extends_to_15_days():
    assert P.TIER2_MAX_DAYS == 15.0
    assert P._tier_for(12.0) == 2 and P._tier_for(16.0) == 3


def test_quality_columns_in_xlsx(tmp_path):
    from openpyxl import load_workbook
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=25.0)
    f = tmp_path / "q.xlsx"; f.write_bytes(P.write_xlsx(pred))
    hdr = [c.value for c in load_workbook(f)["Passes"][1]]
    for col in ("Quality", "Node", "Eff. GSD (m)", "Sun Glint (°)", "TCA ± (s)"):
        assert col in hdr


# ---------------------------------------------------------------- B3 soft p_clear
def test_soft_p_clear_avoids_hard_0_1():
    # a set of members split around the threshold should give a smooth ~0.5,
    # not a hard fraction; members all far clear/cloudy still saturate
    now = P.datetime(2026, 7, 14, tzinfo=P.timezone.utc)
    pred = _cloud_pred()
    P.attach_cloud(pred, forecast_json=FORECAST_JSON, ensemble_json=ENSEMBLE_JSON,
                   threshold=30.0, now=now)
    t2 = [p.cloud.p_clear for p in pred.passes + pred.marginal if p.cloud.tier == 2]
    assert t2, "need tier-2 passes"
    # soft membership keeps probabilities off the hard 0/1 rails for mixed members
    assert all(0.0 <= v <= 1.0 for v in t2)
    assert any(0.05 < v < 0.95 for v in t2)


def test_tier1_uncertainty_band():
    now = P.datetime(2026, 7, 14, tzinfo=P.timezone.utc)
    pred = _cloud_pred()
    P.attach_cloud(pred, forecast_json=FORECAST_JSON, ensemble_json=ENSEMBLE_JSON, now=now)
    t1 = [p.cloud for p in pred.passes + pred.marginal if p.cloud.tier == 1]
    assert t1
    for c in t1:
        assert c.total_band is not None
        lo, hi = c.total_band
        assert lo <= c.total <= hi and 0 <= lo and hi <= 100


# ---------------------------------------------------------------- B2 best-window
def test_campaign_summary_fields():
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=20.0,
                    marginal_off_nadir_deg=20.0)
    s = pred.summary
    assert s["total_opportunities"] == len(pred.passes) + len(pred.marginal)
    assert s["good_quality"] == sum(p.quality == "good"
                                    for p in pred.passes + pred.marginal)
    if s["good_quality"] >= 2:
        assert s["median_revisit_h"] > 0 and s["max_gap_h"] >= s["median_revisit_h"]
    if s.get("best_window"):
        bw = s["best_window"]
        assert bw["good_passes"] >= 1
        assert bw["start_utc"] <= bw["end_utc"]
        assert bw["mean_off_nadir_deg"] >= 0


def test_summary_in_json():
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=20.0)
    body = P.prediction_json([pred])
    assert "summary" in body and "total_opportunities" in body["summary"]


# ---------------------------------------------------------------- B1 swath coverage
def test_ellipsoid_intersect_nadir():
    import numpy as np
    # straight down from 550 km over the equator hits ~Earth radius
    sat = P.geodetic_to_ecef(0.0, 0.0, 550.0)
    hit = P.ellipsoid_intersect(sat, np.array([-1.0, 0.0, 0.0]))
    assert hit is not None
    lat, lon = P.ecef_to_geodetic_latlon(hit)
    assert abs(lat) < 1e-6 and abs(lon) < 1e-6
    assert abs(np.linalg.norm(hit) - P._A) < 1.0
    # a ray pointing away from Earth misses
    assert P.ellipsoid_intersect(sat, np.array([1.0, 0.0, 0.0])) is None


def test_coverage_full_for_small_aoi():
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=20.0,
                    marginal_off_nadir_deg=20.0)
    assert pred.passes
    for p in pred.passes:
        assert p.coverage_pct is not None and 0.0 <= p.coverage_pct <= 1.0
        assert len(p.footprint_lonlat) == 4
    # a 10×10 km AOI is fully inside a 20 km swath at these low rolls
    assert all(p.coverage_pct >= 0.99 for p in pred.passes)


def test_coverage_partial_for_oversized_aoi():
    # a very wide AOI (~1.2° ≈ 130 km E–W) cannot fit in a 20 km swath
    big = make_kmz(SITEA[0], SITEA[1], half_deg=0.6, name="BIG")
    pred = P.predict(big, days=14.0, start_utc=START, terrain_alt_m=400.0,
                     min_sun_elev_deg=-90, marginal_sun_elev_deg=-90, tles=_tles(),
                     max_off_nadir_deg=25.0, marginal_off_nadir_deg=25.0)
    rows = pred.passes + pred.marginal
    assert rows and any(p.coverage_pct is not None and p.coverage_pct < 0.6
                        for p in rows)


# ------------------------------------------- B1: push-broom swath sits on the track
def _equatorial_state(alt_km: float):
    """A satellite over (0,0) at `alt_km`, moving north. theta=0 makes the TEME
    axes coincide with ECEF, so the geometry is checkable by hand: the
    sub-satellite point is (0,0) and cross-track is (near enough) east-west."""
    import numpy as np
    r_t = P.geodetic_to_ecef(0.0, 0.0, alt_km)      # on the +X axis
    v_t = np.array([0.0, 0.0, 7.5])                 # +Z is north at the equator
    return r_t, v_t, 0.0


def _track_frame(r_t, v_t, theta):
    """(geodetic sub-satellite point, track unit, cross unit, geodetic alt km)."""
    import numpy as np
    sat = P.teme_to_ecef(r_t, theta)
    lat0, lon0 = P.ecef_to_geodetic_latlon(sat)       # geodetic sub-satellite point
    ssp = P.geodetic_to_ecef(lat0, lon0, 0.0)
    up = P.geodetic_up(lat0, lon0)
    v_ecef = P.teme_to_ecef(v_t, theta) - np.cross([0.0, 0.0, P._OMEGA_E], sat)
    track = v_ecef - (v_ecef @ up) * up; track /= np.linalg.norm(track)
    cross = np.cross(up, track); cross /= np.linalg.norm(cross)
    return ssp, track, cross, float(np.linalg.norm(sat - ssp))


def _aoi_at_cross_track_offset(r_t, v_t, theta, offset_km, box_km=10.0,
                               along_km=0.0):
    """A `box_km` square AOI whose centre sits `offset_km` cross-track (and
    `along_km` along-track) of the ground track, in the satellite's own frame."""
    ssp, track, cross, _ = _track_frame(r_t, v_t, theta)
    c = ssp + offset_km * cross + along_km * track
    clat, clon = P.ecef_to_geodetic_latlon(c)
    h = box_km / 2.0
    verts = []
    for da, dc in [(-h, -h), (-h, h), (h, h), (h, -h)]:
        la, lo = P.ecef_to_geodetic_latlon(c + da * track + dc * cross)
        verts.append((lo, la))
    return P.AOI("probe", verts, clon, clat, 0.0)


@pytest.mark.parametrize("alt_km", [
    705.0,     # nominal
    731.7,     # Landsat-9's real apogee-ish geodetic altitude; the half-swath is
])              # 96.0 km here, not 92.5 — a hardcoded swath is 3.5 km wrong
@pytest.mark.parametrize("delta_km,expected", [
    (-40.0, 1.00),    # AOI well inside the swath  -> fully imaged
    (-5.0, 1.00),     # AOI's outer edge just touches the swath edge
    (0.0, 0.50),      # AOI centred exactly ON the swath edge -> half imaged
    (2.5, 0.25),      # AOI mostly outside the swath
])
def test_pushbroom_coverage_follows_the_ground_track(delta_km, expected, alt_km):
    """A fixed nadir push-broom images what falls under its TRACK, so an AOI at the
    edge of the swath is only partly imaged. The agile (AOI-centred) model reports
    1.0 for every case here — a plausible number that cannot be false.

    Offsets are expressed relative to the *actual* swath edge rather than a
    hardcoded 92.5 km, because the ground half-swath is not a constant: it follows
    the geodetic altitude. Pinning against 92.5 would only be true at the one
    altitude where the model's dominant error vanishes. [SESSION 2026-07-16]

    Analytic truth: a 10 km AOI centred `delta` beyond the swath edge E spans
    [E+delta-5, E+delta+5], of which [.., E] is imaged => (5-delta)/10, capped at 1.
    """
    fov = P.LANDSAT.max_off_nadir_deg
    r_t, v_t, theta = _equatorial_state(alt_km)
    _, _, _, alt = _track_frame(r_t, v_t, theta)
    edge = P.ground_half_swath_km(alt, fov)
    aoi = _aoi_at_cross_track_offset(r_t, v_t, theta, edge + delta_km)
    fp = P.swath_footprint_lonlat(r_t, v_t, theta, aoi, 185.0, agile=False,
                                  fov_half_deg=fov)
    assert P.aoi_coverage_fraction(fp, aoi) == pytest.approx(expected, abs=0.02)


def test_pushbroom_half_swath_tracks_geodetic_altitude():
    """The ground swath of a fixed-FOV push-broom is NOT a constant: Landsat-9's
    geodetic altitude swings ~704.6-731.7 km over one orbit (circular
    geocentrically, but the ellipsoid is not), moving the half-swath 92.5->96.0 km.
    Hardcoding swath_km/2 against an angular gate let the two disagree by 3.5 km
    and report an AOI wholly inside the FOV as 0% covered. [VERIFIED 2026-07-16]"""
    fov = P.LANDSAT.max_off_nadir_deg
    # inverse of the derivation the profile itself uses
    assert P.ground_half_swath_km(705.0, fov) == pytest.approx(92.5, abs=0.1)
    lo = P.ground_half_swath_km(704.6, fov)
    hi = P.ground_half_swath_km(731.7, fov)
    assert hi - lo > 3.0, "altitude swing must move the swath materially"
    assert lo < hi                                  # monotonic in altitude
    # a look direction past the limb has no ground intersection
    assert P.ground_half_swath_km(705.0, 89.0) is None


@pytest.mark.parametrize("prof_key,lat,lon", [
    ("landsat", -55.0, 0.0),
    ("sentinel2", -70.0, 90.0),
])
def test_pushbroom_coverage_never_contradicts_the_access_gate(prof_key, lat, lon):
    """The regression this whole model change risks: `predict()` must never report
    an AOI lying wholly inside the sensor's FOV as less than fully covered.

    These sites are chosen, not swept. The defect is driven by the geodetic-altitude
    swing, so it only shows where a pass puts the AOI *near the FOV edge* at a
    *high* point of the orbit — high southern latitudes here. A tidy sweep of
    round-number latitudes at one longitude never lands on such a geometry and
    passes happily against the bug. [VERIFIED 2026-07-16]

    The <=2% tolerance is the gate's own geocentric-nadir convention (the validated
    default) meeting the swath's geodetic ground track: bounded at ~0.62 km
    cross-track, and it vanishes entirely under nadir_ellipsoid=True."""
    prof = P.get_profile(prof_key)
    pred = P.predict(make_kmz(lat, lon), days=6.0, start_utc=START,
                     polygon_name="TEST_AOI",
                     offline_tle_file=str(FIX / "tles_landsat_sentinel2_20260715.txt"),
                     profile=prof, min_sun_elev_deg=-90, marginal_sun_elev_deg=-90)
    rows = [p for p in pred.passes + pred.marginal + pred.nonoperational
            if p.coverage_pct is not None]
    near_edge = 0
    for p in rows:
        if p.max_off_nadir_aoi_deg > 0.9 * prof.max_off_nadir_deg:
            near_edge += 1
        if p.max_off_nadir_aoi_deg <= prof.max_off_nadir_deg:
            assert p.coverage_pct >= 0.98, (
                f"{prof_key} @ ({lat},{lon}): every AOI vertex is inside the "
                f"{prof.max_off_nadir_deg} deg FOV (max {p.max_off_nadir_aoi_deg}) "
                f"yet coverage reads {p.coverage_pct}")
    # Guard the test's own premise: no near-edge pass => nothing here can bite.
    assert near_edge, (f"{prof_key} @ ({lat},{lon}) produced no pass near the FOV "
                       "edge; this test is asleep and must be retargeted")


def test_pushbroom_swath_centres_on_the_geodetic_ground_track():
    """The swath centreline must sit on the *geodetic* sub-satellite point — the
    standard ground track, and where a local-vertical-pointing push-broom's
    boresight lands.

    Uses the real Landsat-9 element set rather than a synthetic state, deliberately:
    the geodetic and geocentric sub-points coincide exactly at the equator (so the
    parametrized test above cannot see this at all), and for a *due-north* track
    their ~2 km separation is almost entirely ALONG-track, which a continuous strip
    does not care about. It is Landsat's 98.2 deg retrograde inclination that turns
    part of it cross-track — the only component that moves the swath edge.
    [SESSION 2026-07-16]"""
    import numpy as np
    from sgp4.api import Satrec
    tle = P._parse_3le_file((FIX / "tles_landsat_sentinel2_20260715.txt").read_text(),
                            {"LANDSAT9": 49260})["LANDSAT9"]
    sr = Satrec.twoline2rv(tle.line1, tle.line2)
    jd, fr = P.dt_to_jd(tle.epoch_utc)
    # +74 min: where the cross-track separation of the two candidate sub-points
    # peaks over this orbit (0.623 km, near the southern turn). Chosen for maximum
    # discrimination — the model is what is under test here, not the AU sites.
    fr += 74.0 / 1440.0
    err, r, v = sr.sgp4(jd, fr)
    assert err == 0
    r_t, v_t = np.array(r), np.array(v)
    theta = float(P.gmst_rad(jd + fr))
    sat = P.teme_to_ecef(r_t, theta)

    ssp_lat, ssp_lon = P.ecef_to_geodetic_latlon(sat)
    geodetic_pt = P.geodetic_to_ecef(ssp_lat, ssp_lon, 0.0)
    geocentric_pt = P.ellipsoid_intersect(sat, -sat / np.linalg.norm(sat))

    up = P.geodetic_up(ssp_lat, ssp_lon)
    v_ecef = P.teme_to_ecef(v_t, theta) - np.cross([0.0, 0.0, P._OMEGA_E], sat)
    track = v_ecef - (v_ecef @ up) * up; track /= np.linalg.norm(track)
    cross = np.cross(up, track); cross /= np.linalg.norm(cross)

    # Guard the test's own premise: if the two candidates ever stop being
    # distinguishable cross-track here, this test is asleep and must be retuned.
    separation = abs(float((geodetic_pt - geocentric_pt) @ cross))
    assert separation > 0.3, (
        f"geodetic and geocentric sub-points differ by only {separation:.3f} km "
        "cross-track here — this test no longer discriminates; retune the epoch")

    # Put the AOI well off the track, so the swath centre is decided by the model
    # rather than by where the AOI happens to sit.
    aoi = _aoi_at_cross_track_offset(r_t, v_t, theta, 60.0)
    fp = P.swath_footprint_lonlat(r_t, v_t, theta, aoi, 185.0, agile=False)

    # Coverage cannot pin this — a sub-km shift leaves a 185 km swath reading 1.0 —
    # so assert on the centreline's offset directly. Reconstructing the centre from
    # the four surface-projected corners costs ~0.02 km of chord/curvature error,
    # well inside the ~0.5 km signal.
    corners = np.array([P.geodetic_to_ecef(la, lo, 0.0) for lo, la in fp])
    centreline = corners.mean(axis=0)
    off_geodetic = abs(float((centreline - geodetic_pt) @ cross))
    off_geocentric = abs(float((centreline - geocentric_pt) @ cross))

    assert off_geodetic < 0.1, f"centreline sits {off_geodetic:.3f} km off the track"
    assert off_geocentric > separation / 2, (
        "centreline is on the geocentric sub-point, not the geodetic ground track")


def test_pushbroom_keeps_the_aois_along_track_station():
    """Only the CROSS-track centre moves onto the track — the footprint must still
    straddle the AOI along-track. Pins the docstring's load-bearing claim, which
    `center = ssp` (deleting the behaviour outright) would otherwise satisfy: the
    AOI sits ~1.6-2.0 km along-track of the sub-satellite point at TCA, so this is
    a live invariant, not dead code. [SESSION 2026-07-16]"""
    import numpy as np
    r_t, v_t, theta = _equatorial_state(705.0)
    ssp, track, cross, _ = _track_frame(r_t, v_t, theta)
    ALONG = 30.0
    aoi = _aoi_at_cross_track_offset(r_t, v_t, theta, 40.0, along_km=ALONG)
    fp = P.swath_footprint_lonlat(r_t, v_t, theta, aoi, 185.0, agile=False,
                                  fov_half_deg=P.LANDSAT.max_off_nadir_deg)
    corners = np.array([P.geodetic_to_ecef(la, lo, 0.0) for lo, la in fp])
    centre = corners.mean(axis=0)
    assert float((centre - ssp) @ track) == pytest.approx(ALONG, abs=0.5)
    # and the AOI is still fully imaged, being well inside the swath
    assert P.aoi_coverage_fraction(fp, aoi) == pytest.approx(1.0, abs=1e-3)


def test_agile_coverage_stays_centred_on_the_aoi():
    """The agile model is unchanged and still deliberately AOI-centred: a rolling
    sensor points at the target, so coverage asks 'does my AOI fit cross-track'.
    This is the behaviour DEVELOPMENT.md records as mislabelled-not-broken; it is a
    human's call to rename or re-model, so it is pinned here, not changed."""
    r_t, v_t, theta = _equatorial_state(525.0)
    aoi = _aoi_at_cross_track_offset(r_t, v_t, theta, 95.0)
    fp = P.swath_footprint_lonlat(r_t, v_t, theta, aoi, 185.0, agile=True)
    assert P.aoi_coverage_fraction(fp, aoi) == pytest.approx(1.0, abs=1e-3)


def test_swath_footprint_defaults_to_agile():
    """Dragonette is the default sensor; the default must not silently change."""
    r_t, v_t, theta = _equatorial_state(525.0)
    aoi = _aoi_at_cross_track_offset(r_t, v_t, theta, 95.0)
    assert (P.swath_footprint_lonlat(r_t, v_t, theta, aoi, 185.0)
            == P.swath_footprint_lonlat(r_t, v_t, theta, aoi, 185.0, agile=True))


def test_predict_uses_the_profile_agility_for_coverage():
    """predict() must thread the profile's agility through, not hardcode agile."""
    import numpy as np
    seen = {}
    real = P.swath_footprint_lonlat

    def spy(r_t, v_t, theta, aoi, swath_km=P.SWATH_KM, agile=True, **kw):
        seen[swath_km] = agile
        return real(r_t, v_t, theta, aoi, swath_km, agile, **kw)

    P.swath_footprint_lonlat = spy
    try:
        for prof in (P.DRAGONETTE, P.LANDSAT):
            tles = P._parse_3le_file(Path(DEMO_TLES).read_text(), P.SATELLITES)
            # reuse the synthetic constellation under each profile's optics
            P.predict(SITEA_KMZ, days=3.0, start_utc=START, terrain_alt_m=400.0,
                      polygon_name="SITEA_100sqkm", tles=tles,
                      profile=prof, min_sun_elev_deg=-90,
                      marginal_sun_elev_deg=-90,
                      max_off_nadir_deg=25.0, marginal_off_nadir_deg=25.0)
    finally:
        P.swath_footprint_lonlat = real
    assert seen[P.DRAGONETTE.swath_km] is True
    assert seen[P.LANDSAT.swath_km] is False


def test_coverage_in_json_and_xlsx(tmp_path):
    from openpyxl import load_workbook
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=20.0)
    body = P.prediction_json([pred])
    assert "coverage_pct" in body["passes"][0]
    assert "footprint_lonlat" in body["passes"][0]
    f = tmp_path / "cov.xlsx"; f.write_bytes(P.write_xlsx(pred))
    assert "AOI Coverage %" in [c.value for c in load_workbook(f)["Passes"][1]]


# ---------------------------------------------------------------- R6
import json  # noqa: E402

FORECAST_JSON = json.loads((FIX / "openmeteo_forecast_sample.json").read_text())
ENSEMBLE_JSON = json.loads((FIX / "openmeteo_ensemble_sample.json").read_text())


def _cloud_pred():
    # Wide filters so passes span the whole 14 d window (all three tiers).
    return _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=60.0,
                    marginal_off_nadir_deg=60.0)


# The committed sample fixtures were captured when fetch_real_data.py still
# requested forecast_days=11, so they end 11 d after the window start even though
# TIER2_MAX_DAYS is 15. The live API does return 16 d (verified 2026-07-15), so
# this is a fixture limitation, not a provider cap: leads beyond _FIXTURE_DAYS
# legitimately fall off the end of the series and must read n/a rather than
# inherit the last available hour. [SESSION 2026-07-15]
_FIXTURE_DAYS = 11.0


def test_cloud_tiers_assigned_by_lead_time():
    pred = _cloud_pred()
    P.attach_cloud(pred, forecast_json=FORECAST_JSON, ensemble_json=ENSEMBLE_JSON,
                   now=P.datetime(2026, 7, 14, tzinfo=P.timezone.utc))
    ref = pred.start_utc
    seen_t2 = 0
    for p in pred.passes + pred.marginal + pred.nonoperational:
        lead = (p.tca_utc - ref).total_seconds() / 86400.0
        assert p.cloud is not None
        if lead < 5:
            assert p.cloud.tier == 1 and p.cloud.label == "forecast"
            assert p.cloud.total is not None
            assert p.cloud.p_clear is None          # never a bare % dressed as prob
        elif lead < _FIXTURE_DAYS:                  # A5: Tier 2 now to 15 days
            assert p.cloud.tier == 2
            assert p.cloud.label == "outlook (probabilistic)"
            assert 0.0 <= p.cloud.p_clear <= 1.0
            assert p.cloud.total is None            # never a bare deterministic %
            seen_t2 += 1
        elif lead < 15:
            # Inside Tier 2's lead range but past the fixture's series.
            assert p.cloud.tier == 0 and p.cloud.label == "n/a"
        else:
            assert p.cloud.tier == 3 and p.cloud.total is None and p.cloud.p_clear is None
    assert seen_t2, "fixture should still exercise at least one real Tier-2 pass"


def test_cloud_beyond_returned_series_is_na_not_carried_over():
    """A pass past the end of the series a provider actually returned must read
    n/a, with a warning — never a value snapped from the last available hour.

    [SESSION 2026-07-15] `_nearest_hour_index` was an unbounded nearest-neighbour,
    so with an 11 d series a day-14 pass silently inherited the last hour of day
    11 — a 73 h stale sample published as that pass's outlook.
    """
    pred = _cloud_pred()
    ref = P.datetime(2026, 7, 14, tzinfo=P.timezone.utc)
    P.attach_cloud(pred, forecast_json=FORECAST_JSON, ensemble_json=ENSEMBLE_JSON,
                   now=ref)
    ens_end = P._parse_hourly_times(ENSEMBLE_JSON["hourly"])[-1]
    beyond = [p for p in pred.passes + pred.marginal + pred.nonoperational
              if p.tca_utc > ens_end + P.timedelta(hours=2)
              and (p.tca_utc - ref).total_seconds() / 86400.0 < 15.0]
    assert beyond, "fixture window should contain a Tier-2 pass past the series end"
    for p in beyond:
        assert p.cloud.tier == 0 and p.cloud.label == "n/a"
        assert p.cloud.total is None and p.cloud.p_clear is None
    assert any("beyond the end of the cloud series" in w for w in pred.warnings)


def test_nearest_hour_index_rejects_out_of_range_tca():
    times = P._parse_hourly_times(ENSEMBLE_JSON["hourly"])
    assert P._nearest_hour_index(times, times[10]) == 10          # exact hit
    assert P._nearest_hour_index(times, times[-1]) == len(times) - 1
    # 3 days past the end of the series -> no match, not the last index
    assert P._nearest_hour_index(times, times[-1] + P.timedelta(days=3)) is None
    assert P._nearest_hour_index([], times[0]) is None


def test_cloud_single_call_per_api_batched():
    pred = _cloud_pred()
    calls = []

    def fake_get(url):
        calls.append(url)
        return json.dumps(FORECAST_JSON if "v1/forecast" in url else ENSEMBLE_JSON)

    P.attach_cloud(pred, http_get=fake_get,
                   now=P.datetime(2026, 7, 14, tzinfo=P.timezone.utc))
    assert sum("v1/forecast" in u for u in calls) == 1      # one forecast call
    assert sum("ensemble" in u for u in calls) == 1         # one ensemble call


def test_cloud_graceful_offline_na():
    pred = _cloud_pred()

    def dead(url):
        raise ConnectionError("no route")

    P.attach_cloud(pred, http_get=dead,
                   now=P.datetime(2026, 7, 14, tzinfo=P.timezone.utc))
    # never blocks; tier-1/2 passes fall back to n/a, tier-3 still climatology
    assert any(p.cloud.label == "n/a" for p in pred.passes + pred.marginal)
    assert any("unavailable" in w for w in pred.warnings)


def test_cloud_threshold_affects_p_clear():
    p10 = _cloud_pred(); p90 = _cloud_pred()
    now = P.datetime(2026, 7, 14, tzinfo=P.timezone.utc)
    P.attach_cloud(p10, ensemble_json=ENSEMBLE_JSON, forecast_json=FORECAST_JSON,
                   threshold=10.0, now=now)
    P.attach_cloud(p90, ensemble_json=ENSEMBLE_JSON, forecast_json=FORECAST_JSON,
                   threshold=90.0, now=now)
    t2a = [p.cloud.p_clear for p in p10.marginal + p10.passes if p.cloud.tier == 2]
    t2b = [p.cloud.p_clear for p in p90.marginal + p90.passes if p.cloud.tier == 2]
    assert t2a and t2b
    # a looser "clear" threshold can only raise (never lower) P(clear)
    assert sum(t2b) >= sum(t2a)


def test_cloud_columns_in_xlsx(tmp_path):
    from openpyxl import load_workbook
    pred = _cloud_pred()
    P.attach_cloud(pred, forecast_json=FORECAST_JSON, ensemble_json=ENSEMBLE_JSON,
                   now=P.datetime(2026, 7, 14, tzinfo=P.timezone.utc))
    blob = P.write_xlsx(pred, tz_name="Australia/Brisbane")
    f = tmp_path / "c.xlsx"; f.write_bytes(blob)
    ws = load_workbook(f)["Passes"]
    hdr = [c.value for c in ws[1]]
    for col in P.CLOUD_COLUMNS:
        assert col in hdr
    method = " ".join(str(c.value) for row in load_workbook(f)["Method"].iter_rows()
                      for c in row if c.value)
    assert "Open-Meteo" in method            # CC BY attribution


def test_cloud_json_shape():
    pred = _cloud_pred()
    P.attach_cloud(pred, forecast_json=FORECAST_JSON, ensemble_json=ENSEMBLE_JSON,
                   now=P.datetime(2026, 7, 14, tzinfo=P.timezone.utc))
    body = P.prediction_json([pred])
    sample = (body["passes"] + body["marginal"])[0]
    assert "cloud" in sample and "tier" in sample["cloud"]


# ---------------------------------------------------------------- R8
def test_campaign_timeline_spans_all_aois():
    preds = [_predict(SITEC_KMZ), _predict(SITEB_KMZ, "100sqkm")]
    n = sum(len(p.passes) + len(p.marginal) + len(p.nonoperational) for p in preds)
    fig, ax = P.build_timeline_figure(preds, tz_name="Australia/Brisbane")
    assert len(ax.patches) == n            # bars for every AOI's passes on one axis
    import matplotlib.pyplot as plt
    plt.close(fig)


def test_multi_aoi_workbook_has_aoi_column(tmp_path):
    from openpyxl import load_workbook
    preds = [_predict(SITEC_KMZ), _predict(SITEB_KMZ, "100sqkm")]
    blob = P.write_xlsx_multi(preds, tz_name="Australia/Brisbane")
    f = tmp_path / "camp.xlsx"; f.write_bytes(blob)
    wb = load_workbook(f)
    ws = wb["Passes"]
    assert ws.cell(1, 1).value == "AOI"
    aois = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value}
    assert any(a and a.startswith("Site C") for a in aois)


def test_cli_multi_kmz_campaign(tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / "campaign.xlsx"
    r = subprocess.run(
        [sys.executable, str(SRC / "cli.py"),
         str(FIX / "SiteC.kmz"),
         str(FIX / "SiteB.kmz"), "--all-polygons",
         "--alt", "50", "--min-sun", "-90", "--tle-file", DEMO_TLES,
         "--start", "2026-07-14T00:00:00", "-o", str(out)],
        capture_output=True, text=True)
    assert r.returncode == 0, r.stderr
    wb = load_workbook(out)
    assert wb["Passes"].cell(1, 1).value == "AOI"
    assert "Timeline" in wb.sheetnames


def test_cli_bad_polygon_name_clean_exit(tmp_path):
    r = subprocess.run(
        [sys.executable, str(SRC / "cli.py"),
         str(FIX / "SiteC.kmz"),
         "--polygon", "does-not-exist", "--tle-file", DEMO_TLES,
         "--start", "2026-07-14T00:00:00", "-o", str(tmp_path / "x.xlsx")],
        capture_output=True, text=True)
    assert r.returncode == 2
    assert "Traceback" not in r.stderr and "No polygon named" in r.stderr


def test_api_multi_aoi_json(monkeypatch):
    from fastapi.testclient import TestClient
    import app as A
    monkeypatch.setattr(P, "fetch_tles", lambda *a, **k: (_tles(), []))
    client = TestClient(A.app)
    # all_polygons across two files: Site C(1) + Site B(2) = 3 AOIs
    r = client.post("/predict/json",
                    files=[("kmz", ("m.kmz", SITEC_KMZ)),
                           ("kmz", ("g.kmz", SITEB_KMZ))],
                    data={"days": "14", "alt": "20", "min_sun": "-90",
                          "all_polygons": "true"})
    assert r.status_code == 200
    body = r.json()
    assert body["schema_version"] == "2.1" and len(body["aois"]) == 3


# ---------------------------------------------------------------- R9 (DT contract)
# Freeze the JSON shape: the digital-twin front-end depends on these keys.
# Adding keys is a minor bump; removing/renaming one is a breaking change.
PASS_KEYS = {"satellite", "tca_utc", "off_nadir_deg", "sun_elev_deg",
             "max_off_nadir_aoi_deg", "slant_range_km", "tle_epoch_utc",
             "category", "operational", "node", "local_solar_time_h",
             "quality", "timing_sigma_s", "coverage_pct", "footprint_lonlat",
             "geometry"}
SINGLE_KEYS = {"schema_version", "sensor", "aoi", "window_utc", "passes", "marginal",
               "nonoperational", "cloud_daily", "cloud_horizon_utc", "summary",
               "warnings", "params"}
AOI_KEYS = {"name", "centroid_lat", "centroid_lon", "terrain_alt_m", "vertices_lonlat"}
CLOUD_KEYS = {"tier", "label", "total_pct", "low_pct", "mid_pct", "high_pct",
              "p_clear", "threshold_pct", "spread_pct", "clim_clear_pct",
              "total_band_pct", "optical_obstruction_pct", "cirrus_flag",
              "likely_cloudy"}


def test_json_schema_version_and_single_shape():
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=45.0)
    body = P.prediction_json([pred])
    assert body["schema_version"] == "2.1"
    assert set(body) == SINGLE_KEYS
    assert set(body["aoi"]) == AOI_KEYS
    assert body["passes"], "need a pass to check row shape"
    assert set(body["passes"][0]) == PASS_KEYS


def test_json_multi_shape():
    preds = [_predict(SITEC_KMZ), _predict(SITEB_KMZ, "100sqkm")]
    body = P.prediction_json(preds)
    assert set(body) == {"schema_version", "aois"}
    assert all(set(a) == SINGLE_KEYS - {"schema_version"} for a in body["aois"])


# ---------------------------------------------------- combined "all sensors" view
def _empty_pred(sensor):
    aoi = P.parse_kmz(SITEC_KMZ)
    return P.Prediction(aoi=aoi, start_utc=START, end_utc=START + P.timedelta(days=14),
                        params={"sensor": sensor, "max_off_nadir_deg": 20,
                                "min_sun_elev_deg": 20})


def test_combined_roster_unions_every_profile():
    sats, op = P.combined_roster()
    assert set(sats) == {"DRAG01", "DRAG02", "DRAG03", "DRAG04", "DRAG05",
                         "LANDSAT8", "LANDSAT9", "SENTINEL2A", "SENTINEL2B", "SENTINEL2C"}
    # non-op flag survives the union — DRAG05 must stay non-taskable
    assert op["DRAG05"] is False
    assert op["LANDSAT8"] is True and op["LANDSAT9"] is True and op["SENTINEL2A"] is True


def test_merge_predictions_marks_all_and_exposes_union_roster():
    parts = [_empty_pred("dragonette"), _empty_pred("landsat"), _empty_pred("sentinel2")]
    merged = P.merge_predictions(parts)
    assert merged.params["sensor"] == "all"
    body = P.prediction_json([merged])
    assert body["sensor"]["key"] == "all"
    assert set(body["sensor"]["satellites"]) == set(P.combined_roster()[0])
    assert body["sensor"]["operational"]["DRAG05"] is False
    assert body["sensor"]["operational"]["LANDSAT8"] is True


def test_combined_timeline_figure_labels_every_constellation():
    merged = P.merge_predictions([_empty_pred(k) for k in
                                  ("dragonette", "landsat", "sentinel2")])
    fig, ax = P.build_timeline_figure(merged)
    labels = [t.get_text() for t in ax.get_yticklabels()]
    assert "LANDSAT8" in labels and "LANDSAT9" in labels    # both operational, no suffix
    assert "DRAG05 (non-op)" in labels                       # the surviving non-op lane


def test_combined_method_sheet_describes_all_sensors_not_one():
    merged = P.merge_predictions([_empty_pred(k) for k in
                                  ("dragonette", "landsat", "sentinel2")])
    rows = dict(P._method_rows(merged))
    assert "LANDSAT9=49260" in rows["Satellites (NORAD)"]
    assert "SENTINEL2A=40697" in rows["Satellites (NORAD)"]
    assert "DRAG01=56225" in rows["Satellites (NORAD)"]


def test_json_cloud_block_shape():
    pred = _predict(SITEA_KMZ, "SITEA_100sqkm", max_off_nadir_deg=60.0,
                    marginal_off_nadir_deg=60.0)
    P.attach_cloud(pred, forecast_json=FORECAST_JSON, ensemble_json=ENSEMBLE_JSON,
                   now=P.datetime(2026, 7, 14, tzinfo=P.timezone.utc))
    body = P.prediction_json([pred])
    row = next(r for r in body["passes"] + body["marginal"] if "cloud" in r)
    assert set(row["cloud"]) == CLOUD_KEYS


# ---------------------------------------------------------------- R7 API
def test_api_ambiguous_422(monkeypatch):
    from fastapi.testclient import TestClient
    import app as A
    monkeypatch.setattr(P, "fetch_tles", lambda *a, **k: (_tles(), []))
    client = TestClient(A.app)
    r = client.post("/predict/json", files={"kmz": ("SiteA.kmz", SITEA_KMZ)},
                    data={"days": "14", "alt": "400", "min_sun": "-90"})
    assert r.status_code == 422
    detail = r.json()["detail"]
    assert "SITEA_100sqkm" in detail["polygons"]


def test_api_all_polygons_json(monkeypatch):
    from fastapi.testclient import TestClient
    import app as A
    monkeypatch.setattr(P, "fetch_tles", lambda *a, **k: (_tles(), []))
    client = TestClient(A.app)
    r = client.post("/predict/json", files={"kmz": ("SiteA.kmz", SITEA_KMZ)},
                    data={"days": "14", "alt": "400", "min_sun": "-90",
                          "all_polygons": "true"})
    assert r.status_code == 200
    body = r.json()
    assert body["schema_version"] == "2.1"
    assert len(body["aois"]) == 3
